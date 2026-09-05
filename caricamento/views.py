"""
API Views per il sistema di ottimizzazione carico tridimensionale.

Fornisce endpoint REST per:
- CRUD di Contenitori, Oggetti, PianiDiCarico
- Endpoint dettagliato per il frontend Three.js
- Avvio ottimizzazione (sincrona / asincrona)
"""

import hashlib
import io
import json
import logging
import os
import random
import re
import math
import secrets
import uuid
from datetime import timedelta
from decimal import Decimal
from urllib.parse import unquote

from PIL import Image

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.signing import BadSignature, Signer
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme

logger = logging.getLogger(__name__)
from django.db import models, transaction
from django.shortcuts import redirect, render
from rest_framework import status, viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.throttling import UserRateThrottle


class OttimizzaRateThrottle(UserRateThrottle):
    """Throttle specifico per l'endpoint /ottimizza/ (10 richieste/minuto)."""
    scope = "ottimizza"


class PanelRateThrottle(UserRateThrottle):
    """Throttle generoso per le operazioni pannello (DELETE + POST righe).

    Il flusso "Elabora" esegue 1 DELETE + N POST sequenziali.  Con un
    limite di 60/min il limite viene superato rapidamente con piano di
    50+ righe.  Questo throttle alza il limite a 200/min.
    """
    scope = "panel"

from .models import (
    Contenitore,
    ImpostazioniSistema,
    Oggetto,
    OggettoDaCaricare,
    OggettoPosizionato,
    PianoDiCarico,
    SezioneCarico,
    UserProfile,
    StatoPiano,
    VincoloOggetto,
    VincoloTraOggetti,
    TipoRelazione,
)
from .serializers import (
    AvviaOttimizzazioneSerializer,
    ContenitoreCreateSerializer,
    ContenitoreDetailSerializer,
    ContenitoreListSerializer,
    OggettoCreateSerializer,
    OggettoDaCaricareSerializer,
    OggettoDaCaricareUpdateSerializer,
    OggettoDetailSerializer,
    OggettoListSerializer,
    PianoDiCaricoCreateSerializer,
    PianoDiCaricoDetailSerializer,
    PianoDiCaricoListSerializer,
    SezioneCaricoSerializer,
    SezioneCaricoWriteSerializer,
    ImpostazioniOttimizzatoreSerializer,
    VincoloOggettoUpdateSerializer,
    VincoloTraOggettiSerializer,
)
from .tasks import accoda_ottimizzazione, esegui_ottimizzazione_sincrona
from .engine.orchestratore_tre_d import stima_ops_piano
from .engine.tre_d.constants import (
    OPTIMIZATION_TIME_BUDGET_ASYNC_SECONDS,
    OPTIMIZATION_TIME_BUDGET_SECONDS,
)


# ===========================================================================
# Helper: Calcola distribuzione pesi per sezioni
# ===========================================================================

def _calcola_distribuzione_pesi(piano):
    """Calcola la distribuzione del carico sulle sezioni per un piano.

    Replica la logica di _calcola_distribuzione_sezioni() di bin_packing.py
    ma opera su un piano salvato, senza dover ri-eseguire l'ottimizzazione.
    Restituisce anche i dati degli oggetti posizionati per il profilo di peso.
    """
    from .engine.sezione_weight_tracker import SezioneWeightTracker

    sezioni = list(piano.contenitore.sezioni.all())
    oggetti_posizionati = list(
        piano.oggetti_posizionati.select_related("oggetto").all()
    )

    oggetti = []
    for op in oggetti_posizionati:
        oggetti.append({
            "codice": op.oggetto.codice,
            "posizione_x_mm": op.coordinata_x_mm,
            "dimensione_x_mm": op.dimensione_x_mm,
            "peso_kg": float(op.oggetto.peso_kg),
        })

    if not sezioni:
        return {"sezioni": [], "oggetti": oggetti}

    if not oggetti_posizionati:
        return {"sezioni": [], "oggetti": oggetti}

    tracker = SezioneWeightTracker(sezioni)
    for op in oggetti_posizionati:
        x_start = op.coordinata_x_mm
        x_end = op.coordinata_x_mm + op.dimensione_x_mm
        peso = float(op.oggetto.peso_kg)
        tracker.applica(x_start, x_end, peso)

    # Riepilogo con nome sezione incluso
    risultati = []
    for s in sezioni:
        limite = float(s.carico_massimo_kg)
        carico = tracker.carico_attuale.get(s.id, 0.0)
        risultati.append({
            "sezione_id": s.id,
            "nome": s.nome,
            "inizio_x_mm": s.inizio_x_mm,
            "fine_x_mm": s.fine_x_mm,
            "carico_massimo_kg": limite,
            "carico_attuale_kg": round(carico, 2),
            "margine_kg": round(max(0.0, limite - carico), 2),
        })
    return {"sezioni": risultati, "oggetti": oggetti}


def _colore_posizionamento(op):
    """Colore da mostrare per un posizionamento.

    Preferisce il colore della riga ``OggettoDaCaricare`` (quando il
    posizionamento è collegato a una riga), perché il colore per-riga ha la
    precedenza su quello dell'anagrafica. In questo modo, modificare il colore
    di una riga dal pannello (o salvarla con colore automatico distinto) si
    riflette subito in scena 3D ed export, anche senza ri-elaborare.
    """
    riga = op.riga_origine
    if riga is not None and riga.colore:
        return riga.colore
    return op.colore or "#4488ff"


# ===========================================================================
# VIEW: Pagina Visualizzatore 3D
# ===========================================================================

# ---------------------------------------------------------------------------
# Helper: Controlli anti-abuso demo (3 segnali)
# ---------------------------------------------------------------------------

def _client_ip(request):
    """Restituisce l'IP reale del client dietro il reverse proxy nginx.

    La logica vive in caricamento.client_ip ed è condivisa con django-axes
    (AXES_CLIENT_IP_CALLABLE): entrambi i controlli usano lo stesso IP.
    """
    from .client_ip import get_client_ip
    return get_client_ip(request)


def _demo_cookie_token(request):
    """Restituisce o prepara il token persistente anti-abuso per la richiesta."""
    token = request.COOKIES.get("cb_demo", "") or getattr(request, "_cb_demo_token", "")
    if not token:
        token = secrets.token_urlsafe(32)
        request._cb_demo_token = token
    return token


def _attach_demo_cookie(request, response):
    """Imposta il cookie anti-abuso sulla risposta iniziale della landing."""
    if not request.COOKIES.get("cb_demo"):
        response.set_cookie(
            "cb_demo",
            _demo_cookie_token(request),
            max_age=365 * 24 * 60 * 60,
            httponly=True,
            secure=request.is_secure(),
            samesite="Lax",
        )
    return response


def _check_demo_abuse(request, user=None):
    """Verifica se i segnali sono già associati a una demo utilizzata.

    Il controllo viene usato quando si sta per assegnare un NUOVO trial,
    non durante il login di un account già esistente. I fingerprint di tutte
    le demo non paganti restano rilevanti anche quando il trial è ancora attivo:
    così lo stesso browser/rete non può aprire più trial contemporaneamente.
    """
    from .models import DemoFingerprint, ImpostazioniSistema

    imp = ImpostazioniSistema.get()
    if not imp.controlli_demo_attivi:
        return False
    soglia = imp.soglia_controlli_demo  # default 1

    ip_raw = _client_ip(request)
    ip_hash = hashlib.sha256(ip_raw.encode()).hexdigest() if ip_raw else ""
    browser_hash = request.COOKIES.get("cb_fp", "") or None
    cookie_token = _demo_cookie_token(request) or None

    # Una demo già utilizzata, attiva o scaduta, consuma il trial del segnale.
    matches = DemoFingerprint.objects.filter(
        user_profile__is_paying=False,
        user_profile__user__is_staff=False,
        user_profile__user__is_superuser=False,
    )

    match_count = 0
    if ip_hash and matches.filter(ip_hash=ip_hash).exists():
        match_count += 1
    if browser_hash and matches.filter(browser_hash=browser_hash).exists():
        match_count += 1
    if cookie_token and matches.filter(cookie_token=cookie_token).exists():
        match_count += 1

    blocked = match_count >= soglia
    if blocked and user is not None:
        logger.info(
            "Demo BLOCKED: user=%s match_count=%d soglia=%d ip=%s fp=%s",
            user.username, match_count, soglia,
            ip_hash[:12], (browser_hash or "")[:12],
        )
    return blocked


def _check_fingerprint_expired_trial(request, user):
    """Verifica se un utente con trial scaduto ha fingerprint che matchano
    con un'altra demo esistente (stesso device/rete).

    Questo previene il "cambio identità": l'utente con trial scaduto non
    può creare un nuovo account dallo stesso dispositivo/rete.
    Restituisce True se l'accesso deve essere bloccato.
    """
    from .models import DemoFingerprint, ImpostazioniSistema, UserProfile

    imp = ImpostazioniSistema.get()
    if not imp.controlli_demo_attivi:
        return False

    # Non bloccare staff o utenti paganti
    if user.is_staff or user.is_superuser:
        return False
    try:
        profile = user.profile
        if profile.is_paying:
            return False
    except UserProfile.DoesNotExist:
        return False

    soglia = imp.soglia_controlli_demo  # default 1

    ip_raw = _client_ip(request)
    ip_hash = hashlib.sha256(ip_raw.encode()).hexdigest() if ip_raw else ""
    browser_hash = request.COOKIES.get("cb_fp", "") or None
    cookie_token = _demo_cookie_token(request) or None

    # Cerca fingerprint di demo che non appartengono a questo utente
    # (altri utenti demo non paganti con trial scaduto)
    matches = DemoFingerprint.objects.filter(
        user_profile__is_paying=False,
        user_profile__user__is_staff=False,
        user_profile__user__is_superuser=False,
    ).exclude(user_profile__user=user)

    match_count = 0
    matched_signals = []
    if ip_hash and matches.filter(ip_hash=ip_hash).exists():
        match_count += 1
        matched_signals.append("ip")
    if browser_hash and matches.filter(browser_hash=browser_hash).exists():
        match_count += 1
        matched_signals.append("browser")
    if cookie_token and matches.filter(cookie_token=cookie_token).exists():
        match_count += 1
        matched_signals.append("cookie")

    blocked = match_count >= soglia
    if blocked:
        logger.warning(
            "Fingerprint EXPIRED TRIAL BLOCK: user=%s match_count=%d "
            "soglia=%d signals=%s ip=%s",
            user.username, match_count, soglia,
            ",".join(matched_signals), ip_hash[:12],
        )
    return blocked


def _setup_trial_for_user(user):
    """Crea o aggiorna il profilo utente con periodo di prova.

    Se il profilo esiste già (es. da precedente login Google),
    non sovrascrive trial_start/trial_end.
    """
    from .models import ImpostazioniSistema, UserProfile

    imp = ImpostazioniSistema.get()
    profile, created = UserProfile.objects.get_or_create(user=user)

    if created or profile.trial_start is None:
        profile.trial_start = timezone.now()
        profile.trial_end = timezone.now() + timedelta(days=imp.giorni_prova)
        profile.save(update_fields=["trial_start", "trial_end"])

    return profile


def _send_verification_email(request, user, profile):
    """Invia email di verifica con link token-based (monouso, 24h)."""
    import secrets
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    from django.utils.html import strip_tags

    token = secrets.token_urlsafe(32)
    profile.email_verification_token = token
    profile.email_verification_sent_at = timezone.now()
    profile.save(update_fields=["email_verification_token", "email_verification_sent_at"])

    verify_url = request.build_absolute_uri(f"/verify-email/{token}/")

    subject = "Verifica la tua email — Carico 3D"
    html_message = render_to_string("caricamento/verify_email.html", {
        "user": user,
        "verify_url": verify_url,
    })
    plain_message = strip_tags(html_message)

    send_mail(subject, plain_message, None, [user.email], html_message=html_message)
    logger.info("Email di verifica inviata: user=%s email=%s", user.username, user.email)


def _save_demo_fingerprints(request, user):
    """Salva i segnali solo quando l'anti-abuso è attivo."""
    from .models import DemoFingerprint, ImpostazioniSistema, UserProfile

    if not ImpostazioniSistema.get().controlli_demo_attivi:
        return

    try:
        profile = user.profile
    except UserProfile.DoesNotExist:
        return

    ip_raw = _client_ip(request)
    ip_hash = hashlib.sha256(ip_raw.encode()).hexdigest() if ip_raw else ""
    browser_hash = request.COOKIES.get("cb_fp", "") or None
    cookie_token = _demo_cookie_token(request) or None

    DemoFingerprint.objects.get_or_create(
        user_profile=profile,
        ip_hash=ip_hash,
        defaults={
            "browser_hash": browser_hash,
            "cookie_token": cookie_token,
        },
    )


# ===========================================================================
# VIEW: Landing Page / Homepage
# ===========================================================================

def homepage(request):
    """Landing page pubblica con workspace in vetrina + login.

    - Utenti autenticati: verifica trial, poi redirect al workspace.
    - Google OAuth: gestito da django-allauth (callback su /accounts/google/).
    - POST demo: autentica con account demo + 3 controlli anti-abuso.
    - GET: mostra landing page con form login e pulsante Google.
    """
    from django.contrib.auth import authenticate, login as auth_login
    from .models import ImpostazioniSistema, UserProfile

    imp = ImpostazioniSistema.get()
    _demo_cookie_token(request)

    # Helper per validare il parametro next (anti open-redirect)
    def _safe_next(request, default="caricamento:workspace"):
        next_url = request.POST.get("next") or request.GET.get("next") or ""
        if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts=None):
            return next_url
        return default

    # Se già autenticato, verifica trial e redirect. Un account esistente
    # resta utilizzabile da qualunque dispositivo: il fingerprint controlla
    # soltanto la richiesta di un nuovo trial.
    next_url = _safe_next(request)

    if request.user.is_authenticated:
        profile = _setup_trial_for_user(request.user)
        # Se email non verificata e l'utente ha un'email, blocca l'accesso
        # (ma non logout: l'utente può ancora vedere la pagina pagamenti)
        email_not_verified = (
            request.user.email
            and not profile.email_verified
            and not request.user.is_staff
        )
        if not profile.is_trial_active and not request.user.is_staff:
            from django.contrib.auth import logout
            logout(request)
            response = render(request, "caricamento/landing.html", {
                "login_error": False,
                "trial_expired": True,
                "trial_used": False,
                "email_verification_sent": False,
                "demo_user": "",
                "demo_pass": "",
                "oggetti_demo": [],
                "google_oauth_attivo": imp.google_oauth_attivo,
                "demo_attiva": imp.demo_attiva,
                "switch_account": False,
            })
            return _attach_demo_cookie(request, response)
        # Blocca accesso al workspace se email non verificata.
        # L'utente rimane autenticato per vedere la pagina pagamenti.
        if email_not_verified:
            response = render(request, "caricamento/landing.html", {
                "login_error": False,
                "trial_expired": False,
                "trial_used": False,
                "email_verification_sent": True,
                "demo_user": "",
                "demo_pass": "",
                "oggetti_demo": [],
                "google_oauth_attivo": imp.google_oauth_attivo,
                "demo_attiva": imp.demo_attiva,
                "switch_account": False,
            })
            return _attach_demo_cookie(request, response)
        return _attach_demo_cookie(request, redirect(next_url))

    login_error = False
    password_confirmation_error = False
    email_verification_sent = False
    trial_expired = request.GET.get("trial") == "expired"
    trial_used = request.GET.get("trial") == "used"
    trial_blocked = request.GET.get("trial") == "blocked"  # trial scaduto + fingerprint sovrapposto
    account_disabled = request.GET.get("account") == "disabled"
    demo_disabled = request.GET.get("demo") == "disabled"
    switch_account = request.GET.get("switch_account", "") == "1"

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()
        password_confirm = request.POST.get("password_confirm", "")
        email = request.POST.get("email", "").strip()

        if not username or not password:
            login_error = True
        else:
            user = authenticate(request, username=username, password=password)

            if user is not None:
                # Account esistente: nessun blocco fingerprint. Il trial già
                # scaduto resta scaduto, mentre un account pagante passa sempre.
                profile = _setup_trial_for_user(user)
                # Email non verificata: blocca il login finché l'utente non
                # clicca il link ricevuto via email. Se il vecchio link è
                # scaduto, ne inviamo uno nuovo al volo.
                if user.email and not profile.email_verified and not user.is_staff:
                    _send_verification_email(request, user, profile)
                    email_verification_sent = True
                    logger.info("Login bloccato: email non verificata user=%s", user.username)
                # ``demo_attiva`` controlla solo la creazione di nuovi
                # account demo. Un account già esistente deve poter accedere
                # finché è attivo e il suo trial non è scaduto (come già
                # avviene per un'identità Google esistente).
                # Se il trial è scaduto, verifica anche il fingerprint per
                # bloccare il cambio identità dallo stesso device/rete.
                elif not profile.is_trial_active and not user.is_staff:
                    if _check_fingerprint_expired_trial(request, user):
                        trial_blocked = True
                        logger.warning(
                            "Login bloccato: trial scaduto + fingerprint "
                            "sovrapposto user=%s ip=%s",
                            user.username, _client_ip(request),
                        )
                    else:
                        trial_expired = True
                else:
                    auth_login(request, user)
                    _save_demo_fingerprints(request, user)
                    return _attach_demo_cookie(request, redirect(_safe_next(request)))
            else:
                # authenticate() restituisce None anche per un account
                # disattivato. Verifica prima l'esistenza dello username, così
                # non tentiamo erroneamente di ricrearlo con lo stesso nome.
                existing_user = User.objects.filter(username=username).first()
                if existing_user is not None:
                    if not existing_user.is_active:
                        account_disabled = True
                    else:
                        login_error = True
                # Utente realmente nuovo: il fingerprint decide se è possibile
                # assegnare il primo e unico trial per questa impronta.
                elif not imp.demo_attiva:
                    login_error = True
                else:
                    # Un nuovo account richiede sempre la conferma password.
                    # Il controllo server-side evita account creati con un refuso,
                    # indipendentemente da quanto mostrato dal frontend.
                    if not email:
                        login_error = True
                    elif not password_confirm or password != password_confirm:
                        password_confirmation_error = True
                    else:
                        # Serializza il controllo e la registrazione locale:
                        # senza questo lock due POST simultanei potrebbero
                        # superare entrambi il controllo prima del salvataggio.
                        with transaction.atomic():
                            ImpostazioniSistema.objects.select_for_update().get(pk=1)
                            if _check_demo_abuse(request):
                                trial_used = True
                                logger.warning("Demo auto-create blocked: username=%s", username)
                            else:
                                user = User.objects.create_user(
                                    username=username,
                                    password=password,
                                    email=email,
                                )
                                profile = _setup_trial_for_user(user)
                                # Invia email di verifica
                                _send_verification_email(request, user, profile)
                                # Messaggio di conferma registrazione + verifica email
                                response = render(request, "caricamento/landing.html", {
                                    "login_error": False,
                                    "password_confirmation_error": False,
                                    "trial_expired": False,
                                    "trial_used": False,
                                    "account_disabled": False,
                                    "demo_disabled": False,
                                    "email_verification_sent": True,
                                    "demo_user": "",
                                    "demo_pass": "",
                                    "oggetti_demo": [],
                                    "google_oauth_attivo": imp.google_oauth_attivo,
                                    "demo_attiva": imp.demo_attiva,
                                    "switch_account": False,
                                })
                                return _attach_demo_cookie(request, response)

    # GET: prepara il contesto per la landing page

    # Dati dimostrativi statici: la landing è pubblica e non deve leggere
    # l'anagrafica privata degli oggetti di alcun utente.
    oggetti_demo = [
        {
            "codice": "DEMO-A01",
            "descrizione": "Scatola ricambi auto",
            "colore": "#447e9b",
            "dimensioni": "60×40×30 cm",
            "peso_kg": "35 kg",
            "quantita": 6,
        },
        {
            "codice": "DEMO-B01",
            "descrizione": "Quadro elettrico",
            "colore": "#cc44ff",
            "dimensioni": "80×60×120 cm",
            "peso_kg": "120 kg",
            "quantita": 2,
        },
        {
            "codice": "DEMO-C01",
            "descrizione": "Pallet industriale",
            "colore": "#ff6644",
            "dimensioni": "120×80×100 cm",
            "peso_kg": "250 kg",
            "quantita": 4,
        },
    ]

    return render(request, "caricamento/landing.html", {
        "login_error": login_error,
        "password_confirmation_error": password_confirmation_error,
        "trial_expired": trial_expired,
        "trial_used": trial_used,
        "trial_blocked": trial_blocked,
        "account_disabled": account_disabled,
        "demo_disabled": demo_disabled,
        "email_verification_sent": email_verification_sent,
        "demo_user": "",
        "demo_pass": "",
        "oggetti_demo": oggetti_demo,
        "google_oauth_attivo": imp.google_oauth_attivo,
        "demo_attiva": imp.demo_attiva,
        "switch_account": switch_account,
    })


def verify_email(request, token):
    """Verifica email tramite token monouso (scadenza 24h).

    Dopo la verifica, fa il login automatico e reindirizza al workspace.
    """
    from .models import UserProfile
    from django.contrib.auth import login as auth_login

    profile = UserProfile.objects.filter(email_verification_token=token).select_related("user").first()

    if not profile or not profile.email_verification_sent_at:
        return render(request, "caricamento/verify_email_confirm.html", {"success": False})

    # Scadenza 24 ore
    from datetime import timedelta
    if timezone.now() - profile.email_verification_sent_at > timedelta(hours=24):
        return render(request, "caricamento/verify_email_confirm.html", {"success": False})

    profile.email_verified = True
    profile.email_verification_token = ""
    profile.email_verification_sent_at = None
    profile.save(update_fields=["email_verified", "email_verification_token", "email_verification_sent_at"])

    logger.info("Email verificata: user=%s email=%s", profile.user.username, profile.user.email)

    # Login automatico dopo la verifica
    user = profile.user
    user.backend = "django.contrib.auth.backends.ModelBackend"
    auth_login(request, user)
    _setup_trial_for_user(user)

    return redirect("caricamento:workspace")


def check_username(request):
    """Endpoint JSON per il form login/registrazione della landing.

    Restituisce {"exists": true/false}: il frontend usa il risultato per
    mostrare il campo "Conferma password" solo quando lo username non esiste
    ancora (registrazione) e renderlo opaco/opzionale per l'accesso a un
    account esistente.
    """
    username = (request.GET.get("username") or "").strip()
    exists = bool(username) and User.objects.filter(username=username).exists()
    return JsonResponse({"exists": exists})


def logout_completo(request):
    """Logout completo: cancella la sessione Django e forza la scelta account
    Google al prossimo accesso (per utenti Google). Per utenti non-Google
    il comportamento è identico al logout normale.

    Reindirizza alla landing page con ?switch_account=1.
    """
    from django.contrib.auth import logout
    logout(request)
    return redirect("/?switch_account=1")


# ---------------------------------------------------------------------------
# Pagine legali (Privacy, Cookie Policy, Termini, Rimborsi)
# ---------------------------------------------------------------------------

# Slug -> (template, titolo pagina). Solo questi slug sono ammessi.
_PAGINE_LEGALI = {
    # English (default)
    "privacy": ("caricamento/privacy.html", "Privacy Policy"),
    "cookie-policy": ("caricamento/cookie_policy.html", "Cookie Policy"),
    "termini": ("caricamento/termini.html", "Terms of Service"),
    "rimborsi": ("caricamento/rimborsi.html", "Refund Policy"),
    # Italian
    "privacy-it": ("caricamento/privacy_it.html", "Informativa Privacy"),
    "cookie-policy-it": ("caricamento/cookie_it.html", "Cookie Policy"),
    "termini-it": ("caricamento/termini_it.html", "Termini di Servizio"),
    "rimborsi-it": ("caricamento/rimborsi_it.html", "Informativa Rimborsi"),
}


def pagina_legale(request, slug):
    """Rende le pagine legali pubbliche (senza login).

    I dati del titolare (nome, email, sede, P.IVA, URL base) arrivano dal
    singleton ``ImpostazioniSistema``, configurabile dall'admin nella sezione
    "Privacy — Dati del Titolare".
    """
    pagina = _PAGINE_LEGALI.get(slug)
    if pagina is None:
        return redirect("/")
    template, titolo = pagina
    imp = ImpostazioniSistema.get()
    context = {
        "titolo_pagina": titolo,
        "titolare": imp.privacy_titolare,
        "email_titolare": imp.privacy_email,
        "sede_titolare": imp.privacy_sede,
        "piva_titolare": imp.privacy_piva,
        "sito_url": imp.privacy_sito_url,
    }
    return render(request, template, context)


@login_required
def workspace(request, piano_id=None):
    """Pagina Workspace unificata (Single-Page App).

    Sostituisce l'admin per la gestione quotidiana: permette di
    creare piani, aggiungere oggetti, eseguire ottimizzazione e
    vedere il risultato 3D tutto in un'unica pagina.

    Piani e oggetti sono paginati (50 per pagina) per evitare di
    caricare centinaia di record nel frontend.
    Per le parti che necessitano di tutti gli oggetti (dropdown,
    colonne Vincoli-tra), viene fornito un catalogo leggero
    (id, codice, dimensioni).
    """
    from django.core.paginator import Paginator
    from django.db.models import Prefetch

    piani_qs = PianoDiCarico.objects.filter(
        owner=request.user,
    ).select_related("contenitore").order_by("-created_at")
    paginator = Paginator(piani_qs, 50)
    page_number = request.GET.get("page", 1)
    piani = paginator.get_page(page_number)

    contenitori = Contenitore.objects.filter(
        owner=request.user,
    ).prefetch_related("sezioni").order_by("nome")

    # --- Oggetti: paginati (50/pag) con vincoli prefetched ---
    oggetti_qs = Oggetto.objects.filter(
        owner=request.user,
    ).prefetch_related(
        Prefetch("vincoli", queryset=VincoloOggetto.objects.all())
    ).order_by("codice")
    oggetti_paginator = Paginator(oggetti_qs, 50)
    oggetti_page_number = request.GET.get("oggetti_page", 1)
    oggetti = oggetti_paginator.get_page(oggetti_page_number)

    # --- Vincoli: TUTTI (tabella piccola, serve lookup per oggetto_id nel frontend) ---
    vincoli = VincoloOggetto.objects.filter(oggetto__owner=request.user)

    # --- Catalogo oggetti leggero: TUTTI (serve per dropdown e colonne Vincoli-tra) ---
    oggetti_catalog = list(Oggetto.objects.filter(owner=request.user).values(
        "id", "codice", "descrizione", "lunghezza_mm", "larghezza_mm", "altezza_mm",
        "peso_kg", "quantita_disponibile", "colore",
    ).order_by("codice"))

    vincoli_tra_oggetti = VincoloTraOggetti.objects.filter(
        oggetto_a__owner=request.user,
        oggetto_b__owner=request.user,
    ).select_related(
        "oggetto_a", "oggetto_b"
    ).order_by("-created_at")

    # Serializza dettagli_posizionamento come JSON per il frontend
    vincoli_tra_js = []
    for v in vincoli_tra_oggetti:
        dett_json = json.dumps(v.dettagli_posizionamento) if v.dettagli_posizionamento else "null"
        vincoli_tra_js.append({
            "v": v,
            "dettagli_json": dett_json,
        })

    # La configurazione icone viene iniettata nella pagina (json_script nel
    # template) così il frontend la applica subito, senza fetch né flash di
    # icone Bootstrap. Stessa fonte dell'endpoint /api/icone-config/.
    icon_config = _load_icon_config()

    # Offer Fungies per la vista Abbonamento (vuote finché non configurate)
    fungies_offer_ids = {
        "mensile": getattr(settings, "FUNGIES_OFFER_MENSILE_ID", ""),
        "annuale": getattr(settings, "FUNGIES_OFFER_ANNUALE_ID", ""),
    }

    # Get user language preference
    lingua = 'en'
    profile = getattr(request.user, 'profile', None)
    if profile and profile.lingua:
        lingua = profile.lingua

    return render(request, "caricamento/workspace.html", {
        "piani": piani,
        "contenitori": contenitori,
        "oggetti": oggetti,
        "oggetti_catalog": oggetti_catalog,
        "vincoli": vincoli,
        "vincoli_tra_oggetti": vincoli_tra_js,
        "piano_id": piano_id,
        "icon_config": icon_config,
        "fungies_offer_ids": fungies_offer_ids,
        "fungies_portal_url": (
            getattr(settings, "FUNGIES_PORTAL_URL", "").strip()
            or f"{getattr(settings, 'FUNGIES_STORE_URL', '').strip().rstrip('/')}/portal"
        ),
        "user_email": request.user.email or "",
        "user_id": request.user.id,
        "lingua": lingua,
    })


# ===========================================================================
# API: Preferenze personali del workspace e dell'ottimizzatore
# ===========================================================================

@api_view(["GET", "PUT"])
@permission_classes([IsAuthenticated])
def api_impostazioni_ottimizzatore(request):
    """Legge o salva le preferenze dell'ottimizzatore dell'utente autenticato.

    GET /api/impostazioni_ottimizzatore/
        Restituisce ``{"impostazioni": {...}}``.
    PUT /api/impostazioni_ottimizzatore/
        Aggiorna le sezioni inviate della configurazione personale,
        preservando le sezioni e i campi non inclusi nel payload.

    Le preferenze sono volutamente associate a ``UserProfile`` e non a
    ``ImpostazioniSistema``: strategie e opzioni di visualizzazione sono
    personali, mentre ImpostazioniSistema contiene la configurazione globale
    dell'applicazione.
    """
    profile, _ = UserProfile.objects.get_or_create(user=request.user)

    if request.method == "GET":
        return Response({
            "impostazioni": profile.impostazioni_ottimizzatore or {},
        })

    serializer = ImpostazioniOttimizzatoreSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    payload = serializer.validated_data

    # PUT aggiorna le sezioni inviate e preserva le altre per consentire
    # aggiornamenti parziali senza perdere preferenze già salvate.
    impostazioni = dict(profile.impostazioni_ottimizzatore or {})
    impostazioni.update({
        nome: dict(valori) for nome, valori in payload.items()
    })
    profile.impostazioni_ottimizzatore = impostazioni
    profile.save(update_fields=["impostazioni_ottimizzatore", "updated_at"])

    return Response({
        "success": True,
        "impostazioni": impostazioni,
    })


def visualizzatore_3d(request, piano_id=None):
    """Redirect alla nuova pagina Workspace (sostituisce il vecchio visualizzatore)."""
    if piano_id:
        return redirect("caricamento:workspace_piano", piano_id=piano_id)
    return redirect("caricamento:workspace")


# ===========================================================================
# ViewSet: Contenitore
# ===========================================================================

class ContenitoreViewSet(viewsets.ModelViewSet):
    """
    API endpoint per la gestione dei contenitori/veicoli.
    Richiede autenticazione."""
    permission_classes = [IsAuthenticated]

    queryset = Contenitore.objects.prefetch_related("sezioni").all()

    def get_queryset(self):
        qs = super().get_queryset().filter(owner=self.request.user)
        if self.action == "list":
            mostra_archiviati = self.request.query_params.get("archiviati", "") == "1"
            qs = qs.filter(archiviato=mostra_archiviati)
        return qs

    def perform_create(self, serializer):
        serializer.save(owner=self.request.user)

    def perform_update(self, serializer):
        serializer.save(owner=self.request.user)

    def get_serializer_class(self):
        if self.action == "list":
            return ContenitoreListSerializer
        if self.action == "create":
            return ContenitoreCreateSerializer
        return ContenitoreDetailSerializer

    def destroy(self, request, *args, **kwargs):
        """Elimina un contenitore. Blocca con 409 se ci sono piani di carico collegati."""
        instance = self.get_object()
        try:
            self.perform_destroy(instance)
        except ProtectedError:
            num_piani = instance.piani_di_carico.count()
            return Response(
                {
                    "detail": (
                        f"Impossibile eliminare '{instance.nome}': "
                        f"è usato in {num_piani} piano/i di carico. "
                        f"Archivia il mezzo o elimina prima i piani collegati."
                    ),
                },
                status=status.HTTP_409_CONFLICT,
            )
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=["get", "put"])
    def sezioni(self, request, pk=None):
        """
        Legge o sostituisce in blocco le sezioni di carico di un contenitore.

        GET  /api/contenitori/{id}/sezioni/   → Restituisce le sezioni attuali
        PUT  /api/contenitori/{id}/sezioni/   → Sostituisce TUTTE le sezioni

        Corpo PUT (JSON):
            [
                {"nome": "Zona 1", "inizio_x_mm": 0, "fine_x_mm": 10000, "carico_massimo_kg": 9000},
                {"nome": "Zona 2", "inizio_x_mm": 10000, "fine_x_mm": 13600, "carico_massimo_kg": 9000}
            ]
        """
        contenitore = self.get_object()

        if request.method == "GET":
            sezioni = contenitore.sezioni.all()
            ser = SezioneCaricoSerializer(sezioni, many=True)
            return Response(ser.data)

        # PUT: sostituisci tutte le sezioni (atomico: o tutto o niente)
        ser = SezioneCaricoWriteSerializer(data=request.data, many=True)
        ser.is_valid(raise_exception=True)

        with transaction.atomic():
            contenitore.sezioni.all().delete()
            nuove_sezioni = []
            for item in ser.validated_data:
                nuove_sezioni.append(SezioneCarico(
                    contenitore=contenitore,
                    nome=item["nome"],
                    inizio_x_mm=item["inizio_x_mm"],
                    fine_x_mm=item["fine_x_mm"],
                    carico_massimo_kg=item["carico_massimo_kg"],
                ))
            SezioneCarico.objects.bulk_create(nuove_sezioni)

        # Rileggi e restituisci
        contenitore.refresh_from_db()
        sezioni = contenitore.sezioni.all()
        out = SezioneCaricoSerializer(sezioni, many=True)
        return Response(out.data, status=status.HTTP_200_OK)


# ===========================================================================
# Excel: esportazione/importazione anagrafica oggetti
# ===========================================================================

_EXCEL_OGGETTI_SHEET = "Oggetti"
_EXCEL_ROTazioni_SHEET = "Rotazioni"
_EXCEL_VINCOLI_SHEET = "Vincoli"


def _excel_valore_cella(value):
    return "" if value is None else str(value).strip()


def _excel_bool(value, default=False):
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "vero", "sì", "si", "yes", "y"}:
        return True
    if normalized in {"0", "false", "falso", "no", "n"}:
        return False
    raise ValueError("valore booleano non valido")


def _excel_int(value, field, row, minimum=0):
    try:
        number = Decimal(str(value).replace(",", "."))
    except (TypeError, ValueError, ArithmeticError):
        raise ValueError(f"{field} non è un numero intero")
    if number != number.to_integral_value():
        raise ValueError(f"{field} non è un numero intero")
    number = int(number)
    if number < minimum:
        raise ValueError(f"{field} deve essere almeno {minimum}")
    return number


def _excel_decimal(value, field, row, minimum=None):
    try:
        number = Decimal(str(value).replace(",", "."))
    except (TypeError, ValueError, ArithmeticError):
        raise ValueError(f"{field} non è un numero valido")
    if minimum is not None and number < minimum:
        raise ValueError(f"{field} deve essere almeno {minimum}")
    return number


def _excel_json(value, field):
    text = _excel_valore_cella(value)
    if not text:
        return None
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        raise ValueError(f"{field} non contiene JSON valido")
    if not isinstance(parsed, dict):
        raise ValueError(f"{field} deve contenere un oggetto JSON")
    return parsed


def _excel_workbook_headers(sheet, expected):
    headers = [_excel_valore_cella(cell.value) for cell in sheet[1]]
    if headers != expected:
        raise ValueError(
            f"Il foglio '{sheet.title}' deve avere le colonne: {', '.join(expected)}"
        )


def _excel_rows(sheet, expected):
    _excel_workbook_headers(sheet, expected)
    rows = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if not any(value not in (None, "") for value in values):
            continue
        if len(values) < len(expected):
            values = tuple(values) + (None,) * (len(expected) - len(values))
        rows.append((row_number, dict(zip(expected, values))))
    return rows


def _excel_headers_and_widths(sheet):
    from openpyxl.styles import Font

    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        values = [len(str(cell.value or "")) for cell in column]
        sheet.column_dimensions[column[0].column_letter].width = min(max(max(values, default=10) + 2, 12), 45)


def _crea_excel_oggetti(oggetti, vincoli_tra):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError(
            "La libreria openpyxl non è installata sul server. "
            "Installa le dipendenze del progetto e riprova."
        ) from exc

    wb = Workbook()
    ws_oggetti = wb.active
    ws_oggetti.title = _EXCEL_OGGETTI_SHEET
    headers_oggetti = [
        "Codice", "Descrizione", "Lunghezza_mm", "Larghezza_mm", "Altezza_mm",
        "Peso_kg", "Quantita_disponibile", "Colore", "Archiviato",
    ]
    ws_oggetti.append(headers_oggetti)
    for oggetto in oggetti:
        ws_oggetti.append([
            oggetto.codice, oggetto.descrizione, oggetto.lunghezza_mm,
            oggetto.larghezza_mm, oggetto.altezza_mm, float(oggetto.peso_kg),
            oggetto.quantita_disponibile, oggetto.colore or "", oggetto.archiviato,
        ])

    ws_rotazioni = wb.create_sheet(_EXCEL_ROTazioni_SHEET)
    headers_rotazioni = [
        "Codice", "Rotazione_consentita", "Rotazione_su_X", "Rotazione_su_Y",
        "Rotazione_su_Z", "Sovrapponibile", "Peso_massimo_tetto_kg", "Fragile",
        "Merce_pericolosa", "Solo_su_piano", "Aggancio_forche", "Note",
    ]
    ws_rotazioni.append(headers_rotazioni)
    for oggetto in oggetti:
        try:
            vincolo = oggetto.vincoli
        except VincoloOggetto.DoesNotExist:
            vincolo = VincoloOggetto(oggetto=oggetto)
        ws_rotazioni.append([
            oggetto.codice, vincolo.rotazione_consentita, vincolo.rotazione_su_x,
            vincolo.rotazione_su_y, vincolo.rotazione_su_z, vincolo.sovrapponibile,
            float(vincolo.peso_massimo_tetto_kg), vincolo.fragile,
            vincolo.merce_pericolosa, vincolo.solo_su_piano, vincolo.aggancio_forche,
            vincolo.note or "",
        ])

    ws_vincoli = wb.create_sheet(_EXCEL_VINCOLI_SHEET)
    headers_vincoli = [
        "Oggetto_A", "Oggetto_B", "Tipo_relazione", "Attivo",
        "Dettagli_posizionamento", "Note",
    ]
    ws_vincoli.append(headers_vincoli)
    for vincolo in vincoli_tra:
        dettagli = (
            json.dumps(vincolo.dettagli_posizionamento, ensure_ascii=False)
            if vincolo.dettagli_posizionamento is not None else ""
        )
        ws_vincoli.append([
            vincolo.oggetto_a.codice, vincolo.oggetto_b.codice,
            vincolo.tipo_relazione, vincolo.attivo, dettagli, vincolo.note or "",
        ])

    for sheet in wb.worksheets:
        _excel_headers_and_widths(sheet)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _leggi_excel_oggetti(uploaded_file, user, mode):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "La libreria openpyxl non è installata sul server. "
            "Installa le dipendenze del progetto e riprova."
        ) from exc

    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"File Excel non leggibile: {exc}")

    required_sheets = {_EXCEL_OGGETTI_SHEET, _EXCEL_ROTazioni_SHEET, _EXCEL_VINCOLI_SHEET}
    if set(workbook.sheetnames) != required_sheets:
        raise ValueError("Il file deve contenere esattamente i fogli Oggetti, Rotazioni e Vincoli")

    object_headers = [
        "Codice", "Descrizione", "Lunghezza_mm", "Larghezza_mm", "Altezza_mm",
        "Peso_kg", "Quantita_disponibile", "Colore", "Archiviato",
    ]
    rotation_headers = [
        "Codice", "Rotazione_consentita", "Rotazione_su_X", "Rotazione_su_Y",
        "Rotazione_su_Z", "Sovrapponibile", "Peso_massimo_tetto_kg", "Fragile",
        "Merce_pericolosa", "Solo_su_piano", "Aggancio_forche", "Note",
    ]
    constraint_headers = [
        "Oggetto_A", "Oggetto_B", "Tipo_relazione", "Attivo",
        "Dettagli_posizionamento", "Note",
    ]

    errors = []
    objects = []
    object_codes = set()
    for row_number, row in _excel_rows(workbook[_EXCEL_OGGETTI_SHEET], object_headers):
        try:
            code = _excel_valore_cella(row["Codice"])
            if not code:
                raise ValueError("Codice obbligatorio")
            if code in object_codes:
                raise ValueError(f"Codice duplicato nel file: {code}")
            object_codes.add(code)
            objects.append({
                "codice": code,
                "descrizione": _excel_valore_cella(row["Descrizione"]),
                "lunghezza_mm": _excel_int(row["Lunghezza_mm"], "Lunghezza_mm", row_number, 1),
                "larghezza_mm": _excel_int(row["Larghezza_mm"], "Larghezza_mm", row_number, 1),
                "altezza_mm": _excel_int(row["Altezza_mm"], "Altezza_mm", row_number, 1),
                "peso_kg": _excel_decimal(row["Peso_kg"], "Peso_kg", row_number, Decimal("0.01")),
                "quantita_disponibile": _excel_int(row["Quantita_disponibile"], "Quantita_disponibile", row_number, 1),
                "colore": _excel_valore_cella(row["Colore"]),
                "archiviato": _excel_bool(row["Archiviato"]),
            })
        except ValueError as exc:
            errors.append(f"Oggetti riga {row_number}: {exc}")

    rotations = {}
    for row_number, row in _excel_rows(workbook[_EXCEL_ROTazioni_SHEET], rotation_headers):
        try:
            code = _excel_valore_cella(row["Codice"])
            if not code:
                raise ValueError("Codice obbligatorio")
            if code not in object_codes:
                raise ValueError(f"codice non presente nel foglio Oggetti: {code}")
            if code in rotations:
                raise ValueError(f"Codice duplicato nel file: {code}")
            rotations[code] = {
                "rotazione_consentita": _excel_bool(row["Rotazione_consentita"], True),
                "rotazione_su_x": _excel_bool(row["Rotazione_su_X"], True),
                "rotazione_su_y": _excel_bool(row["Rotazione_su_Y"], True),
                "rotazione_su_z": _excel_bool(row["Rotazione_su_Z"], True),
                "sovrapponibile": _excel_bool(row["Sovrapponibile"], True),
                "peso_massimo_tetto_kg": _excel_decimal(row["Peso_massimo_tetto_kg"] or 0, "Peso_massimo_tetto_kg", row_number, Decimal("0")),
                "fragile": _excel_bool(row["Fragile"]),
                "merce_pericolosa": _excel_bool(row["Merce_pericolosa"]),
                "solo_su_piano": _excel_bool(row["Solo_su_piano"]),
                "aggancio_forche": _excel_bool(row["Aggancio_forche"]),
                "note": _excel_valore_cella(row["Note"]),
            }
        except ValueError as exc:
            errors.append(f"Rotazioni riga {row_number}: {exc}")

    constraints = []
    valid_relation_types = {choice[0] for choice in TipoRelazione.choices}
    for row_number, row in _excel_rows(workbook[_EXCEL_VINCOLI_SHEET], constraint_headers):
        try:
            code_a = _excel_valore_cella(row["Oggetto_A"])
            code_b = _excel_valore_cella(row["Oggetto_B"])
            relation = _excel_valore_cella(row["Tipo_relazione"])
            if code_a not in object_codes or code_b not in object_codes:
                raise ValueError("entrambi i codici devono essere presenti nel foglio Oggetti")
            if relation not in valid_relation_types:
                raise ValueError(f"Tipo_relazione non valido: {relation}")
            constraints.append({
                "oggetto_a": code_a,
                "oggetto_b": code_b,
                "tipo_relazione": relation,
                "attivo": _excel_bool(row["Attivo"], True),
                "dettagli_posizionamento": _excel_json(row["Dettagli_posizionamento"], "Dettagli_posizionamento"),
                "note": _excel_valore_cella(row["Note"]),
            })
        except ValueError as exc:
            errors.append(f"Vincoli riga {row_number}: {exc}")

    existing = set(Oggetto.objects.filter(owner=user, codice__in=object_codes).values_list("codice", flat=True))
    if mode == "add":
        for code in sorted(existing):
            errors.append(f"Oggetti: il codice esiste già e non può essere aggiunto: {code}")

    existing_pairs = set(
        VincoloTraOggetti.objects.filter(
            oggetto_a__owner=user, oggetto_b__owner=user,
            oggetto_a__codice__in=object_codes, oggetto_b__codice__in=object_codes,
        ).values_list("oggetto_a__codice", "oggetto_b__codice", "tipo_relazione")
    )
    if mode == "add":
        for constraint in constraints:
            pair = (constraint["oggetto_a"], constraint["oggetto_b"], constraint["tipo_relazione"])
            if pair in existing_pairs:
                errors.append("Vincoli: relazione già esistente: " + " / ".join(pair))

    if errors:
        raise ValueError("Importazione non eseguita:\n" + "\n".join(errors[:30]))
    return objects, rotations, constraints


class OggettiExcelMixin:
    """Funzioni condivise dalle azioni Excel del ViewSet Oggetto."""

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        ids_text = request.query_params.get("ids", "")
        try:
            ids = {int(value) for value in ids_text.split(",") if value.strip()}
        except ValueError:
            return Response({"errore": "Selezione oggetti non valida."}, status=status.HTTP_400_BAD_REQUEST)
        if not ids:
            return Response({"errore": "Seleziona almeno un oggetto."}, status=status.HTTP_400_BAD_REQUEST)
        oggetti = list(Oggetto.objects.filter(owner=request.user, id__in=ids).prefetch_related("vincoli").order_by("codice"))
        if len(oggetti) != len(ids):
            return Response({"errore": "Uno o più oggetti selezionati non appartengono all'utente."}, status=status.HTTP_400_BAD_REQUEST)
        selected_ids = {obj.id for obj in oggetti}
        vincoli = VincoloTraOggetti.objects.filter(
            oggetto_a_id__in=selected_ids, oggetto_b_id__in=selected_ids,
        ).select_related("oggetto_a", "oggetto_b").order_by("oggetto_a__codice", "oggetto_b__codice")
        try:
            output = _crea_excel_oggetti(oggetti, vincoli)
        except RuntimeError as exc:
            return Response({"errore": str(exc)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="oggetti_loadplanner3d.xlsx"'
        return response

    @action(detail=False, methods=["post"], url_path="import-excel")
    def import_excel(self, request):
        uploaded_file = request.FILES.get("file")
        mode = request.data.get("modalita", "add")
        if not uploaded_file:
            return Response({"errore": "Seleziona un file Excel .xlsx."}, status=status.HTTP_400_BAD_REQUEST)
        if mode not in {"add", "update", "restore"}:
            return Response({"errore": "Modalità di importazione non valida."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            objects, rotations, constraints = _leggi_excel_oggetti(uploaded_file, request.user, mode)
        except RuntimeError as exc:
            return Response({"errore": str(exc)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        except ValueError as exc:
            return Response({"errore": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        created_objects = updated_objects = 0
        created_constraints = updated_constraints = 0
        with transaction.atomic():
            object_map = {}
            for data in objects:
                obj, created = Oggetto.objects.get_or_create(
                    owner=request.user, codice=data["codice"], defaults=data,
                )
                if created:
                    created_objects += 1
                else:
                    if mode in {"update", "restore"}:
                        for field, value in data.items():
                            setattr(obj, field, value)
                        obj.save()
                        updated_objects += 1
                object_map[data["codice"]] = obj

            for code, data in rotations.items():
                vincolo, _ = VincoloOggetto.objects.get_or_create(oggetto=object_map[code])
                for field, value in data.items():
                    setattr(vincolo, field, value)
                vincolo.save()

            codes = set(object_map)
            if mode == "restore":
                VincoloTraOggetti.objects.filter(
                    oggetto_a__owner=request.user, oggetto_b__owner=request.user,
                    oggetto_a__codice__in=codes, oggetto_b__codice__in=codes,
                ).delete()

            for data in constraints:
                vincolo, created = VincoloTraOggetti.objects.update_or_create(
                    oggetto_a=object_map[data["oggetto_a"]],
                    oggetto_b=object_map[data["oggetto_b"]],
                    tipo_relazione=data["tipo_relazione"],
                    defaults={
                        "attivo": data["attivo"],
                        "dettagli_posizionamento": data["dettagli_posizionamento"],
                        "note": data["note"],
                    },
                )
                if created:
                    created_constraints += 1
                else:
                    updated_constraints += 1

        return Response({
            "successo": True,
            "modalita": mode,
            "oggetti_aggiunti": created_objects,
            "oggetti_aggiornati": updated_objects,
            "vincoli_aggiunti": created_constraints,
            "vincoli_aggiornati": updated_constraints,
            "vincoli_importati": len(constraints),
        })


class OggettoViewSet(OggettiExcelMixin, viewsets.ModelViewSet):
    """
    API endpoint per la gestione degli oggetti/pacchi da caricare.
    Richiede autenticazione."""
    permission_classes = [IsAuthenticated]

    queryset = Oggetto.objects.all()

    def get_queryset(self):
        qs = super().get_queryset().filter(owner=self.request.user)
        if self.action == "list":
            mostra_archiviati = self.request.query_params.get("archiviati", "") == "1"
            qs = qs.filter(archiviato=mostra_archiviati)
        return qs

    def perform_create(self, serializer):
        serializer.save(owner=self.request.user)

    def perform_update(self, serializer):
        serializer.save(owner=self.request.user)

    def get_serializer_class(self):
        if self.action == "list":
            return OggettoListSerializer
        if self.action == "create":
            return OggettoCreateSerializer
        return OggettoDetailSerializer

    def destroy(self, request, *args, **kwargs):
        """Elimina un oggetto. Blocca con 409 se è posizionato in piani di carico."""
        instance = self.get_object()
        try:
            self.perform_destroy(instance)
        except ProtectedError:
            num_piani = instance.posizionamenti.count()
            return Response(
                {
                    "detail": (
                        f"Impossibile eliminare '{instance.codice}': "
                        f"è posizionato in {num_piani} piano/i di carico. "
                        f"Elimina prima i posizionamenti o i piani collegati."
                    ),
                },
                status=status.HTTP_409_CONFLICT,
            )
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=["get", "patch"])
    def vincoli(self, request, pk=None):
        """
        Legge o aggiorna i vincoli di un oggetto.

        GET  /api/oggetti/{id}/vincoli/  → Restituisce i vincoli attuali
        PATCH /api/oggetti/{id}/vincoli/ → Aggiorna i vincoli (parziale)

        Esempio corpo PATCH:
            { "rotazione_su_x": false, "sovrapponibile": false }
        """
        oggetto = self.get_object()
        vincolo, created = VincoloOggetto.objects.get_or_create(oggetto=oggetto)

        if request.method == "GET":
            from .serializers import VincoloOggettoSerializer
            ser = VincoloOggettoSerializer(vincolo)
            return Response(ser.data)

        # PATCH
        serializer = VincoloOggettoUpdateSerializer(
            vincolo, data=request.data, partial=True
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @action(detail=False, methods=["post"])
    def bulk_delete(self, request):
        """
        Elimina multipli oggetti in blocco.

        POST /api/oggetti/bulk_delete/
        Corpo: { "ids": [1, 2, 3] }
        """
        ids = request.data.get("ids", [])
        if not ids:
            return Response(
                {"errore": "Nessun ID fornito."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        qs = Oggetto.objects.filter(id__in=ids, owner=request.user)
        deleted_count, _ = qs.delete()
        return Response({"eliminati": deleted_count})

    @action(detail=False, methods=["post"])
    def bulk_vincoli(self, request):
        """
        Aggiorna i vincoli di multipli oggetti in blocco.

        POST /api/oggetti/bulk_vincoli/
        Corpo: { "ids": [1, 2, 3], "vincoli": { "fragile": true, "sovrapponibile": false } }
        """
        ids = request.data.get("ids", [])
        vincoli_data = request.data.get("vincoli", {})
        if not ids:
            return Response(
                {"errore": "Nessun ID fornito."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not vincoli_data:
            return Response(
                {"errore": "Nessun vincolo da aggiornare."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Prende o crea i VincoloOggetto per tutti gli ID
        vincoli_qs = VincoloOggetto.objects.filter(
            oggetto_id__in=ids,
            oggetto__owner=request.user,
        )
        vincoli_map = {v.oggetto_id: v for v in vincoli_qs}
        owned_ids = set(
            Oggetto.objects.filter(id__in=ids, owner=request.user)
            .values_list("id", flat=True)
        )
        if owned_ids != set(ids):
            raise ValidationError({"ids": "Uno o più oggetti non appartengono all'utente."})

        aggiornati = 0
        for oid in ids:
            vincolo = vincoli_map.get(oid)
            if vincolo is None:
                vincolo = VincoloOggetto(oggetto_id=oid)
            # Il serializer applica SOLO i campi della whitelist (vincoli fisici).
            # I campi strutturali inviati da un client malevolo (oggetto_id, id, ...)
            # vengono ignorati, impedendo la riassegnazione cross-tenant del vincolo.
            ser = VincoloOggettoUpdateSerializer(
                vincolo, data=vincoli_data, partial=True
            )
            ser.is_valid(raise_exception=True)
            ser.save()
            aggiornati += 1

        return Response({"aggiornati": aggiornati})


# ===========================================================================
# Barriera di priorità nel salvataggio posizioni
# ===========================================================================


def _applica_barriera_priorita(nuovi, piano):
    """Ripiazza in coda gli oggetti a priorità 0 che risultano posizionati
    prima della fine X dei prioritari.

    La priorità ha la precedenza su tutto: un lotto a priorità 0 non può
    mai restare dentro i blocchi delle priorità 1 (non sarebbe scaricabile
    senza smontare il carico). Se il payload salvato (soluzione alternativa
    generata prima del fix, client legacy) viola la barriera, i posizionamenti
    a priorità 0 vengono ricompattati nella fascia X terminale con un
    first-fit Y → X → Z, rispettando il contenitore e le collisioni.

    ``nuovi`` è la lista di ``OggettoPosizionato`` non ancora salvati;
    la funzione la modifica in-place e la restituisce.
    """
    righe_prio = {
        riga.pk: (riga.priorita or 0)
        for riga in piano.oggetti_da_caricare.all()
    }

    def _priorita(op):
        return righe_prio.get(op.riga_origine_id, 0)

    prio1 = [op for op in nuovi if _priorita(op) > 0]
    prio0 = [op for op in nuovi if _priorita(op) == 0]
    if not prio1 or not prio0:
        return nuovi

    barriera = max(
        op.coordinata_x_mm + op.dimensione_x_mm for op in prio1
    )
    violatori = [op for op in prio0 if op.coordinata_x_mm < barriera - 0.5]
    if not violatori:
        return nuovi

    contenitore = piano.contenitore
    L = contenitore.lunghezza_mm
    W = contenitore.larghezza_mm
    H = contenitore.altezza_mm

    # Riferimento per collisioni: i prioritari + i prio0 già in coda
    # (non violatori) che restano al loro posto.
    placed = list(prio1) + [
        op for op in prio0 if op not in violatori
    ]

    def _collide(x, y, z, dx, dy, dz):
        if x + dx > L + 0.5 or y + dy > W + 0.5 or z + dz > H + 0.5:
            return True
        for p in placed:
            if (
                x < p.coordinata_x_mm + p.dimensione_x_mm - 0.5
                and x + dx > p.coordinata_x_mm + 0.5
                and y < p.coordinata_y_mm + p.dimensione_y_mm - 0.5
                and y + dy > p.coordinata_y_mm + 0.5
                and z < p.coordinata_z_mm + p.dimensione_z_mm - 0.5
                and z + dz > p.coordinata_z_mm + 0.5
            ):
                return True
        return False

    def _ha_supporto(x, y, z, dx, dy):
        if z <= 0.5:
            return True
        for p in placed:
            if abs(p.coordinata_z_mm + p.dimensione_z_mm - z) < 0.5:
                if (
                    x < p.coordinata_x_mm + p.dimensione_x_mm - 0.5
                    and x + dx > p.coordinata_x_mm + 0.5
                    and y < p.coordinata_y_mm + p.dimensione_y_mm - 0.5
                    and y + dy > p.coordinata_y_mm + 0.5
                ):
                    return True
        return False

    # Primo fit: i più ingombranti prima, per saturare meglio la coda.
    violatori.sort(key=lambda op: -op.dimensione_x_mm * op.dimensione_y_mm)

    for op in violatori:
        dx, dy, dz = op.dimensione_x_mm, op.dimensione_y_mm, op.dimensione_z_mm
        piazzato = False

        # Candidati X: la barriera e i bordi destri degli oggetti in coda.
        xs = {barriera}
        for p in placed:
            xs.add(p.coordinata_x_mm + p.dimensione_x_mm)

        for x in sorted(xs):
            if x < barriera - 0.5:
                continue
            if x + dx > L + 0.5:
                continue

            # Candidati Y: 0 e i bordi di chi si sovrappone in X.
            ys = {0}
            for p in placed:
                if p.coordinata_x_mm < x + dx and p.coordinata_x_mm + p.dimensione_x_mm > x:
                    ys.add(p.coordinata_y_mm + p.dimensione_y_mm)

            for y in sorted(ys):
                if y + dy > W + 0.5:
                    continue

                # Candidati Z: pavimento o sommità di chi copre il footprint.
                zs = {0}
                for p in placed:
                    if (
                        p.coordinata_x_mm < x + dx
                        and p.coordinata_x_mm + p.dimensione_x_mm > x
                        and p.coordinata_y_mm < y + dy
                        and p.coordinata_y_mm + p.dimensione_y_mm > y
                    ):
                        zs.add(p.coordinata_z_mm + p.dimensione_z_mm)

                for z in sorted(zs):
                    if z + dz > H + 0.5:
                        continue
                    if not _ha_supporto(x, y, z, dx, dy):
                        continue
                    if _collide(x, y, z, dx, dy, dz):
                        continue
                    op.coordinata_x_mm = x
                    op.coordinata_y_mm = y
                    op.coordinata_z_mm = z
                    placed.append(op)
                    piazzato = True
                    break
                if piazzato:
                    break
            if piazzato:
                break

        if not piazzato:
            raise ValidationError({
                "oggetti": (
                    "Oggetti a priorità 0 non collocabili in coda senza "
                    "sovrapporre i prioritari: riduci le quantità o rivedi "
                    "le priorità."
                )
            })

    return nuovi


# ===========================================================================
# ViewSet: PianoDiCarico (con azioni personalizzate)
# ===========================================================================

class PianoDiCaricoViewSet(viewsets.ModelViewSet):
    """
    API endpoint per la gestione dei piani di carico.
    Richiede autenticazione.

    ## Endpoint standard
    list:       Elenco di tutti i piani di carico
    create:     Crea un nuovo piano di carico
    read:       Dettaglio completo (contenitore + oggetti posizionati) ← ★
    update:     Aggiorna un piano di carico
    delete:     Elimina un piano di carico

    ## Azioni personalizzate
    /{id}/ottimizza/     Avvia l'ottimizzazione 3D (sincrona o asincrona) — limitato a 10/min
    /{id}/stato/         Stato del task di ottimizzazione
    """
    permission_classes = [IsAuthenticated]
    queryset = PianoDiCarico.objects.prefetch_related(
        "oggetti_posizionati__oggetto",
        "contenitore",
    ).annotate(
        num_oggetti=models.Count("oggetti_posizionati"),
    )

    def get_queryset(self):
        return super().get_queryset().filter(owner=self.request.user)

    def _verifica_contenitore_utente(self, contenitore):
        if contenitore.owner_id != self.request.user.id:
            raise ValidationError({"contenitore": "Contenitore non appartenente all'utente."})

    def perform_create(self, serializer):
        contenitore = serializer.validated_data["contenitore"]
        self._verifica_contenitore_utente(contenitore)
        serializer.save(owner=self.request.user)

    def perform_update(self, serializer):
        contenitore = serializer.validated_data.get(
            "contenitore", serializer.instance.contenitore
        )
        self._verifica_contenitore_utente(contenitore)
        serializer.save(owner=self.request.user)

    def get_serializer_class(self):
        if self.action == "list":
            return PianoDiCaricoListSerializer
        if self.action == "create":
            return PianoDiCaricoCreateSerializer
        return PianoDiCaricoDetailSerializer

    # ---- Azioni personalizzate ----

    @action(detail=True, methods=["post"], throttle_classes=[OttimizzaRateThrottle])
    def ottimizza(self, request, pk=None):
        """Avvia l'ottimizzazione 3D per un piano di carico.

        Corpo richiesta (JSON):
            {
                "asincrono": true,           // default true
                "config": { ... }             // configurazione opzionale
            }

        Il campo `config` segue il formato delle impostazioni dell'ottimizzatore
        (Strategia, Performance, Output). Se omesso, vengono usati i default.
        """
        piano = self.get_object()

        # Legge i parametri dal body
        data = request.data if request.data else {}
        logger.info("OTTIMIZZA richiesta — piano_id=%s — request.data=%s", pk, data)
        serializer = AvviaOttimizzazioneSerializer(data=data)
        if not serializer.is_valid():
            logger.warning("OTTIMIZZA validazione fallita — errori=%s", serializer.errors)
            raise ValidationError(serializer.errors)
        validated = serializer.validated_data
        asincrono = validated.get("asincrono")  # None = decisione automatica
        salva_risultato = validated["salva_risultato"]
        config = validated.get("config", None)  # dict o None

        # Stima del costo (ops = quantità × orientamenti) per decidere
        # sync/async: i carichi pesanti finiscono in coda per non sforare
        # il timeout del browser.
        ops = stima_ops_piano(piano)
        ms_per_op = float(getattr(settings, "OPTIMIZZATORE_MS_PER_OP", 20.0))
        soglia_sync_s = float(getattr(settings, "OPTIMIZZATORE_SOGLIA_SYNC_S", 15.0))
        tempo_stimato_s = (ops * ms_per_op) / 1000.0

        if asincrono is None:
            asincrono = bool(
                salva_risultato and tempo_stimato_s > soglia_sync_s
            )

        if asincrono and not salva_risultato:
            raise ValidationError(
                {"salva_risultato": "La preview deve essere eseguita in modo sincrono."}
            )

        if asincrono:
            # Task asincrono via Django Q2
            task_id = accoda_ottimizzazione(
                piano.id,
                config=config,
                budget_seconds=OPTIMIZATION_TIME_BUDGET_ASYNC_SECONDS,
            )

            piano.task_id = task_id
            piano.stato = StatoPiano.IN_ELABORAZIONE
            piano.save(update_fields=["task_id", "stato"])

            return Response(
                {
                    "successo": True,
                    "piano_id": piano.id,
                    "task_id": task_id,
                    "config_applicata": config is not None,
                    "asincrono": True,
                    "ops": ops,
                    "tempo_stimato_s": round(tempo_stimato_s, 2),
                    "messaggio": (
                        f"Ottimizzazione avviata in coda asincrona. "
                        f"Task ID: {task_id}"
                    ),
                },
                status=status.HTTP_202_ACCEPTED,
            )
        else:
            # Esecuzione sincrona (bloccante)
            risultato = esegui_ottimizzazione_sincrona(
                piano.id,
                config=config,
                salva_risultato=salva_risultato,
                budget_seconds=OPTIMIZATION_TIME_BUDGET_SECONDS,
            )

            # Ricarica il piano per dati aggiornati
            piano.refresh_from_db()

            # Costruisci risposta con metriche estese
            response_data = {
                "successo": risultato["successo"],
                "piano_id": piano.id,
                "piano_nome": piano.nome,
                "oggetti_posizionati": risultato["oggetti_posizionati"],
                "oggetti_non_posizionati": risultato["oggetti_non_posizionati"],
                "saturazione": round(risultato["saturazione_percentuale"], 1),
                "messaggio": risultato["messaggio"],
                "report_priorita": risultato.get("report_priorita", {}),
                "posizioni_preview": risultato.get("posizioni_preview", []),
                "salva_risultato": salva_risultato,
            }

            # Aggiungi metriche estese se configurate per output dettagliato
            if risultato.get("metriche"):
                response_data["metriche"] = risultato["metriche"]

            # Aggiungi eventuali soluzioni alternative generate dal Simulated Annealing
            if risultato.get("soluzioni_alternative"):
                response_data["soluzioni_alternative"] = risultato["soluzioni_alternative"]

            # Telemetria (ops, passate, tempi, motivo di stop)
            if risultato.get("telemetria"):
                response_data["telemetria"] = risultato["telemetria"]

            return Response(response_data, status=status.HTTP_200_OK)

    def _soluzioni_alternative_transitorie(self, piano):
        """Legge le alternative dal risultato del task asincrono.

        Le alternative NON sono più persistite sul piano (sono sempre
        temporanee): per il polling del frontend vengono lette dal risultato
        del task Django Q2 associato al piano (``piano.task_id``).
        """
        if not piano.task_id:
            return []
        try:
            from django_q.models import Task
            task = Task.objects.filter(id=piano.task_id).first()
            if task and task.success:
                risultato = task.result
                if isinstance(risultato, dict):
                    return risultato.get("soluzioni_alternative") or []
        except Exception:
            pass
        return []

    def _task_pronto(self, piano):
        """True se il task Q2 ha già scritto il suo risultato.

        Django Q scrive ``success`` (True/False) solo a task terminato: il
        worker marca il piano ``completato`` PRIMA di scrivere il result del
        task, dove stanno le soluzioni alternative. Il frontend usa questo
        flag per continuare il polling finché il risultato non è disponibile,
        senza tempi fissi: appena il task è pronto, le alternative ci sono o
        non ci saranno (strategia senza Monte Carlo).
        """
        if not piano.task_id:
            return True
        try:
            from django_q.models import Task
            task = Task.objects.filter(id=piano.task_id).first()
            if task is None:
                # Task non ancora scritto dal worker: risultato non disponibile.
                return False
            return task.success is not None
        except Exception:
            # Backend Q non interrogabile: non bloccare il polling.
            return True

    @action(detail=True, methods=["get"])
    def stato(self, request, pk=None):
        """Restituisce lo stato corrente del piano di carico.

        Utile per il polling del frontend dopo aver avviato
        un'ottimizzazione asincrona.
        """
        piano = self.get_object()
        num_posizionati = piano.oggetti_posizionati.count()
        stato = piano.stato
        messaggio_errore = piano.messaggio_errore

        # Django Q registra il fallimento del task, ma in caso di eccezione
        # fuori dall'orchestratore il modello del piano può essere rimasto in
        # ``in_elaborazione``. Intercetta quel caso durante il polling, così
        # il frontend riceve uno stato finale invece di attendere inutilmente.
        if stato == StatoPiano.IN_ELABORAZIONE and piano.task_id:
            try:
                from django_q.models import Task
                task = Task.objects.filter(id=piano.task_id).first()
                if task is not None and task.success is False:
                    logger.error(
                        "Task Q fallito durante il polling: piano_id=%s task_id=%s",
                        piano.id,
                        piano.task_id,
                    )
                    stato = StatoPiano.ERRORE
                    messaggio_errore = (
                        "Il worker non ha completato l'ottimizzazione. "
                        "Riprova l'operazione."
                    )
            except Exception:
                # Il polling dello stato non deve fallire se il backend Q non
                # è disponibile: continua a restituire lo stato del piano.
                pass

        return Response(
            {
                "id": piano.id,
                "nome": piano.nome,
                "stato": stato,
                "task_id": piano.task_id,
                "oggetti_posizionati": num_posizionati,
                "saturazione": piano.saturazione_volumetrica,
                "peso_totale_kg": piano.peso_totale_kg,
                "completato_at": piano.completato_at,
                "messaggio_errore": messaggio_errore,
                "task_pronto": self._task_pronto(piano),
                "soluzioni_alternative": self._soluzioni_alternative_transitorie(piano),
            }
        )

    @action(detail=True, methods=["post", "patch", "delete"],
             throttle_classes=[PanelRateThrottle])
    def oggetti_da_caricare(self, request, pk=None):
        """Gestisce gli oggetti da caricare per questo piano.

        POST   /api/piani/{id}/oggetti_da_caricare/  → Crea una nuova riga
        PATCH  /api/piani/{id}/oggetti_da_caricare/  → Aggiorna una riga esistente
        DELETE /api/piani/{id}/oggetti_da_caricare/  → Rimuove TUTTI gli oggetti dal piano

        Corpo POST (JSON):
            { "oggetto_id": 1, "quantita": 5, "priorita": 0, "colore": "#ff0000", "note": "" }

        Corpo PATCH (JSON, campi opzionali):
            { "riga_id": 12, "quantita": 5, "priorita": 2, "colore": "" }
        """
        piano = self.get_object()

        if request.method == "DELETE":
            count = piano.oggetti_da_caricare.count()
            piano.oggetti_da_caricare.all().delete()
            return Response({"successo": True, "rimossi": count})

        if request.method == "PATCH":
            riga_id = request.data.get("riga_id")
            if not riga_id:
                raise ValidationError({"riga_id": "Campo obbligatorio per il PATCH."})
            try:
                riga = piano.oggetti_da_caricare.get(pk=int(riga_id))
            except (TypeError, ValueError, OggettoDaCaricare.DoesNotExist):
                raise ValidationError({"riga_id": "Riga non trovata in questo piano."})

            serializer = OggettoDaCaricareUpdateSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)

            for campo, valore in serializer.validated_data.items():
                setattr(riga, campo, valore)
            riga.save()

            return Response(
                {
                    "id": riga.id,
                    "oggetto_id": riga.oggetto_id,
                    "quantita": riga.quantita,
                    "priorita": riga.priorita,
                    "colore": riga.colore or "",
                    "note": riga.note,
                    "created": False,
                }
            )

        # POST
        serializer = OggettoDaCaricareSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        oggetto_id = serializer.validated_data["oggetto_id"]
        if not Oggetto.objects.filter(id=oggetto_id, owner=request.user).exists():
            raise ValidationError({"oggetto_id": "Oggetto non appartenente all'utente."})

        # Ogni POST rappresenta un lotto distinto. Non usare update_or_create:
        # lo stesso oggetto può comparire su più righe con priorità diverse.
        oggetto_da_caricare = OggettoDaCaricare.objects.create(
            piano_di_carico=piano,
            oggetto_id=oggetto_id,
            quantita=serializer.validated_data.get("quantita", 1),
            priorita=serializer.validated_data.get("priorita", 0),
            note=serializer.validated_data.get("note", ""),
            colore=serializer.validated_data.get("colore", ""),
        )

        return Response(
            {
                "id": oggetto_da_caricare.id,
                "oggetto_id": oggetto_da_caricare.oggetto_id,
                "quantita": oggetto_da_caricare.quantita,
                "priorita": oggetto_da_caricare.priorita,
                "colore": oggetto_da_caricare.colore or "",
                "note": oggetto_da_caricare.note,
                "created": True,
            },
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["get"])
    def export_posizioni(self, request, pk=None):
        """Esporta le posizioni degli oggetti in formato testo leggibile.

        Restituisce un file .txt con tutte le coordinate, dimensioni,
        rotazioni e codici degli oggetti posizionati nel piano.
        """
        from django.http import HttpResponse

        piano = self.get_object()
        contenitore = piano.contenitore
        oggetti_posizionati = list(
            piano.oggetti_posizionati.select_related("oggetto", "riga_origine").order_by(
                "coordinata_x_mm", "coordinata_y_mm", "coordinata_z_mm"
            )
        )

        lines = []
        lines.append("=" * 85)
        lines.append(f"PIANO DI CARICO: {piano.nome}")
        lines.append(f"Contenitore: {contenitore.nome} ({contenitore.tipo_mezzo})")
        lines.append(f"Dimensioni contenitore (mm): X={contenitore.lunghezza_mm} Y={contenitore.larghezza_mm} Z={contenitore.altezza_mm}")
        lines.append(f"Stato: {piano.get_stato_display()}")
        lines.append(f"Algoritmo: {piano.algoritmo or 'N/D'}")
        lines.append(f"Oggetti posizionati: {len(oggetti_posizionati)}")
        lines.append(f"Saturazione: {piano.saturazione_volumetrica:.1f}%" if piano.saturazione_volumetrica else "Saturazione: N/D")
        lines.append(f"Peso totale: {piano.peso_totale_kg} kg" if piano.peso_totale_kg else "Peso totale: N/D")
        lines.append("=" * 85)
        lines.append("")
        lines.append(f"{'#':>3} {'Codice':<12} {'X':>6} {'Y':>6} {'Z':>6} {'dX':>5} {'dY':>5} {'dZ':>5} {'Rot':<15} {'Colore'}")
        lines.append("-" * 85)

        for i, op in enumerate(oggetti_posizionati, 1):
            lines.append(
                f"{i:>3} {op.oggetto.codice:<12} "
                f"{op.coordinata_x_mm:>6} {op.coordinata_y_mm:>6} {op.coordinata_z_mm:>6} "
                f"{op.dimensione_x_mm:>5} {op.dimensione_y_mm:>5} {op.dimensione_z_mm:>5} "
                f"{op.rotazione_applicata or 'nessuna':<15} "
                f"{_colore_posizionamento(op) or '-'}"
            )

        lines.append("-" * 85)
        lines.append("")

        # Riepilogo per codice + colore (posizione)
        from collections import Counter
        codici = Counter(
            (op.oggetto.codice, _colore_posizionamento(op) or '-')
            for op in oggetti_posizionati
        )
        lines.append("RIEPILOGO PER CODICE:")
        for (codice, colore), qty in sorted(codici.items()):
            lines.append(f"  {codice}: {qty} pezzi (colore: {colore})")

        # Riepilogo occupazione X
        if oggetti_posizionati:
            max_x = max(op.coordinata_x_mm + op.dimensione_x_mm for op in oggetti_posizionati)
            lines.append(f"\nOccupazione X max: {max_x} mm = {max_x/10:.1f} cm")

        content = "\n".join(lines)

        response = HttpResponse(content, content_type="text/plain; charset=utf-8")
        response["Content-Disposition"] = (
            f'attachment; filename="piano_{piano.id}_{piano.nome.replace(" ", "_")}.txt"'
        )
        return response

    @action(detail=True, methods=["get"])
    def dati_3d(self, request, pk=None):
        """Restituisce i dati pronti per il rendering Three.js.

        Endpoint ottimizzato per il frontend: sfrutta la
        prefetch_related già presente nella queryset del viewset
        per evitare query N+1.
        """
        piano = self.get_object()
        contenitore = piano.contenitore

        # Usa la prefetch cache (evita select_related extra)
        # IMPORTANTE: stesso ordinamento di export_posizioni per avere
        # numeri corrispondenti tra tabella esportata e scene 3D.
        oggetti_posizionati = list(
            piano.oggetti_posizionati.select_related("riga_origine").order_by(
                "coordinata_x_mm", "coordinata_y_mm", "coordinata_z_mm"
            )
        )

        # Includi anche gli oggetti_da_caricare (q.tà salvate dall'utente)
        oggetti_da_caricare = list(
            piano.oggetti_da_caricare.select_related("oggetto").all()
        )

        payload = {
            "piano": {
                "id": piano.id,
                "nome": piano.nome,
                "stato": piano.stato,
                "algoritmo": piano.algoritmo or "",
            },
            "contenitore": {
                "nome": contenitore.nome,
                "dimensioni_mm": {
                    "x": contenitore.lunghezza_mm,
                    "y": contenitore.larghezza_mm,
                    "z": contenitore.altezza_mm,
                },
                "dimensioni_cm": {
                    "x": round(contenitore.lunghezza_mm / 10, 1),
                    "y": round(contenitore.larghezza_mm / 10, 1),
                    "z": round(contenitore.altezza_mm / 10, 1),
                },
            },
            "oggetti": [],
            "metriche": {
                "peso_totale_kg": (
                    float(piano.peso_totale_kg)
                    if piano.peso_totale_kg else 0
                ),
                "saturazione": (
                    round(piano.saturazione_volumetrica, 1)
                    if piano.saturazione_volumetrica else 0
                ),
                "oggetti_posizionati": len(oggetti_posizionati),
            },
            "oggetti_da_caricare": [
                {
                    "id": odc.id,
                    "oggetto_id": odc.oggetto_id,
                    "codice": odc.oggetto.codice,
                    "quantita": odc.quantita,
                    "priorita": odc.priorita,
                    "colore": odc.colore or "",
                    "note": odc.note,
                }
                for odc in oggetti_da_caricare
            ],
        }

        for op in oggetti_posizionati:
            payload["oggetti"].append(
                {
                    "id": op.id,
                    "oggetto_id": op.oggetto_id,
                    "riga_id": op.riga_origine_id,
                    "codice": op.oggetto.codice,
                    "descrizione": op.oggetto.descrizione,
                    "posizione_mm": {
                        "x": op.coordinata_x_mm,
                        "y": op.coordinata_y_mm,
                        "z": op.coordinata_z_mm,
                    },
                    "dimensioni_mm": {
                        "x": op.dimensione_x_mm,
                        "y": op.dimensione_y_mm,
                        "z": op.dimensione_z_mm,
                    },
                    "posizione_cm": {
                        "x": round(op.coordinata_x_mm / 10, 1),
                        "y": round(op.coordinata_y_mm / 10, 1),
                        "z": round(op.coordinata_z_mm / 10, 1),
                    },
                    "dimensioni_cm": {
                        "x": round(op.dimensione_x_mm / 10, 1),
                        "y": round(op.dimensione_y_mm / 10, 1),
                        "z": round(op.dimensione_z_mm / 10, 1),
                    },
                    "rotazione": op.rotazione_applicata,
                    "colore": _colore_posizionamento(op),
                    "peso_kg": float(op.oggetto.peso_kg),
                    "peso_sopra_kg": float(op.peso_posato_sopra_kg),
                }
            )

        return Response(payload)

    @action(detail=True, methods=["get"])
    def distribuzione_pesi(self, request, pk=None):
        """Restituisce la distribuzione del carico sulle sezioni.

        Calcola per ogni sezione del contenitore quanto carico (in kg)
        ricade su di essa, basandosi sugli oggetti posizionati.
        Utile per il grafico della toolbar dopo aver caricato un piano salvato.
        Restituisce anche gli oggetti posizionati per disegnare il profilo
        di peso lungo l'asse X del veicolo.
        """
        piano = self.get_object()
        dati = _calcola_distribuzione_pesi(piano)
        return Response({
            "distribuzione_pesi": dati["sezioni"],
            "oggetti": dati["oggetti"],
        })

    @action(detail=True, methods=["post"])
    def salva_posizioni_manuali(self, request, pk=None):
        """Salva o sincronizza le posizioni 3D del piano.

        POST /api/piani/{id}/salva_posizioni_manuali/

        Il campo ``origine`` distingue i due flussi:
        - ``manuale``: l'utente ha modificato direttamente la scena 3D e il
          piano viene marcato come manuale;
        - ``sincronizzazione``: il frontend aggiorna le posizioni per allineare
          la lista di carico, senza cambiare l'origine dell'elaborazione.

        Per compatibilità con i client precedenti, l'assenza di ``origine``
        mantiene il comportamento manuale.

        Corpo richiesta (JSON):
            {
                "oggetti": [
                    {
                        "codice": "CART-102",
                        "posizione_cm": {"x": 0, "y": 100, "z": 200},
                        "dimensioni_cm": {"x": 40, "y": 30, "z": 25},
                        "colore": "#447e9b",
                        "rotazione": "XYZ"
                    }
                ]
            }
        """
        piano = self.get_object()
        oggetti_data = request.data.get("oggetti", [])
        origine = request.data.get("origine", "manuale")
        if origine not in {"manuale", "sincronizzazione"}:
            return Response(
                {"errore": "Origine non valida. Usare 'manuale' o 'sincronizzazione'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not oggetti_data:
            return Response(
                {"errore": "Nessun oggetto fornito."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Il salvataggio rappresenta la scena visibile, non una nuova
        # esecuzione dell'ottimizzatore: la lista ``oggetti_da_caricare`` e i
        # relativi vincoli/quantità non devono quindi rifiutare il payload.
        # L'unico controllo d'appartenenza necessario è la proprietà degli
        # oggetti, per impedire di salvare record appartenenti a un altro utente.
        # La preview di "Elabora" conserva l'ID dell'anagrafica. Usalo come
        # chiave primaria stabile dopo l'eliminazione del piano tecnico; il
        # codice resta un fallback per i client precedenti che non lo inviano.
        codici = [str(item.get("codice", "")).strip() for item in oggetti_data]
        oggetto_ids = [item.get("oggetto_id") for item in oggetti_data if item.get("oggetto_id")]
        oggetti_qs = Oggetto.objects.filter(owner=request.user).filter(
            models.Q(id__in=oggetto_ids) | models.Q(codice__in=codici)
        )
        oggetti_per_id = {str(o.id): o for o in oggetti_qs}
        oggetti_per_codice = {o.codice: o for o in oggetti_qs}

        nuovi = []
        # Traccia quali righe sono già state assegnate come riga_origine
        # per evitare che il fallback assegni la stessa riga a più
        # posizionamenti dello stesso oggetto (causando colori identici).
        righe_assegnate = set()
        for item in oggetti_data:
            codice = str(item.get("codice", "")).strip()
            oggetto_id = item.get("oggetto_id")
            oggetto = oggetti_per_id.get(str(oggetto_id)) if oggetto_id else None
            if oggetto is None:
                oggetto = oggetti_per_codice.get(codice)
            if not oggetto:
                raise ValidationError({
                    "oggetti": "Uno o più oggetti non appartengono all'utente."
                })

            pc = item.get("posizione_cm", {})
            dc = item.get("dimensioni_cm", {})
            try:
                x, y, z = (float(pc[key]) for key in ("x", "y", "z"))
                dx, dy, dz = (float(dc[key]) for key in ("x", "y", "z"))
            except (KeyError, TypeError, ValueError):
                raise ValidationError({
                    "oggetti": "Posizione e dimensioni devono contenere valori numerici completi."
                })
            if (
                not all(math.isfinite(value) for value in (x, y, z, dx, dy, dz))
                or min(x, y, z) < 0
                or min(dx, dy, dz) <= 0
            ):
                raise ValidationError({
                    "oggetti": "Coordinate non negative e dimensioni positive richieste."
                })

            x_mm, y_mm, z_mm = round(x * 10), round(y * 10), round(z * 10)
            dx_mm, dy_mm, dz_mm = round(dx * 10), round(dy * 10), round(dz * 10)
            if min(dx_mm, dy_mm, dz_mm) < 1:
                raise ValidationError({
                    "oggetti": "Le dimensioni devono essere almeno 1 mm."
                })
            if (
                x_mm + dx_mm > piano.contenitore.lunghezza_mm
                or y_mm + dy_mm > piano.contenitore.larghezza_mm
                or z_mm + dz_mm > piano.contenitore.altezza_mm
            ):
                raise ValidationError({
                    "oggetti": "Un posizionamento supera le dimensioni del contenitore."
                })

            riga_id = item.get("riga_id")
            riga_origine = None
            if riga_id not in (None, ""):
                try:
                    riga_origine = piano.oggetti_da_caricare.get(pk=int(riga_id))
                except (TypeError, ValueError, OggettoDaCaricare.DoesNotExist):
                    # Preview e client legacy possono inviare un id non più
                    # valido; prova a trovare la riga corretta per oggetto_id.
                    riga_origine = None
            # Fallback: se la riga non è stata trovata per ID, prova a
            # trovare una riga con lo stesso oggetto_id che non sia già
            # stata assegnata a un altro posizionamento.
            if riga_origine is None and riga_id not in (None, ""):
                try:
                    riga_origine = (
                        piano.oggetti_da_caricare
                        .filter(oggetto=oggetto)
                        .exclude(pk__in=righe_assegnate)
                        .order_by('id')
                        .first()
                    )
                except Exception:
                    riga_origine = None
            if riga_origine is not None and riga_origine.oggetto_id != oggetto.id:
                raise ValidationError({
                    "oggetti": "La riga origine non corrisponde all'oggetto posizionato."
                })

            # Traccia la riga assegnata per evitare duplicati nel fallback.
            if riga_origine is not None:
                righe_assegnate.add(riga_origine.pk)

            # Il colore della riga ha la precedenza sul colore inviato dal
            # frontend: quando la riga è nota, usa il suo colore per-riga;
            # altrimenti usa il colore dello snapshot (preview o legacy).
            colore_finale = item.get("colore", "#4488ff")
            if riga_origine is not None and riga_origine.colore:
                colore_finale = riga_origine.colore

            nuovi.append(OggettoPosizionato(
                piano_di_carico=piano,
                oggetto=oggetto,
                riga_origine=riga_origine,
                coordinata_x_mm=x_mm,
                coordinata_y_mm=y_mm,
                coordinata_z_mm=z_mm,
                dimensione_x_mm=dx_mm,
                dimensione_y_mm=dy_mm,
                dimensione_z_mm=dz_mm,
                colore=colore_finale,
                rotazione_applicata=item.get("rotazione", "XYZ"),
            ))

        # Le priorità hanno la precedenza su tutto nei flussi AUTOMATICI:
        # la barriera di fase deve valere per i payload generati dal motore
        # (Elabora, soluzioni alternative, client legacy): un oggetto a
        # priorità 0 che risulterebbe posizionato prima della fine X dei
        # prioritari viene ripiazzato in coda, altrimenti non sarebbe
        # scaricabile senza smontare il carico.
        #
        # Nel flusso MANUALE invece l'utente ha la precedenza: se ha spostato
        # un oggetto a mano nella scena 3D, la sua scelta va salvata così
        # com'è (o al massimo segnalata, mai rifiutata). Applicare qui la
        # barriera bloccherebbe il salvataggio con un 400 quando un prio0
        # spostato a mano non trova posto in coda.
        if origine == "sincronizzazione":
            _applica_barriera_priorita(nuovi, piano)

        # Cancella solo dopo aver validato tutto il payload. La transazione
        # mantiene comunque l'operazione atomica anche in caso di errore DB.
        with transaction.atomic():
            piano.oggetti_posizionati.all().delete()
            OggettoPosizionato.objects.bulk_create(nuovi)

            # Calcola metriche (peso, volume, saturazione) dalle posizioni salvate
            peso_totale = sum(
                float(op.oggetto.peso_kg or 0) for op in nuovi
            )
            volume_totale = sum(
                op.dimensione_x_mm * op.dimensione_y_mm * op.dimensione_z_mm
                for op in nuovi
            )
            dims = (
                piano.contenitore.lunghezza_mm,
                piano.contenitore.larghezza_mm,
                piano.contenitore.altezza_mm,
            )
            volume_container = dims[0] * dims[1] * dims[2]
            saturazione = (
                (volume_totale / volume_container * 100) if volume_container > 0 else 0
            )

            # La sincronizzazione non deve alterare lo stato di un piano
            # automatico parziale; una modifica manuale, invece, completa il
            # piano perché le posizioni sono state confermate dall'utente.
            update_fields = [
                "updated_at",
                "peso_totale_kg",
                "volume_utilizzato_mm3",
                "completato_at",
            ]
            piano.peso_totale_kg = Decimal(str(round(peso_totale, 2)))
            piano.volume_utilizzato_mm3 = volume_totale
            piano.completato_at = timezone.now()
            if origine == "manuale":
                piano.stato = StatoPiano.COMPLETATO
                piano.algoritmo = "manuale"
                update_fields.extend(["stato", "algoritmo"])
            elif origine == "sincronizzazione":
                # Applicare una soluzione alternativa (o salvare un piano
                # automatico) riporta il piano ad "automatico": rimuove
                # l'etichetta "manuale" che un salvataggio manuale precedente
                # aveva impostato.
                algoritmo = request.data.get("algoritmo")
                if algoritmo:
                    piano.algoritmo = str(algoritmo)[:128]
                elif not piano.algoritmo:
                    piano.algoritmo = "Algoritmo 3D Semplificato"
                update_fields.append("algoritmo")
                # Una sincronizzazione di una soluzione automatica completa il
                # piano, ma non deve trasformare un risultato già dichiarato
                # parziale in completato: l'utente deve continuare a vedere
                # che alcune istanze non erano state posizionate.
                if piano.stato not in (StatoPiano.COMPLETATO, StatoPiano.PARZIALE):
                    piano.stato = StatoPiano.COMPLETATO
                    update_fields.append("stato")
            piano.save(update_fields=update_fields)

        return Response({
            "successo": True,
            "oggetti_salvati": len(nuovi),
        })


# ===========================================================================
# ViewSet: VincoloTraOggetti
# ===========================================================================

class VincoloTraOggettiViewSet(viewsets.ModelViewSet):
    """API endpoint per la gestione dei vincoli tra coppie di oggetti.
    Richiede autenticazione."""
    permission_classes = [IsAuthenticated]
    queryset = VincoloTraOggetti.objects.select_related("oggetto_a", "oggetto_b").all()
    serializer_class = VincoloTraOggettiSerializer

    def get_queryset(self):
        return super().get_queryset().filter(
            oggetto_a__owner=self.request.user,
            oggetto_b__owner=self.request.user,
        )

    def perform_create(self, serializer):
        oggetto_a = serializer.validated_data["oggetto_a"]
        oggetto_b = serializer.validated_data["oggetto_b"]
        if oggetto_a.owner_id != self.request.user.id or oggetto_b.owner_id != self.request.user.id:
            raise ValidationError({"oggetto_a": "Gli oggetti devono appartenere all'utente."})
        serializer.save()

    def perform_update(self, serializer):
        oggetto_a = serializer.validated_data.get(
            "oggetto_a", serializer.instance.oggetto_a
        )
        oggetto_b = serializer.validated_data.get(
            "oggetto_b", serializer.instance.oggetto_b
        )
        if (
            oggetto_a.owner_id != self.request.user.id
            or oggetto_b.owner_id != self.request.user.id
        ):
            raise ValidationError({"oggetto_a": "Gli oggetti devono appartenere all'utente."})
        serializer.save()


# ===========================================================================
# API: Gestione Icone (Admin only)
# ===========================================================================

# Percorsi sovrascrivibili via env per il deploy Docker:
# - ICON_CONFIG_PATH → volume persistente per la config icone
# - ICON_UPLOAD_DIR   → volume condiviso per i PNG caricati dagli admin
ICON_CONFIG_PATH = os.environ.get(
    "ICON_CONFIG_PATH",
    os.path.join(settings.BASE_DIR, "icon_config.json"),
)
ICON_UPLOAD_DIR = os.environ.get(
    "ICON_UPLOAD_DIR",
    os.path.join(settings.BASE_DIR, "caricamento", "static", "caricamento", "img"),
)

# Firma binaria di un PNG valido (8 byte iniziali). Usata per verificare il
# contenuto reale degli upload, oltre alla sola estensione del nome file.
PNG_SIGNATURE = b"\x89PNG\r\n\x1a\n"


def _load_icon_config():
    """Carica la configurazione icone dal file JSON."""
    if os.path.exists(ICON_CONFIG_PATH):
        try:
            with open(ICON_CONFIG_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {"config": {}}


def _save_icon_config(data):
    """Salva la configurazione icone nel file JSON."""
    os.makedirs(os.path.dirname(ICON_CONFIG_PATH), exist_ok=True)
    with open(ICON_CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# Chiavi consentite per la sezione "colori" del config icone (colori aree/slider
# + "base": colore di riferimento per la generazione delle tonalità coordinate).
COLORI_CATALOG = {
    "base", "header", "sidebar", "panel-bg", "panel-border", "main-bg",
    "accent", "accent-hover",
    "slider-track-start", "slider-track", "slider-thumb", "strategia-thumb",
}

# Chiavi consentite per la sezione "header" del config icone (altezze in px).
HEADERI_CATALOG = {
    "sidebar-tabs", "panel-destro",
    "pv-list-header", "pv-list-articoli", "pv-list-trasporti",
    "pv-list-impostazioni",
    "pv-form-header", "pv-form-articoli", "pv-form-vincoli",
    "pv-form-trasporti", "pv-form-impostazioni",
}


def _normalizza_nome_file_icona(nome):
    """Restituisce il nome PNG nello stesso formato usato dall'upload."""
    file = os.path.basename(unquote(str(nome or "")).strip())
    return file.lower().replace(" ", "_").replace("..", "")


def _normalizza_file_config(data):
    """Normalizza i nomi PNG nelle sezioni di configurazione ricevute."""
    for section in ("config", "bottoni"):
        section_data = data.get(section)
        if not isinstance(section_data, dict):
            continue
        for entry in section_data.values():
            if isinstance(entry, dict) and entry.get("file"):
                entry["file"] = _normalizza_nome_file_icona(entry["file"])


def _verifica_file_icone_esistono(data):
    """Raccoglie i riferimenti PNG mancanti nella cartella img."""
    mancanti = []
    riferimenti = []
    for section in ("config", "bottoni"):
        section_data = data.get(section)
        if not isinstance(section_data, dict):
            continue
        for entry in section_data.values():
            if isinstance(entry, dict) and entry.get("file"):
                riferimenti.append(_normalizza_nome_file_icona(entry["file"]))
    for fn in sorted(set(riferimenti)):
        if not os.path.exists(os.path.join(ICON_UPLOAD_DIR, fn)):
            mancanti.append(fn)
    return mancanti


def _json_api_error(request, message, status_code, code="api_error", **extra):
    """Risposta errore coerente per gli endpoint legacy non-DRF."""
    request_id = getattr(request, "request_id", "-")
    payload = {
        "success": False,
        "error": {"code": code, "message": message, "request_id": request_id},
        # Mantieni detail per i client legacy che leggono il formato DRF.
        "detail": message,
    }
    payload.update(extra)
    return JsonResponse(payload, status=status_code)


@login_required
def api_icone_config(request):
    """
    GET  /api/icone-config/  → Restituisce la configurazione icone attuale
    POST /api/icone-config/  → Salva configurazione icone e bottoni (solo admin).

    Valida che tutti i file PNG referenziati (icone e bottoni) esistano
    nella cartella img prima di salvare.
    """
    if request.method == "GET":
        config = _load_icon_config()
        return JsonResponse(config)

    if request.method == "POST":
        if not request.user.is_staff:
            return _json_api_error(request, "Solo amministratori.", 403, "permission_denied")
        try:
            data = json.loads(request.body.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError):
            return _json_api_error(request, "JSON non valido.", 400, "invalid_json")
        if not isinstance(data, dict):
            return _json_api_error(request, "Il payload deve essere un oggetto JSON.", 400, "invalid_payload")

        if "config" not in data and "bottoni" not in data and "colori" not in data and "header" not in data:
            return _json_api_error(
                request,
                "Campi 'config', 'bottoni', 'colori' o 'header' richiesti.",
                400,
                "validation_error",
            )

        # Merge difensivo: un payload parziale non deve mai perdere la
        # sezione non inviata della configurazione esistente.
        existing = _load_icon_config()
        if "config" not in data:
            data["config"] = existing.get("config", {})
        if "bottoni" not in data:
            data["bottoni"] = existing.get("bottoni", {})
        if "colori" not in data:
            data["colori"] = existing.get("colori", {})
        if "header" not in data:
            data["header"] = existing.get("header", {})

        # L'upload sostituisce gli spazi con underscore: applica la stessa
        # normalizzazione anche ai nomi digitati nella tabella, evitando 400
        # e URL 404 per file come "icons8-trash-96 (1).png".
        _normalizza_file_config(data)

        # Sanitizzazione del colore dei bottoni: solo #rrggbb, maiuscole
        # normalizzate in minuscole. Un valore non valido viene scartato
        # (mai un payload di stile arbitrario nel file di configurazione).
        hex_color_re = re.compile(r"^#[0-9a-fA-F]{6}$")
        bottoni = data.get("bottoni")
        if isinstance(bottoni, dict):
            for entry in bottoni.values():
                if not isinstance(entry, dict):
                    continue
                if entry.get("color"):
                    color = str(entry["color"]).strip()
                    if hex_color_re.match(color):
                        entry["color"] = color.lower()
                    else:
                        entry.pop("color", None)
                # Dimensioni bottone: larghezza % (1-100) e altezza px (20-120).
                # Il minimo 1% permette bottoni molto stretti, se necessario.
                if "width_pct" in entry:
                    try:
                        entry["width_pct"] = max(1, min(100, int(entry["width_pct"])))
                    except (TypeError, ValueError):
                        entry.pop("width_pct", None)
                if "height_px" in entry:
                    try:
                        entry["height_px"] = max(20, min(120, int(entry["height_px"])))
                    except (TypeError, ValueError):
                        entry.pop("height_px", None)

        # Sanitizzazione dei colori: solo le chiavi del catalogo e solo valori
        # #rrggbb validi (minuscoli). Qualsiasi altra chiave/valore viene scartato.
        colori = data.get("colori")
        if isinstance(colori, dict):
            colori_validi = {}
            for key, value in colori.items():
                if key not in COLORI_CATALOG:
                    continue
                if isinstance(value, str) and hex_color_re.match(value.strip()):
                    colori_validi[key] = value.strip().lower()
            data["colori"] = colori_validi

        # Sanitizzazione altezze header: solo chiavi del catalogo, interi
        # clampati nell'intervallo consentito (20-120 px).
        header = data.get("header")
        if isinstance(header, dict):
            header_validi = {}
            for key, value in header.items():
                if key not in HEADERI_CATALOG:
                    continue
                try:
                    header_validi[key] = max(20, min(120, int(value)))
                except (TypeError, ValueError):
                    continue
            data["header"] = header_validi

        # Validazione: i file PNG referenziati devono esistere nella cartella img
        mancanti = _verifica_file_icone_esistono(data)
        if mancanti:
            return _json_api_error(
                request,
                "File PNG non trovati: " + ", ".join(mancanti),
                400,
                "missing_files",
                file_mancanti=mancanti,
            )

        _save_icon_config(data)
        return JsonResponse({"success": True, "message": "Configurazione salvata."})

    return _json_api_error(request, "Metodo non consentito.", 405, "method_not_allowed")


def api_privacy_settings(request):
    """GET/POST dei dati Privacy/Titolare usati dalle pagine legali.

    GET  /api/privacy-settings/  → dati correnti (lettura libera: sono gli
                                   stessi valori mostrati dalle pagine legali)
    POST /api/privacy-settings/  → salva (solo admin)

    I valori vivono nel singleton ``ImpostazioniSistema`` e vengono
    mostrati dalle pagine pubbliche privacy/cookie-policy/termini/rimborsi.
    """
    imp = ImpostazioniSistema.get()

    if request.method == "GET":
        return JsonResponse({
            "privacy": {
                "titolare": imp.privacy_titolare,
                "email": imp.privacy_email,
                "sede": imp.privacy_sede,
                "piva": imp.privacy_piva,
                "sito_url": imp.privacy_sito_url,
            }
        })

    if request.method == "POST":
        if not request.user.is_staff:
            return _json_api_error(request, "Solo amministratori.", 403, "permission_denied")
        try:
            data = json.loads(request.body.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError):
            return _json_api_error(request, "JSON non valido.", 400, "invalid_json")
        payload = (data or {}).get("privacy") or {}
        if not isinstance(payload, dict):
            return _json_api_error(
                request, "Campo 'privacy' deve essere un oggetto JSON.", 400, "invalid_payload"
            )

        titolare = str(payload.get("titolare", "")).strip()
        email = str(payload.get("email", "")).strip()
        sede = str(payload.get("sede", "")).strip()
        piva = str(payload.get("piva", "")).strip()
        sito_url = str(payload.get("sito_url", "")).strip()

        errors = {}
        if not titolare:
            errors["titolare"] = "Obbligatorio."
        if not email:
            errors["email"] = "Obbligatoria."
        else:
            from django.core.validators import validate_email
            from django.core.exceptions import ValidationError as DjangoValidationError
            try:
                validate_email(email)
            except DjangoValidationError:
                errors["email"] = "Email non valida."
        if sito_url and not (sito_url.startswith("http://") or sito_url.startswith("https://")):
            errors["sito_url"] = "Deve iniziare con http:// o https://."
        if errors:
            return _json_api_error(
                request, "Campi non validi.", 400, "validation_error", errors=errors
            )

        imp.privacy_titolare = titolare
        imp.privacy_email = email
        imp.privacy_sede = sede
        imp.privacy_piva = piva
        imp.privacy_sito_url = sito_url
        imp.save(update_fields=[
            "privacy_titolare", "privacy_email", "privacy_sede",
            "privacy_piva", "privacy_sito_url",
        ])
        return JsonResponse({"success": True, "message": "Dati Privacy salvati."})

    return _json_api_error(request, "Metodo non consentito.", 405, "method_not_allowed")


@login_required
def api_icone_file_delete(request):
    """
    DELETE /api/icone-file/  → Elimina un PNG caricato (solo admin).

    Corpo (JSON): {"filename": "icona.png"}
    Rifiuta con 409 se il file è ancora referenziato dalla configurazione
    attuale (icone o bottoni).
    """
    if request.method != "DELETE":
        return _json_api_error(request, "Metodo non consentito.", 405, "method_not_allowed")

    if not request.user.is_staff:
        return _json_api_error(request, "Solo amministratori.", 403, "permission_denied")

    try:
        data = json.loads(request.body.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return _json_api_error(request, "JSON non valido.", 400, "invalid_json")
    if not isinstance(data, dict):
        return _json_api_error(request, "Il payload deve essere un oggetto JSON.", 400, "invalid_payload")

    filename = (data.get("filename") or "").strip()
    if not filename.lower().endswith(".png"):
        return _json_api_error(request, "Solo file PNG.", 400, "invalid_file_type")

    safe_name = _normalizza_nome_file_icona(filename)
    path = os.path.join(ICON_UPLOAD_DIR, safe_name)

    if not os.path.exists(path):
        return _json_api_error(request, "File non trovato.", 404, "not_found")

    # Controlla che il file non sia referenziato nella configurazione attuale
    cfg = _load_icon_config()
    referenziati = set()
    for entry in (cfg.get("config") or {}).values():
        if isinstance(entry, dict) and entry.get("file"):
            referenziati.add(_normalizza_nome_file_icona(entry["file"]))
    for entry in (cfg.get("bottoni") or {}).values():
        if isinstance(entry, dict) and entry.get("file"):
            referenziati.add(_normalizza_nome_file_icona(entry["file"]))

    if safe_name in referenziati:
        return _json_api_error(
            request,
            (
                f"'{safe_name}' è ancora usato nella configurazione. "
                "Rimuovi il riferimento prima di eliminarlo."
            ),
            409,
            "conflict",
        )

    os.remove(path)
    logger.info("Icon PNG deleted: %s by %s", safe_name, request.user.username)
    return JsonResponse({"success": True, "message": f"'{safe_name}' eliminato."})


@login_required
def api_icone_upload(request):
    """
    POST /api/icone-upload/  → Carica un file immagine nella cartella img (solo admin).
    Accetta PNG, JPG, JPEG, WebP, GIF, BMP e li converte automaticamente in PNG.

    Body: multipart/form-data con campo 'file'.
    """
    if request.method != "POST":
        return _json_api_error(request, "Metodo non consentito.", 405, "method_not_allowed")

    if not request.user.is_staff:
        return _json_api_error(request, "Solo amministratori.", 403, "permission_denied")

    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return _json_api_error(request, "Nessun file fornito.", 400, "missing_file")

    # Verifica estensione: accetta formati comuni
    ALLOWED_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.webp', '.gif', '.bmp', '.ico', '.tiff'}
    filename = uploaded_file.name.lower()
    ext = os.path.splitext(filename)[1]
    if ext not in ALLOWED_EXTENSIONS:
        return _json_api_error(
            request,
            f"Formato non supportato: {ext}. Accettati: PNG, JPG, WebP, GIF, BMP.",
            400, "invalid_file_type"
        )

    # Per i PNG controlla subito la firma binaria, così un file rinominato
    # (es. un .exe con estensione .png) restituisce un errore chiaro prima
    # del tentativo di conversione Pillow.
    if ext == ".png":
        try:
            signature = uploaded_file.read(len(PNG_SIGNATURE))
            uploaded_file.seek(0)
        except (AttributeError, IOError):
            signature = b""
        if signature != PNG_SIGNATURE:
            return _json_api_error(
                request,
                "Il file PNG non contiene una firma PNG valida.",
                400,
                "invalid_file_type",
            )

    # Limite dimensione: 2 MB (prima della conversione)
    if uploaded_file.size > 2 * 1024 * 1024:
        return _json_api_error(request, "File troppo grande. Max 2 MB.", 400, "file_too_large")

    # Converti in PNG usando Pillow se non è già PNG
    try:
        img = Image.open(uploaded_file)
        # Converti in RGBA per supportare trasparenza
        if img.mode != 'RGBA':
            img = img.convert('RGBA')
        # Salva come PNG in memoria
        png_buffer = io.BytesIO()
        img.save(png_buffer, format='PNG', optimize=True)
        png_data = png_buffer.getvalue()
        # Limite PNG finale: 500 KB
        if len(png_data) > 500 * 1024:
            return _json_api_error(request, "File troppo grande dopo conversione. Max 500 KB.", 400, "file_too_large")
    except Exception as e:
        return _json_api_error(request, f"Errore conversione immagine: {e}", 400, "conversion_error")

    # Sanitizza il nome file e forza estensione .png
    base_name = _normalizza_nome_file_icona(uploaded_file.name)
    safe_name = os.path.splitext(base_name)[0] + '.png'

    # Crea la directory se non esiste
    os.makedirs(ICON_UPLOAD_DIR, exist_ok=True)

    dest_path = os.path.join(ICON_UPLOAD_DIR, safe_name)
    with open(dest_path, "wb") as dest:
        dest.write(png_data)

    logger.info("Icon uploaded: %s (from %s) by %s", safe_name, uploaded_file.name, request.user.username)
    return JsonResponse({"success": True, "filename": safe_name})


def landing_it(request):
    """Landing page in Italiano."""
    from .models import ImpostazioniSistema
    imp = ImpostazioniSistema.get()
    # Save language preference if user is logged in
    if request.user.is_authenticated:
        profile = getattr(request.user, 'profile', None)
        if profile and profile.lingua != 'it':
            profile.lingua = 'it'
            profile.save(update_fields=['lingua'])
    return render(request, "caricamento/landing_it.html", {
        "google_oauth_attivo": imp.google_oauth_attivo,
        "demo_attiva": imp.demo_attiva,
        "login_error": False,
        "trial_expired": False,
        "trial_used": False,
        "email_verification_sent": False,
        "password_confirmation_error": False,
        "account_disabled": False,
        "demo_disabled": False,
        "switch_account": False,
    })


def landing_en(request):
    """Landing page in English."""
    # Save language preference if user is logged in
    if request.user.is_authenticated:
        profile = getattr(request.user, 'profile', None)
        if profile and profile.lingua != 'en':
            profile.lingua = 'en'
            profile.save(update_fields=['lingua'])
    return render(request, "caricamento/landing_en.html")


def privacy_en(request):
    """Privacy policy in English."""
    return render(request, "caricamento/privacy_en.html")


def cookie_en(request):
    """Cookie policy in English."""
    return render(request, "caricamento/cookie_en.html")


def termini_en(request):
    """Terms of service in English."""
    return render(request, "caricamento/termini_en.html")


def rimborsi_en(request):
    """Refund policy in English."""
    return render(request, "caricamento/rimborsi_en.html")


# ===========================================================================
# API: Lingua utente
# ===========================================================================

@api_view(["GET", "PUT"])
@permission_classes([IsAuthenticated])
def api_user_lingua(request):
    """Legge o salva la lingua preferita dell'utente.

    GET /api/user-lingua/
        Restituisce {"lingua": "it"|"en"}.

    PUT /api/user-lingua/
        Aggiorna la lingua. Payload: {"lingua": "it"} oppure {"lingua": "en"}.
    """
    profile = getattr(request.user, 'profile', None)
    if not profile:
        return Response({"error": "Profile not found"}, status=400)

    if request.method == "GET":
        return Response({"lingua": profile.lingua or "en"})

    lingua = request.data.get("lingua", "en")
    if lingua not in ('it', 'en'):
        return Response({"error": "Invalid language"}, status=400)

    profile.lingua = lingua
    profile.save(update_fields=['lingua'])
    return Response({"lingua": lingua})

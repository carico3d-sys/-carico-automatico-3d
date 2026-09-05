"""
Test suite per il sistema di ottimizzazione carico 3D (Algoritmo 3D Semplificato).

Copre:
- VincoloOggetto: rotazioni, sovrapponibilità, peso massimo tetto, solo su piano
- VincoloTraOggetti: sopra
- ConfigurazioneOttimizzazione
- Rotazioni TreDPacker
"""

import copy
import hashlib
import hmac
import io
import json
import random
import tempfile
from datetime import timedelta
from decimal import Decimal
import unittest
from unittest.mock import Mock, patch

from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import RequestFactory, TestCase, override_settings
from django.utils import timezone
from PIL import Image
from rest_framework.test import APIClient

from caricamento.client_ip import get_client_ip
from caricamento.views import (
    _check_demo_abuse,
    _crea_excel_oggetti,
    _save_demo_fingerprints,
)

from caricamento.engine.common import (
    COLORI_PACCHI,
    ConfigurazioneOttimizzazione,
    ItemPacked,
    _build_lookup_vincoli_tra,
)
from caricamento.models import (
    Contenitore,
    ImpostazioniSistema,
    Oggetto,
    OggettoDaCaricare,
    OggettoPosizionato,
    PianoDiCarico,
    SezioneCarico,
    StatoPiano,
    UserProfile,
    VincoloOggetto,
    VincoloTraOggetti,
)


class TestCatalogoInizialeNuovoUtente(TestCase):
    """Ogni nuovo utente riceve una copia privata del catalogo iniziale."""

    def test_crea_catalogo_e_piano_ottimizzato_per_nuovo_utente(self):
        user = User.objects.create_user(username="starter-user")

        self.assertTrue(UserProfile.objects.filter(user=user).exists())
        self.assertEqual(Oggetto.objects.filter(owner=user).count(), 14)
        self.assertEqual(VincoloOggetto.objects.filter(oggetto__owner=user).count(), 14)
        self.assertEqual(Contenitore.objects.filter(owner=user).count(), 3)
        self.assertEqual(SezioneCarico.objects.filter(contenitore__owner=user).count(), 4)
        self.assertEqual(VincoloTraOggetti.objects.filter(oggetto_a__owner=user).count(), 171)

        piano = PianoDiCarico.objects.get(owner=user)
        self.assertEqual(piano.nome, "Carico 03/09/2026 16:06")
        self.assertEqual(piano.stato, StatoPiano.COMPLETATO)
        self.assertEqual(piano.oggetti_da_caricare.count(), 6)
        self.assertEqual(piano.oggetti_posizionati.count(), 78)

    def test_due_utenti_ricevono_copie_indipendenti(self):
        primo = User.objects.create_user(username="starter-user-one")
        secondo = User.objects.create_user(username="starter-user-two")

        primo_oggetto = Oggetto.objects.get(owner=primo, codice="PLT-001")
        secondo_oggetto = Oggetto.objects.get(owner=secondo, codice="PLT-001")
        self.assertNotEqual(primo_oggetto.pk, secondo_oggetto.pk)

        primo_oggetto.descrizione = "Modificato dal primo utente"
        primo_oggetto.save(update_fields=["descrizione"])
        secondo_oggetto.refresh_from_db()
        self.assertNotEqual(primo_oggetto.descrizione, secondo_oggetto.descrizione)
        self.assertEqual(secondo_oggetto.descrizione, "EPAL EUR 1 - European standard, logistics")

        self.assertEqual(PianoDiCarico.objects.filter(owner=primo).count(), 1)
        self.assertEqual(PianoDiCarico.objects.filter(owner=secondo).count(), 1)



from caricamento.engine.tre_d.packer_3d_v2 import (
    Obj,
    _prova_tutte_orientazioni,
    _trova_singoli_interni,
    _piazza_deferiti_in_coda,
    load_truck_v2,
)
from caricamento.engine.orchestratore_tre_d import (
    _piano_e_parziale,
    esegui_ottimizzazione_tre_d,
)
from caricamento.engine.tre_d.data_adapter import _carica_vincoli_sopra
from caricamento.engine.tre_d.geometry import (
    center_of_mass,
    compute_overhang,
    intersection_area,
    point_inside,
    rect,
)
from caricamento.engine.tre_d.placement_rules import (
    can_stack,
    check_z_collision,
)
from caricamento.engine.tre_d.postprocessing import (
    _trova_vuoti_xy,
    compatta_gradini_x,
    defer_singles,
    fill_xy_voids,
    find_internal_singles,
    has_object_above,
    has_object_ahead,
    place_singles_at_end,
)
from caricamento.engine.tre_d.constraints import (
    column_contains,
    evaluate_relational_constraint,
)
from caricamento.engine.tre_d.strategies import (
    BacktrackingStrategy,
    DeterministicStrategy,
    HybridStrategy,
    MonteCarloStrategy,
    strategy_for_config,
)
from caricamento.engine.tre_d.group_optimizer import (
    _compactness_key,
    _score,
    _group_is_relational,
    _relational_block_valid,
    _relational_blocks,
    candidate_rotation_counts,
)
from caricamento.engine.tre_d.tre_d_packer import TreDPacker
from caricamento.engine.tre_d.random_packer import run_packing_random
from caricamento.engine.tre_d.grid import SpatialGrid


# =============================================================================
# TEST: FACTORY STRATEGIE
# =============================================================================

class TestOggettoColoriAnagrafica(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="color-owner")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def _payload(self, codice):
        return {
            "codice": codice,
            "descrizione": "Oggetto test",
            "lunghezza_cm": 10,
            "larghezza_cm": 10,
            "altezza_cm": 10,
            "peso_kg": "1.00",
            "quantita_disponibile": 1,
        }

    def test_nuovi_oggetti_usano_colori_anagrafica_distinti(self):
        risposte = [
            self.client.post("/api/oggetti/", self._payload(f"COLOR-{i}"), format="json")
            for i in range(3)
        ]
        self.assertTrue(all(response.status_code == 201 for response in risposte))
        self.assertEqual(
            [response.data["colore"] for response in risposte],
            COLORI_PACCHI[:3],
        )

    def test_colore_di_riga_non_occupa_un_colore_anagrafico(self):
        oggetto = Oggetto.objects.create(
            owner=self.user,
            codice="ROW-COLOR",
            lunghezza_mm=100,
            larghezza_mm=100,
            altezza_mm=100,
            peso_kg=Decimal("1"),
        )
        container = Contenitore.objects.create(
            owner=self.user,
            nome="Color container",
            lunghezza_mm=1000,
            larghezza_mm=1000,
            altezza_mm=1000,
            carico_massimo_kg=Decimal("100"),
        )
        piano = PianoDiCarico.objects.create(owner=self.user, nome="Color plan", contenitore=container)
        OggettoDaCaricare.objects.create(piano_di_carico=piano, oggetto=oggetto, colore=COLORI_PACCHI[0])

        response = self.client.post("/api/oggetti/", self._payload("ANAG-COLOR"), format="json")
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["colore"], COLORI_PACCHI[0])




try:
    import openpyxl  # noqa: F401
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl non installato")
class TestOggettiExcel(TestCase):
    """Verifica il formato Excel dell'anagrafica oggetti."""

    def setUp(self):
        self.user = User.objects.create_user(username="excel-owner")

    def crea_oggetto(self, codice):
        return Oggetto.objects.create(
            owner=self.user,
            codice=codice,
            descrizione="Oggetto Excel",
            lunghezza_mm=100,
            larghezza_mm=200,
            altezza_mm=300,
            peso_kg=Decimal("2.50"),
            quantita_disponibile=2,
        )

    def test_export_crea_tre_fogli_e_filtra_i_vincoli(self):
        from openpyxl import load_workbook

        primo = self.crea_oggetto("EXCEL-A")
        secondo = self.crea_oggetto("EXCEL-B")
        esterno = self.crea_oggetto("EXCEL-C")
        VincoloTraOggetti.objects.create(
            oggetto_a=primo,
            oggetto_b=secondo,
            tipo_relazione="sopra",
        )
        VincoloTraOggetti.objects.create(
            oggetto_a=primo,
            oggetto_b=esterno,
            tipo_relazione="sopra",
        )

        output = _crea_excel_oggetti(
            [primo, secondo],
            VincoloTraOggetti.objects.filter(
                oggetto_a_id__in=[primo.id, secondo.id],
                oggetto_b_id__in=[primo.id, secondo.id],
            ).select_related("oggetto_a", "oggetto_b"),
        )
        workbook = load_workbook(output, read_only=True, data_only=True)

        self.assertEqual(workbook.sheetnames, ["Oggetti", "Rotazioni", "Vincoli"])
        self.assertEqual(workbook["Oggetti"].max_row, 3)
        self.assertEqual(workbook["Rotazioni"].max_row, 3)
        self.assertEqual(
            tuple(cell.value for cell in workbook["Vincoli"][2]),
            ("EXCEL-A", "EXCEL-B", "sopra", True, None, None),
        )

    def test_import_add_crea_oggetti_rotazioni_e_vincoli(self):
        from openpyxl import Workbook

        workbook = Workbook()
        objects_sheet = workbook.active
        objects_sheet.title = "Oggetti"
        objects_sheet.append([
            "Codice", "Descrizione", "Lunghezza_mm", "Larghezza_mm", "Altezza_mm",
            "Peso_kg", "Quantita_disponibile", "Colore", "Archiviato",
        ])
        objects_sheet.append(["IMPORT-A", "Importato", 100, 200, 300, 2.5, 1, "#123456", False])
        rotations_sheet = workbook.create_sheet("Rotazioni")
        rotations_sheet.append([
            "Codice", "Rotazione_consentita", "Rotazione_su_X", "Rotazione_su_Y",
            "Rotazione_su_Z", "Sovrapponibile", "Peso_massimo_tetto_kg", "Fragile",
            "Merce_pericolosa", "Solo_su_piano", "Aggancio_forche", "Note",
        ])
        rotations_sheet.append(["IMPORT-A", True, False, True, True, False, 10, True, False, True, False, "nota"])
        constraints_sheet = workbook.create_sheet("Vincoli")
        constraints_sheet.append([
            "Oggetto_A", "Oggetto_B", "Tipo_relazione", "Attivo",
            "Dettagli_posizionamento", "Note",
        ])

        stream = io.BytesIO()
        workbook.save(stream)
        upload = SimpleUploadedFile(
            "oggetti.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = APIClient()
        response.force_authenticate(user=self.user)
        result = response.post(
            "/api/oggetti/import-excel/",
            {"file": upload, "modalita": "add"},
            format="multipart",
        )

        self.assertEqual(result.status_code, 200)
        obj = Oggetto.objects.get(owner=self.user, codice="IMPORT-A")
        self.assertEqual(obj.lunghezza_mm, 100)
        self.assertTrue(obj.vincoli.fragile)
        self.assertFalse(obj.vincoli.sovrapponibile)



class TestStrategyFactory(TestCase):
    def test_seleziona_deterministica(self):
        config = ConfigurazioneOttimizzazione()
        self.assertIsInstance(strategy_for_config(config), DeterministicStrategy)

    def test_seleziona_monte_carlo(self):
        config = ConfigurazioneOttimizzazione(ordinamento_casuale=True)
        strategia = strategy_for_config(config)
        self.assertIsInstance(strategia, MonteCarloStrategy)
        self.assertEqual(strategia.num_restarts, 20)

    def test_seleziona_backtracking(self):
        config = ConfigurazioneOttimizzazione(backtracking_avanzato=True)
        self.assertIsInstance(strategy_for_config(config), BacktrackingStrategy)

    def test_seleziona_ibrida(self):
        config = ConfigurazioneOttimizzazione(
            ordinamento_casuale=True,
            backtracking_avanzato=True,
        )
        self.assertIsInstance(strategy_for_config(config), HybridStrategy)

    def test_tredpacker_delega_alla_strategy_selezionata(self):
        oggetto = Oggetto.objects.create(
            owner=User.objects.create_user(username="strategy-owner"),
            codice="STRATEGY-TEST",
            lunghezza_mm=800,
            larghezza_mm=800,
            altezza_mm=800,
            peso_kg=Decimal("10"),
            quantita_disponibile=1,
        )
        config = ConfigurazioneOttimizzazione(
            ordinamento_casuale=True,
            backtracking_avanzato=True,
            compattazione_aggressiva=True,
        )
        risultato_obj = Obj(
            "STRATEGY-TEST-0", 80, 80, 80, oggetto_id=oggetto.pk
        )
        risultato_obj._peso_kg = 10.0
        strategia = Mock()
        strategia.execute.return_value = [risultato_obj]

        packer = TreDPacker(
            bin_dimensioni=(2000, 2000, 2000),
            peso_max_kg=1000,
            configurazione=config,
        )
        packer.aggiungi_oggetto(oggetto)

        with patch(
            "caricamento.engine.tre_d.tre_d_packer.strategy_for_config",
            return_value=strategia,
        ) as factory:
            packer.esegui()

        factory.assert_called_once_with(config, ops=1)
        strategia.execute.assert_called_once()
        args, kwargs = strategia.execute.call_args
        self.assertEqual(args[1], {})
        self.assertEqual(args[2], (200, 200, 200))
        self.assertIs(kwargs["tracker"], None)
        self.assertTrue(kwargs["compattazione_aggressiva"])
        self.assertEqual(len(packer.results), 1)
        self.assertEqual(packer.results[0].codice, oggetto.codice)

    def test_adattatori_eseguono_e_non_duplicano_istanze(self):
        def make_objects():
            return [
                Obj("A-0", 80, 80, 80, oggetto_id=1),
                Obj("B-0", 70, 70, 70, oggetto_id=2),
            ]

        strategie = [
            DeterministicStrategy(),
            MonteCarloStrategy(num_restarts=1),
            BacktrackingStrategy(iterations=1),
            HybridStrategy(iterations=1, num_restarts=1),
        ]
        for strategia in strategie:
            with self.subTest(strategy=strategia.name):
                risultato = strategia.execute(
                    make_objects(),
                    {},
                    (500, 500, 500),
                )
                ids = [obj.id for obj in risultato if obj.z >= 0]
                self.assertIsInstance(risultato, list)
                self.assertEqual(len(ids), len(set(ids)))
                self.assertGreaterEqual(len(ids), 1)


class TestRigheDuplicateCarico(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="duplicate-rows-user")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)
        self.container = Contenitore.objects.create(
            owner=self.user,
            nome="Duplicate rows container",
            lunghezza_mm=4000,
            larghezza_mm=2000,
            altezza_mm=2000,
            carico_massimo_kg=Decimal("1000"),
        )
        self.oggetto = Oggetto.objects.create(
            owner=self.user,
            codice="DUP-ROW",
            lunghezza_mm=500,
            larghezza_mm=400,
            altezza_mm=300,
            peso_kg=Decimal("10"),
            quantita_disponibile=10,
        )
        self.piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Duplicate rows plan",
            contenitore=self.container,
        )

    def test_post_same_object_creates_two_distinct_rows(self):
        url = f"/api/piani/{self.piano.id}/oggetti_da_caricare/"
        first = self.client.post(
            url,
            {"oggetto_id": self.oggetto.id, "quantita": 2, "priorita": 1},
            format="json",
        )
        second = self.client.post(
            url,
            {"oggetto_id": self.oggetto.id, "quantita": 1, "priorita": 0},
            format="json",
        )

        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 201)
        self.assertNotEqual(first.data["id"], second.data["id"])
        self.assertEqual(self.piano.oggetti_da_caricare.count(), 2)
        self.assertEqual(
            set(self.piano.oggetti_da_caricare.values_list("priorita", flat=True)),
            {0, 1},
        )

    def test_optimizer_preserves_source_row_for_each_instance(self):
        first = OggettoDaCaricare.objects.create(
            piano_di_carico=self.piano,
            oggetto=self.oggetto,
            quantita=2,
            priorita=1,
        )
        second = OggettoDaCaricare.objects.create(
            piano_di_carico=self.piano,
            oggetto=self.oggetto,
            quantita=1,
            priorita=0,
        )

        risultato = esegui_ottimizzazione_tre_d(self.piano.id)

        self.assertTrue(risultato.successo)
        posizionati = list(
            OggettoPosizionato.objects.filter(piano_di_carico=self.piano)
        )
        self.assertEqual(len(posizionati), 3)
        self.assertEqual(
            sum(item.riga_origine_id == first.id for item in posizionati),
            2,
        )
        self.assertEqual(
            sum(item.riga_origine_id == second.id for item in posizionati),
            1,
        )

    def test_post_row_accepts_per_row_colore(self):
        url = f"/api/piani/{self.piano.id}/oggetti_da_caricare/"
        response = self.client.post(
            url,
            {"oggetto_id": self.oggetto.id, "quantita": 2, "colore": "#ff00aa"},
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["colore"], "#ff00aa")
        riga = self.piano.oggetti_da_caricare.get(pk=response.data["id"])
        self.assertEqual(riga.colore, "#ff00aa")

    def test_patch_row_updates_only_that_row(self):
        prima = OggettoDaCaricare.objects.create(
            piano_di_carico=self.piano,
            oggetto=self.oggetto,
            quantita=1,
            priorita=0,
        )
        seconda = OggettoDaCaricare.objects.create(
            piano_di_carico=self.piano,
            oggetto=self.oggetto,
            quantita=1,
            priorita=0,
        )

        url = f"/api/piani/{self.piano.id}/oggetti_da_caricare/"
        response = self.client.patch(
            url,
            {
                "riga_id": prima.id,
                "quantita": 3,
                "priorita": 2,
                "colore": "#00ff55",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        prima.refresh_from_db()
        seconda.refresh_from_db()
        self.assertEqual(prima.quantita, 3)
        self.assertEqual(prima.priorita, 2)
        self.assertEqual(prima.colore, "#00ff55")
        # La seconda riga (stesso codice) non viene toccata.
        self.assertEqual(seconda.quantita, 1)
        self.assertEqual(seconda.priorita, 0)
        self.assertEqual(seconda.colore, "")

    def test_patch_row_rejects_riga_of_another_plan(self):
        altro_piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Altro piano",
            contenitore=self.container,
        )
        riga_altra = OggettoDaCaricare.objects.create(
            piano_di_carico=altro_piano,
            oggetto=self.oggetto,
            quantita=1,
        )
        url = f"/api/piani/{self.piano.id}/oggetti_da_caricare/"
        response = self.client.patch(
            url,
            {"riga_id": riga_altra.id, "colore": "#123456"},
            format="json",
        )
        self.assertEqual(response.status_code, 400)


    def test_optimizer_uses_row_color_for_each_lot(self):
        prima = OggettoDaCaricare.objects.create(
            piano_di_carico=self.piano,
            oggetto=self.oggetto,
            quantita=1,
            priorita=1,
            colore="#ff0000",
        )
        seconda = OggettoDaCaricare.objects.create(
            piano_di_carico=self.piano,
            oggetto=self.oggetto,
            quantita=1,
            priorita=0,
            colore="#00ff00",
        )

        risultato = esegui_ottimizzazione_tre_d(self.piano.id)
        self.assertTrue(risultato.successo)

        posizionati = list(
            OggettoPosizionato.objects.filter(piano_di_carico=self.piano)
        )
        colori_per_riga = {
            op.riga_origine_id: op.colore for op in posizionati
        }
        self.assertEqual(colori_per_riga[prima.id], "#ff0000")
        self.assertEqual(colori_per_riga[seconda.id], "#00ff00")

    def test_optimizer_falls_back_to_anagrafica_color_without_row_color(self):
        # Nessun colore di riga: il motore usa il colore dell'anagrafica.
        self.oggetto.colore = "#abcdef"
        self.oggetto.save(update_fields=["colore"])
        OggettoDaCaricare.objects.create(
            piano_di_carico=self.piano,
            oggetto=self.oggetto,
            quantita=1,
        )

        risultato = esegui_ottimizzazione_tre_d(self.piano.id)
        self.assertTrue(risultato.successo)
        posizionato = OggettoPosizionato.objects.get(
            piano_di_carico=self.piano
        )
        self.assertEqual(posizionato.colore, "#abcdef")

    def test_optimizer_reassigns_shared_custom_color_between_duplicate_rows(self):
        """Due righe con lo STESSO colore custom devono ottenere colori
        diversi dal motore (replica della logica _assegnaColoriAutomatici
        del frontend: un colore condiviso da 2+ righe va riassegnato)."""
        prima = OggettoDaCaricare.objects.create(
            piano_di_carico=self.piano,
            oggetto=self.oggetto,
            quantita=1,
            priorita=1,
            colore="#FF0000",
        )
        seconda = OggettoDaCaricare.objects.create(
            piano_di_carico=self.piano,
            oggetto=self.oggetto,
            quantita=1,
            priorita=0,
            colore="#FF0000",
        )

        risultato = esegui_ottimizzazione_tre_d(self.piano.id)
        self.assertTrue(risultato.successo)

        posizionati = list(
            OggettoPosizionato.objects.filter(piano_di_carico=self.piano)
        )
        colori_per_riga = {
            op.riga_origine_id: op.colore for op in posizionati
        }
        # Due lotti distinti: colori distinti, mai lo stesso colore condiviso.
        self.assertEqual(len(set(colori_per_riga.values())), 2)
        self.assertNotEqual(
            colori_per_riga[prima.id], colori_per_riga[seconda.id]
        )

    def test_optimizer_keeps_unique_custom_color_and_reassigns_shared_one(self):
        """Una riga con colore custom UNICO lo mantiene; l'altra riga dello
        stesso codice senza personalizzazione ottiene un colore diverso dalla
        palette (mai uguale al custom unico né all'anagrafica)."""
        self.oggetto.colore = "#123456"
        self.oggetto.save(update_fields=["colore"])
        unica = OggettoDaCaricare.objects.create(
            piano_di_carico=self.piano,
            oggetto=self.oggetto,
            quantita=1,
            priorita=1,
            colore="#00AA00",
        )
        senza_custom = OggettoDaCaricare.objects.create(
            piano_di_carico=self.piano,
            oggetto=self.oggetto,
            quantita=1,
            priorita=0,
            colore="",
        )

        risultato = esegui_ottimizzazione_tre_d(self.piano.id)
        self.assertTrue(risultato.successo)

        posizionati = list(
            OggettoPosizionato.objects.filter(piano_di_carico=self.piano)
        )
        colori_per_riga = {
            op.riga_origine_id: op.colore for op in posizionati
        }
        # La riga custom unica mantiene il suo colore; l'altra riga ne ha uno
        # diverso (e non l'anagrafica, per distinguere i lotti).
        self.assertEqual(colori_per_riga[unica.id], "#00AA00")
        self.assertNotEqual(
            colori_per_riga[unica.id], colori_per_riga[senza_custom.id]
        )
        self.assertNotEqual(colori_per_riga[senza_custom.id], "#123456")


class TestGroupOptimizer(TestCase):
    """Verifica le configurazioni locali senza coinvolgere MC o v3."""

    def test_piano_parziale_dipende_solo_dalle_istanze_mancanti(self):
        self.assertFalse(_piano_e_parziale(10, 10))
        self.assertTrue(_piano_e_parziale(10, 9))

    def test_vincoli_o_priorita_non_rendono_parziale_un_piano_completo(self):
        # La validazione dei vincoli/priorità è un report separato: con tutte
        # le istanze posizionate il piano non è parziale.
        self.assertFalse(_piano_e_parziale(4, 4))

    def test_configurazioni_dipendono_dalla_capacita_y(self):
        # I05: 3 pezzi originali per Y, 4 dopo la rotazione.
        self.assertEqual(
            candidate_rotation_counts(19, 60, 80, (1360, 248, 270)),
            [4, 8],
        )
        # I02: 2 pezzi originali per Y, 3 dopo la rotazione.
        self.assertEqual(
            candidate_rotation_counts(18, 80, 120, (1360, 248, 270)),
            [3, 6],
        )
        # I01: la rotazione non aumenta le righe Y, quindi nessuna variante.
        self.assertEqual(
            candidate_rotation_counts(13, 100, 120, (1360, 248, 270)),
            [],
        )

    def test_quantita_piccola_non_forza_un_blocco_completo(self):
        self.assertEqual(
            candidate_rotation_counts(2, 60, 80, (1360, 248, 270)),
            [2],
        )

    def test_auto_vincolo_non_esclude_il_gruppo_ma_vincolo_tra_codici_si(self):
        self.assertFalse(
            _group_is_relational((7,), {7: {7: None}})
        )
        self.assertTrue(
            _group_is_relational((7,), {7: {8: None}})
        )
        self.assertTrue(
            _group_is_relational((8,), {7: {8: None}})
        )

    def test_deterministica_usa_il_modulo_dei_gruppi(self):
        objects = [
            Obj("GROUP-0", 60, 80, 100, oggetto_id=1),
            Obj("GROUP-1", 60, 80, 100, oggetto_id=1),
        ]
        with patch(
            "caricamento.engine.tre_d.strategies.deterministic.optimize_deterministic_groups",
            return_value=objects,
        ) as optimizer:
            result = DeterministicStrategy().execute(
                objects, {}, (1360, 248, 270)
            )

        optimizer.assert_called_once()
        self.assertIs(result, objects)

    def test_auto_vincolo_consente_blocco_ruotato_validato(self):
        objects = [
            Obj(
                f"SELF-{index}",
                60,
                80,
                100,
                oggetto_id=7,
                rotazione_su_x=False,
                rotazione_su_y=False,
                rotazione_su_z=True,
            )
            for index in range(4)
        ]
        result = DeterministicStrategy().execute(
            objects, {7: {7: None}}, (500, 120, 200)
        )
        fitted = [obj for obj in result if obj.z >= 0]

        self.assertEqual(len(fitted), 4)
        self.assertLessEqual(max(obj.x + obj.width for obj in fitted), 160)
        self.assertTrue(
            any((obj.width, obj.depth) == (80, 60) for obj in fitted)
        )
        self.assertTrue(
            any(obj.z == 100 for obj in fitted),
            "L'auto-vincolo deve lasciare almeno una coppia impilata.",
        )

    def test_colonna_cross_ruotata_mantiene_allineamento_e_vincolo(self):
        base = Obj(
            "BASE-0", 100, 120, 100, oggetto_id=1,
            rotazione_su_x=False, rotazione_su_y=False, rotazione_su_z=True,
        )
        top = Obj(
            "TOP-0", 80, 120, 100, oggetto_id=2,
            rotazione_su_x=False, rotazione_su_y=False, rotazione_su_z=True,
        )
        base.x = top.x = 0
        base.y = top.y = 0
        base.z = 0
        top.z = 100
        self.assertTrue(
            _relational_block_valid(
                [base, top],
                (base.id, top.id),
                {2: {1: None}},
            )
        )

        top.y = 10
        self.assertFalse(
            _relational_block_valid(
                [base, top],
                (base.id, top.id),
                {2: {1: None}},
            )
        )

    def test_compattezza_preferisce_fasce_y_regolari(self):
        ordinato = [
            Obj(f"O-{index}", 60, 80, 100, oggetto_id=1)
            for index in range(3)
        ]
        frammentato = [
            Obj(f"F-{index}", 60, 80, 100, oggetto_id=1)
            for index in range(3)
        ]
        for obj, y in zip(ordinato, (0, 80, 160)):
            obj.x, obj.y, obj.z = 0, y, 0
        for obj, y in zip(frammentato, (0, 60, 160)):
            obj.x, obj.y, obj.z = 0, y, 0

        self.assertLess(
            _compactness_key(ordinato),
            _compactness_key(frammentato),
        )

    def test_score_completo_preferisce_layout_compatto(self):
        ordinato = [
            Obj(f"S-{index}", 60, 80, 100, oggetto_id=1)
            for index in range(2)
        ]
        frammentato = [
            Obj("T-0", 80, 60, 100, oggetto_id=1),
            Obj("T-1", 60, 80, 100, oggetto_id=1),
        ]
        for obj, y in zip(ordinato, (0, 80)):
            obj.x, obj.y, obj.z = 0, y, 0
        for obj, y in zip(frammentato, (0, 60)):
            obj.x, obj.y, obj.z = 0, y, 0

        self.assertGreater(
            _score(ordinato, (500, 200, 200), ordinato),
            _score(frammentato, (500, 200, 200), frammentato),
        )

    def test_compattezza_penalizza_orientamenti_misti(self):
        uniforme = [
            Obj(f"U-{index}", 60, 80, 100, oggetto_id=1)
            for index in range(2)
        ]
        misto = [
            Obj("M-0", 80, 60, 100, oggetto_id=1),
            Obj("M-1", 60, 80, 100, oggetto_id=1),
        ]
        for obj, y in zip(uniforme, (0, 80)):
            obj.x, obj.y, obj.z = 0, y, 0
        for obj, y in zip(misto, (0, 60)):
            obj.x, obj.y, obj.z = 0, y, 0

        self.assertLess(
            _compactness_key(uniforme),
            _compactness_key(misto),
        )

    def test_blocco_cross_reale_viene_rilevato_dalla_colonna(self):
        base = Obj(
            "BASE-0", 100, 120, 100, oggetto_id=1,
            rotazione_su_x=False, rotazione_su_y=False, rotazione_su_z=True,
        )
        top = Obj(
            "TOP-0", 80, 120, 100, oggetto_id=2,
            rotazione_su_x=False, rotazione_su_y=False, rotazione_su_z=True,
        )
        base.x = top.x = 0
        base.y = top.y = 0
        base.z = 0
        top.z = 100
        blocks = _relational_blocks([base, top], {2: {1: None}})
        self.assertEqual(len(blocks), 1)
        self.assertEqual(blocks[0][0], ("BASE-0", "TOP-0"))

    def test_blocco_ruotato_viene_accettato_solo_se_migliora_x(self):
        # Con Y=120 l'orientamento 60x80 crea una sola riga, mentre 80x60
        # ne crea due: quattro pezzi passano da quattro colonne a due.
        objects = [
            Obj(
                f"BLOCK-{index}",
                60,
                80,
                100,
                oggetto_id=7,
                rotazione_su_x=False,
                rotazione_su_y=False,
                rotazione_su_z=True,
            )
            for index in range(4)
        ]
        result = DeterministicStrategy().execute(
            objects, {}, (500, 120, 100)
        )
        fitted = [obj for obj in result if obj.z >= 0]

        self.assertEqual(len(fitted), 4)
        self.assertLessEqual(max(obj.x + obj.width for obj in fitted), 160)
        self.assertTrue(
            all((obj.width, obj.depth) == (80, 60) for obj in fitted)
        )


class TestRiempimentoFasce(TestCase):
    """Regressione sul riempimento Y/Z prima dell'avanzamento su X."""

    def test_carico_reale_trova_una_disposizione_compatta(self):
        """Il carico del piano reale non deve terminare con una coda lunga.

        Le dimensioni sono quelle del piano 348, espresse in cm. Con tutti
        gli oggetti richiesti e 20 restart, la soluzione deve caricare tutto
        e chiudere la lunghezza entro 12,20 m nel caso deterministico.
        """
        dati = [
            ("CART-I03", 160, 70, 100, 9, 3),
            ("CART-I01", 120, 100, 100, 13, 1),
            ("CART-I02", 120, 80, 100, 18, 2),
            ("cart-i05", 60, 80, 100, 17, 5),
        ]
        objects = []
        for codice, width, depth, height, quantita, oggetto_id in dati:
            for indice in range(quantita):
                objects.append(
                    Obj(
                        f"{codice}-{indice}",
                        width,
                        depth,
                        height,
                        oggetto_id=oggetto_id,
                        orientation_allowed=True,
                        rotazione_su_x=False,
                        rotazione_su_y=False,
                        rotazione_su_z=True,
                    )
                )

        # I vincoli auto-referenziali riproducono il coinvolgimento dei codici
        # nei vincoli relazionali del piano senza imporre coppie tra tipi.
        vincoli_sopra = {
            oggetto_id: {oggetto_id: None}
            for oggetto_id in (1, 2, 3, 5)
        }
        stato_random = random.getstate()
        try:
            random.seed(0)
            risultato = run_packing_random(
                objects,
                vincoli_sopra=vincoli_sopra,
                num_restarts=20,
                container_dim=(1360, 248, 270),
            )
        finally:
            random.setstate(stato_random)

        posizionati = [obj for obj in risultato if obj.z >= 0]
        self.assertEqual(len(posizionati), 57)
        self.assertEqual(len([obj for obj in risultato if obj.z < 0]), 0)
        self.assertLessEqual(
            max(obj.x + obj.width for obj in posizionati),
            1220,
            "Il riempimento Y/Z deve evitare una coda X inutilmente lunga.",
        )


# =============================================================================
# TEST: GEOMETRIA PURA
# =============================================================================

class TestGeometry(TestCase):
    """Regressioni per le primitive geometriche indipendenti dal packer."""

    def test_rettangolo_intersezione_e_centro(self):
        obj = Obj("A", 80, 120, 100, oggetto_id=1)
        obj.x, obj.y = 10, 20
        base = Obj("B", 70, 160, 100, oggetto_id=2)
        base.x, base.y = 0, 0

        self.assertEqual(
            rect(obj),
            {"x1": 10, "x2": 90, "y1": 20, "y2": 140},
        )
        self.assertEqual(intersection_area(rect(obj), rect(base)), 60 * 120)
        self.assertEqual(center_of_mass(obj), (50, 80))

    def test_sbalzo_e_punto_nel_rettangolo(self):
        obj = Obj("A", 80, 120, 100, oggetto_id=1)
        obj.x, obj.y = 10, 0
        base = Obj("B", 70, 160, 100, oggetto_id=2)
        base.x, base.y = 0, 0

        self.assertEqual(compute_overhang(rect(obj), rect(base)), 20)
        self.assertTrue(point_inside(50, 60, rect(base)))
        self.assertFalse(point_inside(80, 60, rect(base)))


# =============================================================================
# TEST: REGOLE DI POSIZIONAMENTO PURE
# =============================================================================

class TestPlacementRules(TestCase):
    """Regressioni per stacking e collisione indipendenti dalla strategia."""

    def test_can_stack_rispetta_area_e_supporto(self):
        base = Obj("BASE", 100, 100, 100, oggetto_id=1)
        base.x = base.y = base.z = 0
        sopra = Obj("SOPRA", 80, 80, 100, oggetto_id=2)
        sopra.x = sopra.y = 0
        sopra.z = 100

        self.assertTrue(can_stack(sopra, base))
        self.assertEqual(sopra.support_ratio, 1.0)

        troppo_grande = Obj("GRANDE", 120, 100, 100, oggetto_id=3)
        troppo_grande.x = troppo_grande.y = 0
        self.assertFalse(can_stack(troppo_grande, base))

    def test_can_stack_rispetta_fragilita_e_sovrapponibilita(self):
        base = Obj("BASE", 100, 100, 100, oggetto_id=1, fragile=True)
        sopra = Obj("SOPRA", 80, 80, 100, oggetto_id=2)
        self.assertFalse(can_stack(sopra, base))

        base.fragile = False
        base.sovrapponibile = False
        self.assertFalse(can_stack(sopra, base))

    def test_collisione_3d_distinta_dal_contatto_sovrapposto(self):
        base = Obj("BASE", 100, 100, 100, oggetto_id=1)
        base.x = base.y = base.z = 0
        sovrapposto = Obj("SOVRAPPOSTO", 50, 50, 50, oggetto_id=2)
        sovrapposto.x = sovrapposto.y = 10
        sovrapposto.z = 50
        self.assertTrue(check_z_collision(sovrapposto, [base]))

        affiancato = Obj("AFFIANCATO", 50, 50, 50, oggetto_id=3)
        affiancato.x, affiancato.y, affiancato.z = 100, 0, 0
        self.assertFalse(check_z_collision(affiancato, [base]))


# =============================================================================
# TEST: VINCOLI RELAZIONALI ISOLATI
# =============================================================================

class TestPostprocessingRules(TestCase):
    """Regressioni per il rilevamento dei singoli e delle fasce terminali."""

    def test_rileva_singolo_interno_indipendentemente_dai_vincoli(self):
        interno = Obj("INTERNO", 100, 100, 100, oggetto_id=1)
        interno.x = interno.y = interno.z = 0
        terminale = Obj("TERMINALE", 100, 100, 100, oggetto_id=2)
        terminale.x, terminale.y, terminale.z = 200, 0, 0

        self.assertFalse(has_object_above([interno, terminale], interno, 270))
        self.assertTrue(has_object_ahead([interno, terminale], interno))
        self.assertEqual(
            find_internal_singles([interno, terminale], 270, {1: {1: set()}}),
            ["INTERNO"],
        )

    def test_oggetto_nella_fascia_terminale_non_e_interno(self):
        primo = Obj("PRIMO", 100, 100, 100, oggetto_id=1)
        primo.x = primo.y = primo.z = 0
        affiancato = Obj("AFFIANCATO", 100, 100, 100, oggetto_id=2)
        affiancato.x, affiancato.y, affiancato.z = 200, 100, 0

        self.assertFalse(has_object_ahead([primo, affiancato], affiancato))
        self.assertEqual(find_internal_singles([primo, affiancato], 270), ["PRIMO"])

    def test_piazzamento_singoli_usa_callback_e_preserva_integrita(self):
        terminale = Obj("TERMINALE", 100, 100, 100, oggetto_id=1)
        terminale.x = terminale.y = terminale.z = 200
        singolo = Obj("SINGOLO", 80, 80, 100, oggetto_id=2)
        singolo.x = singolo.y = singolo.z = 0
        for obj in (terminale, singolo):
            obj._peso_kg = 1

        def try_orientations(obj, x, y, z, placed, container_dim, constraints, **kwargs):
            obj.x, obj.y, obj.z = x, y, z
            return True

        placed = [terminale, singolo]
        place_singles_at_end(
            placed,
            (500, 500, 270),
            {},
            None,
            try_orientations,
        )
        self.assertEqual({obj.id for obj in placed}, {"TERMINALE", "SINGOLO"})
        self.assertEqual(len({obj.id for obj in placed}), 2)

    def test_piazzamento_singoli_non_scarta_rotazione_valida(self):
        terminale = Obj("TERMINALE-ROT", 40, 100, 100, oggetto_id=1)
        terminale.x, terminale.y, terminale.z = 200, 0, 0
        singolo = Obj("SINGOLO-ROT", 100, 40, 100, oggetto_id=2)
        singolo.x = singolo.y = singolo.z = 0

        def try_orientations(obj, x, y, z, placed, container_dim, constraints, **kwargs):
            # Il callback simula il controllo reale: solo l'orientamento
            # ruotato 40x100 entra nella fascia terminale disponibile.
            if x + obj.width > container_dim[0] or y + obj.depth > container_dim[1]:
                if obj.width == 100 and obj.depth == 40:
                    obj.width, obj.depth = 40, 100
                else:
                    return False
            if x + obj.width <= container_dim[0] and y + obj.depth <= container_dim[1]:
                obj.x, obj.y, obj.z = x, y, z
                return True
            return False

        placed = [terminale, singolo]
        self.assertTrue(
            place_singles_at_end(
                placed, (280, 100, 270), {}, None, try_orientations
            )
        )
        self.assertEqual({obj.id for obj in placed}, {"TERMINALE-ROT", "SINGOLO-ROT"})

    def test_compatta_gradini_x_chiude_gradini_tra_fasce_y(self):
        """Le colonne scivolano a sinistra fino al vicino della stessa fascia Y."""
        # Fasce Y diverse che ripartono allineate a colonne di altre fasce:
        # - fascia Y=0..100:  colonna A a X=0..100, poi B a X=200 (gradino 100)
        # - fascia Y=100..200: colonna C a X=0..100, poi D a X=200 (gradino 100)
        placed = []
        for i, (x, y, w, d) in enumerate(
            [(0, 0, 100, 100), (200, 0, 100, 100), (0, 100, 100, 100), (200, 100, 100, 100)]
        ):
            o = Obj(f"SC-{i}", w, d, 50)
            o.x, o.y, o.z = x, y, 0
            placed.append(o)

        def try_orientations(obj, x, y, z, placed, container_dim, constraints, **kwargs):
            obj.x, obj.y, obj.z = x, y, z
            return True

        compatta_gradini_x(
            placed, (500, 300, 270), {}, None, try_orientations,
            max_passes=4,
        )

        # Le colonne a X=200 devono essere scivolate a X=100 (contigue al vicino)
        by_id = {o.id: o for o in placed}
        self.assertEqual(by_id["SC-1"].x, 100.0)
        self.assertEqual(by_id["SC-3"].x, 100.0)

    def test_compatta_gradini_x_scala_intera_colonna_impilata(self):
        """La pila sopra la base scivola insieme alla base, senza ruotare."""
        base = Obj("BASE", 100, 100, 50)
        base.x, base.y, base.z = 200, 0, 0
        sopra = Obj("SOPRA", 80, 80, 50)
        sopra.x, sopra.y, sopra.z = 200, 0, 50
        vicino = Obj("VICINO", 100, 100, 50)
        vicino.x, vicino.y, vicino.z = 0, 0, 0
        placed = [base, sopra, vicino]

        def try_orientations(obj, x, y, z, placed, container_dim, constraints, **kwargs):
            obj.x, obj.y, obj.z = x, y, z
            return True

        compatta_gradini_x(
            placed, (500, 300, 270), {}, None, try_orientations,
            max_passes=4,
        )
        self.assertEqual(base.x, 100.0)
        self.assertEqual(sopra.x, 100.0)
        self.assertEqual(sopra.z, 50.0)  # la pila resta sopra la base

    def test_compatta_gradini_x_non_sposta_prima_colonna_della_fascia(self):
        """Una colonna senza nulla dietro resta dove si trova (non va a X=0)."""
        prima = Obj("PRIMA-FASCIA", 60, 80, 100)
        prima.x, prima.y, prima.z = 580, 0, 0
        placed = [prima]

        def try_orientations(obj, x, y, z, placed, container_dim, constraints, **kwargs):
            obj.x, obj.y, obj.z = x, y, z
            return True

        compatta_gradini_x(
            placed, (1360, 248, 270), {}, None, try_orientations,
            max_passes=4,
        )
        self.assertEqual(prima.x, 580.0)

    def test_compatta_gradini_x_non_sposta_colonna_gia_allineata(self):
        """Una colonna gia' contigua non viene toccata."""
        a = Obj("A", 100, 100, 50)
        a.x, a.y, a.z = 0, 0, 0
        b = Obj("B", 100, 100, 50)
        b.x, b.y, b.z = 100, 0, 0
        placed = [a, b]

        def try_orientations(obj, x, y, z, placed, container_dim, constraints, **kwargs):
            obj.x, obj.y, obj.z = x, y, z
            return True

        compatta_gradini_x(
            placed, (500, 300, 270), {}, None, try_orientations,
            max_passes=4,
        )
        self.assertEqual(b.x, 100.0)

    def test_trova_vuoti_xy_rileva_colonna_intera_vuota(self):
        """Rileva una regione XY senza oggetti a nessuna altezza."""
        placed = []
        for i, (x, y) in enumerate([(0, 0), (150, 0), (0, 100), (150, 100)]):
            o = Obj(f"V-{i}", 50, 100, 50)
            o.x, o.y, o.z = x, y, 0
            placed.append(o)
        vuoti = _trova_vuoti_xy(placed, (300, 200, 270))
        self.assertTrue(
            any(abs(v["w"] - 100) < 1e-6 and abs(v["d"] - 200) < 1e-6 for v in vuoti),
            f"vuoti rilevati: {vuoti}",
        )

    def test_postprocessing_peggiorativo_viene_ripristinato(self):
        """Una riparazione che allunga X non deve sostituire il packing base."""
        oggetti = [
            Obj("BASE-0", 100, 100, 100, oggetto_id=1),
            Obj("BASE-1", 100, 100, 100, oggetto_id=2),
        ]
        risultato_base = load_truck_v2(
            copy.deepcopy(oggetti),
            vincoli_sopra={},
            container_dim=(500, 500, 270),
            preserve_order=True,
            _run_postprocessing=False,
        )
        coordinate_base = {
            obj.id: (obj.x, obj.y, obj.z, obj.width, obj.depth, obj.height)
            for obj in risultato_base
        }

        def peggiora_x(placed, *args, **kwargs):
            placed[0].x += 1000

        with patch(
            "caricamento.engine.tre_d.packer_3d_v2._deferral_pass",
            side_effect=peggiora_x,
        ), patch(
            "caricamento.engine.tre_d.packer_3d_v2._riempi_buchi_sicuro",
            return_value=True,
        ):
            risultato = load_truck_v2(
                copy.deepcopy(oggetti),
                vincoli_sopra={},
                container_dim=(500, 500, 270),
                preserve_order=True,
            )

        self.assertEqual(
            {
                obj.id: (obj.x, obj.y, obj.z, obj.width, obj.depth, obj.height)
                for obj in risultato
            },
            coordinate_base,
        )


class TestConstraintRules(TestCase):
    """Verifica gli stati assente, valido, escluso e in colonna."""

    def _base_pair(self):
        base = Obj("BASE", 70, 160, 100, oggetto_id=3)
        base.x = base.y = base.z = 0
        sopra = Obj("SOPRA", 80, 120, 100, oggetto_id=2)
        sopra.x = sopra.y = 0
        sopra.z = 100
        return sopra, base

    def test_relazione_assente_lascia_regole_standard(self):
        sopra, base = self._base_pair()
        self.assertEqual(
            evaluate_relational_constraint(sopra, base, [base], {}),
            (True, False, None),
        )

    def test_configurazione_valida_autorizza_solo_match(self):
        sopra, base = self._base_pair()
        validi = {((80, 120, 100), (70, 160, 100))}
        self.assertEqual(
            evaluate_relational_constraint(
                sopra, base, [base], {2: {3: validi}}
            ),
            (True, True, validi),
        )

    def test_configurazione_esclusa_blocca(self):
        sopra, base = self._base_pair()
        self.assertEqual(
            evaluate_relational_constraint(
                sopra, base, [base], {2: {3: set()}}
            ),
            (False, False, None),
        )

    def test_relazione_valida_trovata_nella_colonna(self):
        fondo = Obj("FONDO", 100, 100, 100, oggetto_id=1)
        fondo.x = fondo.y = fondo.z = 0
        base = Obj("BASE", 100, 100, 100, oggetto_id=3)
        base.x = base.y = 0
        base.z = 100
        sopra = Obj("SOPRA", 80, 80, 100, oggetto_id=2)
        sopra.x = sopra.y = 0
        sopra.z = 200

        self.assertTrue(column_contains([fondo, base], base, 1))
        allowed, relational, _ = evaluate_relational_constraint(
            sopra, base, [fondo, base], {2: {1: None}}
        )
        self.assertTrue(allowed)
        self.assertTrue(relational)

    def test_configurazione_in_colonna_confronta_la_base_richiesta(self):
        fondo = Obj("FONDO", 70, 160, 100, oggetto_id=1)
        fondo.x = fondo.y = fondo.z = 0
        intermedio = Obj("INTERMEDIO", 100, 100, 100, oggetto_id=3)
        intermedio.x = intermedio.y = 0
        intermedio.z = 100
        sopra = Obj("SOPRA", 80, 120, 100, oggetto_id=2)
        sopra.x = sopra.y = 0
        sopra.z = 200
        validi = {((80, 120, 100), (70, 160, 100))}

        allowed, relational, details = evaluate_relational_constraint(
            sopra, intermedio, [fondo, intermedio], {2: {1: validi}}
        )
        self.assertTrue(allowed)
        self.assertTrue(relational)
        self.assertEqual(details, validi)

        vietati = {((80, 120, 100), (80, 120, 100))}
        self.assertEqual(
            evaluate_relational_constraint(
                sopra, intermedio, [fondo, intermedio], {2: {1: vietati}}
            ),
            (False, False, None),
        )


# =============================================================================
# HELPERS per i test
# =============================================================================

class TestCaseBase(TestCase):
    """Base con metodi helper per creare dati di test rapidamente."""

    def setUp(self):
        self.user = User.objects.create_user(
            username="owner-{}".format(self._testMethodName),
            password="test-password",
        )

    def crea_contenitore(self, nome="Test Box", x=2000, y=2000, z=2000, peso=5000):
        return Contenitore.objects.create(
            owner=self.user,
            nome=nome,
            lunghezza_mm=x,
            larghezza_mm=y,
            altezza_mm=z,
            carico_massimo_kg=Decimal(str(peso)),
        )

    def crea_oggetto(self, codice, x=500, y=400, z=300, peso=20, qta=1):
        return Oggetto.objects.create(
            owner=self.user,
            codice=codice,
            lunghezza_mm=x,
            larghezza_mm=y,
            altezza_mm=z,
            peso_kg=Decimal(str(peso)),
            quantita_disponibile=qta,
        )

    def set_vincoli(self, oggetto, **kwargs):
        """Imposta i vincoli per un oggetto."""
        v, _ = VincoloOggetto.objects.get_or_create(oggetto=oggetto)
        for k, val in kwargs.items():
            setattr(v, k, val)
        v.save()
        return v

    def crea_vincolo_tra(self, a, b, tipo_relazione, distanza_cm=None, attivo=True, dettagli_posizionamento=None):
        """Crea un VincoloTraOggetti tra due oggetti."""
        vincolo = VincoloTraOggetti.objects.create(
            oggetto_a=a,
            oggetto_b=b,
            tipo_relazione=tipo_relazione,
            attivo=attivo,
        )
        if dettagli_posizionamento is not None:
            vincolo.dettagli_posizionamento = dettagli_posizionamento
            vincolo.save()
        return vincolo

    def crea_piano_e_ottimizza(self, contenitore, oggetti_con_quantita, config=None):
        """Crea un PianoDiCarico, aggiunge oggetti e ottimizza."""
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Test Piano",
            contenitore=contenitore,
        )
        for oggetto, qta in oggetti_con_quantita:
            OggettoDaCaricare.objects.create(
                piano_di_carico=piano,
                oggetto=oggetto,
                quantita=qta,
            )

        risultato = esegui_ottimizzazione_tre_d(piano.id, config=config)
        piano.refresh_from_db()
        return piano, risultato


# =============================================================================
# TEST: VINCOLI OGGETTO (VincoloOggetto)
# =============================================================================

class TestVincoliOggettoRotazione(TestCaseBase):
    """Test dei vincoli di rotazione su singolo oggetto."""

    def test_rotazione_consentita_false(self):
        """Se rotazione_consentita=False, l'oggetto mantiene solo orientamento XYZ."""
        contenitore = self.crea_contenitore()
        oggetto = self.crea_oggetto("NO-ROT", 400, 300, 200, peso=10)
        self.set_vincoli(oggetto, rotazione_consentita=False)

        piano, risultato = self.crea_piano_e_ottimizza(contenitore, [(oggetto, 1)])

        self.assertTrue(risultato.successo)
        posizionati = piano.oggetti_posizionati.all()
        self.assertEqual(posizionati.count(), 1)
        self.assertEqual(posizionati[0].rotazione_applicata, "XYZ")


class TestVincoliOggettoSovrapponibile(TestCaseBase):
    """Test del vincolo sovrapponibile e peso_massimo_tetto_kg."""

    def test_sovrapponibile_false_nessun_oggetto_sopra(self):
        """Se sovrapponibile=False, nessun oggetto può essere posizionato sopra."""
        contenitore = self.crea_contenitore()
        base = self.crea_oggetto("BASE", 1000, 1000, 500, peso=50)
        self.set_vincoli(base, sovrapponibile=False)

        sopra = self.crea_oggetto("SOPRA", 400, 400, 400, peso=10)

        piano, risultato = self.crea_piano_e_ottimizza(contenitore, [(base, 1), (sopra, 1)])

        self.assertTrue(risultato.successo)
        posizionati = list(piano.oggetti_posizionati.all())
        self.assertGreaterEqual(len(posizionati), 1)

        for p in posizionati:
            if p.oggetto_id == base.id:
                base_z_top = p.coordinata_z_mm + p.dimensione_z_mm
                for q in posizionati:
                    if q.id == p.id:
                        continue
                    if q.coordinata_z_mm >= base_z_top:
                        overlaps_x = not (
                            q.coordinata_x_mm >= p.coordinata_x_mm + p.dimensione_x_mm
                            or p.coordinata_x_mm >= q.coordinata_x_mm + q.dimensione_x_mm
                        )
                        overlaps_y = not (
                            q.coordinata_y_mm >= p.coordinata_y_mm + p.dimensione_y_mm
                            or p.coordinata_y_mm >= q.coordinata_y_mm + q.dimensione_y_mm
                        )
                        self.assertFalse(
                            overlaps_x and overlaps_y,
                            f"Oggetto {q.oggetto.codice} non dovrebbe stare sopra "
                            f"{p.oggetto.codice} (sovrapponibile=False)"
                        )

    def test_solo_su_piano_solo_z_zero(self):
        """solo_su_piano=True → oggetto posizionato solo a Z=0."""
        contenitore = self.crea_contenitore()
        piano_obj = self.crea_oggetto("PIANO-ONLY", 500, 500, 300, peso=30)
        self.set_vincoli(piano_obj, solo_su_piano=True)

        altro = self.crea_oggetto("ALTRO", 600, 600, 400, peso=20)

        piano, risultato = self.crea_piano_e_ottimizza(contenitore, [(piano_obj, 1), (altro, 1)])

        self.assertTrue(risultato.successo)
        posizionati = piano.oggetti_posizionati.all()

        for p in posizionati:
            if p.oggetto_id == piano_obj.id:
                self.assertEqual(p.coordinata_z_mm, 0,
                                 f"Oggetto solo_su_piano deve avere Z=0, ha Z={p.coordinata_z_mm}")


# =============================================================================
# TEST: ROTAZIONI IN TreDPacker (3D Semplificato)
# =============================================================================

class TestDeferimentoSingoliInterni(TestCase):
    """Verifica che un singolo interno liberi lo spazio per il ripacking."""

    def test_singolo_interno_viene_identificato_e_messo_in_coda(self):
        singolo = Obj("SINGOLO-0", 100, 100, 100, oggetto_id=1)
        altro = Obj("ALTRO-0", 100, 100, 100, oggetto_id=2)
        singolo.x, singolo.y, singolo.z = 0, 0, 0
        altro.x, altro.y, altro.z = 200, 0, 0

        self.assertEqual(
            _trova_singoli_interni([singolo, altro], 200, {}),
            ["SINGOLO-0"],
        )

        deferred = Obj("SINGOLO-0", 100, 100, 100, oggetto_id=1)
        deferred._peso_kg = 1
        placed = [copy.deepcopy(altro)]
        self.assertTrue(
            _piazza_deferiti_in_coda(
                placed,
                [deferred],
                (1000, 1000, 200),
                {},
            )
        )
        self.assertEqual(placed[-1].x, 200)

    def test_carico_reale_deferisce_il_singolo_i05_in_coda(self):
        """Il caso 20/8/13/19 non deve lasciare un singolo I05 interno."""
        dati = [
            ("CART-I03", 70, 160, 20, 3),
            ("CART-I01", 100, 120, 8, 1),
            ("CART-I02", 80, 120, 13, 2),
            ("cart-i05", 60, 80, 19, 5),
        ]
        objects = []
        for codice, width, depth, quantita, oggetto_id in dati:
            for indice in range(quantita):
                objects.append(
                    Obj(
                        f"{codice}-{indice}",
                        width,
                        depth,
                        100,
                        oggetto_id=oggetto_id,
                    )
                )

        risultato = load_truck_v2(
            objects,
            vincoli_sopra={},
            container_dim=(1360, 248, 270),
            preserve_order=True,
        )
        posizionati, non_posizionati = (
            [obj for obj in risultato if obj.z >= 0],
            [obj for obj in risultato if obj.z < 0],
        )

        self.assertEqual(len(posizionati), 60)
        self.assertEqual(non_posizionati, [])
        self.assertEqual(_trova_singoli_interni(posizionati, 270, {}), [])

        singolo_i05 = next(obj for obj in posizionati if obj.id == "cart-i05-18")
        max_x_iniziale = max(obj.x for obj in posizionati)
        self.assertEqual(
            singolo_i05.x,
            max_x_iniziale,
            "Il singolo I05 deve stare nella fascia X terminale, non nel mezzo.",
        )
        # Nel caso reale il singolo viene affiancato in Y alla fascia finale;
        # quindi non deve essere semplicemente accodato dopo la fine del
        # rettangolo più largo.
        self.assertEqual(singolo_i05.y, 0)

    def test_vincoli_tra_tutti_i_tipi_non_bloccano_il_deferimento(self):
        """Regressione: con vincoli 'sopra' su TUTTI i tipi (caso reale),
        il singolo interno deve essere comunque rilevato e deferito.

        In precedenza _trova_singoli_interni escludeva ogni oggetto
        coinvolto in vincoli tra oggetti: nel caso reale tutti i tipi
        hanno almeno un vincolo 'sopra', quindi il deferimento non
        partiva mai e il buco restava nel mezzo del carico.
        """
        dati = [
            ("CART-I03", 70, 160, 20, 3),
            ("CART-I01", 100, 120, 8, 1),
            ("CART-I02", 80, 120, 13, 2),
            ("cart-i05", 60, 80, 19, 5),
        ]
        objects = []
        for codice, width, depth, quantita, oggetto_id in dati:
            for indice in range(quantita):
                objects.append(
                    Obj(
                        f"{codice}-{indice}",
                        width,
                        depth,
                        100,
                        oggetto_id=oggetto_id,
                    )
                )

        # Come nel caso reale: ogni tipo è coinvolto in un vincolo 'sopra'
        # (auto-riferito con dettagli validi). Tutti gli oggetto_id quindi
        # finivano in ids_con_vincoli e il vecchio codice non deferiva nulla.
        # Struttura dettagli: set di coppie ((dimsA), (dimsB)).
        vincoli_sopra = {}
        for o_id, dims in [(3, (70, 160, 100)), (1, (100, 120, 100)),
                           (2, (80, 120, 100)), (5, (60, 80, 100))]:
            w, d, h = dims
            vincoli_sopra[o_id] = {
                o_id: {((w, d, h), (w, d, h)), ((d, w, h), (d, w, h))}
            }

        # Anche con vincoli non vuoti, il singolo deve essere rilevato
        singolo_test = Obj("S-0", 60, 80, 100, oggetto_id=5)
        singolo_test.x, singolo_test.y, singolo_test.z = 0, 0, 0
        altro_test = Obj("A-0", 100, 120, 100, oggetto_id=1)
        altro_test.x, altro_test.y, altro_test.z = 200, 0, 0
        self.assertNotEqual(
            _trova_singoli_interni(
                [singolo_test, altro_test],
                270,
                vincoli_sopra,
            ),
            [],
        )

        risultato = load_truck_v2(
            objects,
            vincoli_sopra=vincoli_sopra,
            container_dim=(1360, 248, 270),
            preserve_order=True,
        )
        posizionati = [obj for obj in risultato if obj.z >= 0]

        self.assertEqual(len(posizionati), 60)
        self.assertEqual(
            _trova_singoli_interni(posizionati, 270, vincoli_sopra),
            [],
            "Nessun singolo deve restare interno anche con vincoli su tutti i tipi",
        )

        singolo_i05 = next(obj for obj in posizionati if obj.id == "cart-i05-18")
        max_x_iniziale = max(obj.x for obj in posizionati)
        self.assertEqual(
            singolo_i05.x,
            max_x_iniziale,
            "Il singolo I05 deve stare nella fascia X terminale anche con vincoli.",
        )
        # Il deferimento non deve violare il vincolo 'sopra': il singolo
        # resta a pavimento (z=0) nella coda.
        self.assertEqual(singolo_i05.z, 0)

    def test_deferimento_sequenziale_isolato_riempie_con_codice_successivo(self):
        """Un isolato viene accodato dopo aver ricomposto il suffisso."""
        oggetti = [
            Obj("A-0", 100, 100, 100, oggetto_id=1, sovrapponibile=False),
            Obj("B-0", 100, 100, 100, oggetto_id=2),
            Obj("B-1", 100, 100, 100, oggetto_id=2),
        ]
        risultato = load_truck_v2(
            oggetti,
            vincoli_sopra={},
            container_dim=(500, 100, 200),
            preserve_order=True,
        )
        posizionati = [obj for obj in risultato if obj.z >= 0]
        isolato = next(obj for obj in posizionati if obj.id == "A-0")
        successivi = [obj for obj in posizionati if obj.oggetto_id == 2]

        self.assertEqual(len(posizionati), 3)
        self.assertGreaterEqual(
            isolato.x,
            max(obj.x for obj in successivi),
            "L'isolato deve essere rimesso dopo il codice che ha riempito il buco.",
        )
        self.assertEqual(max(obj.x + obj.width for obj in posizionati), 200)

    def test_deferimento_fallito_continua_con_il_singolo_successivo(self):
        """Il rollback di A non deve impedire di tentare B."""
        a = Obj("A-0", 100, 100, 100, oggetto_id=1)
        b = Obj("B-0", 100, 100, 100, oggetto_id=2)
        c = Obj("C-0", 100, 100, 100, oggetto_id=3)
        a.x, a.y, a.z = 0, 0, 0
        b.x, b.y, b.z = 200, 0, 0
        c.x, c.y, c.z = 400, 0, 0
        placed = [a, b, c]

        def try_orientations(obj, x, y, z, current, container, constraints, **kwargs):
            if obj.id == "A-0" or z > 0:
                return False
            if any(
                other.z == 0
                and other.x < x + obj.width
                and x < other.x + other.width
                and other.y < y + obj.depth
                and y < other.y + other.depth
                for other in current
            ):
                return False
            obj.x, obj.y, obj.z = x, y, z
            return True

        def columns_info(current):
            return {
                (obj.x, obj.y): {
                    "z_top": obj.z + obj.height,
                    "top_item": obj,
                }
                for obj in current
            }

        def y_candidates(current, try_x, container_d):
            return [0]

        defer_singles(
            placed,
            [a, b, c],
            (600, 100, 200),
            {},
            None,
            try_orientations,
            columns_info,
            y_candidates,
            lambda current, obj: True,
        )

        self.assertEqual({obj.id for obj in placed}, {"A-0", "B-0", "C-0"})
        self.assertEqual(a.x, 0)
        self.assertEqual(len({obj.id for obj in placed}), 3)


class TestRotazioniTreDPacker(TestCase):
    """Test per verificare che _prova_tutte_orientazioni rispetti
    i flag individuali rotazione_su_x/y/z in TreDPacker."""

    def setUp(self):
        self.obj_default = Obj("TEST", 120, 80, 50, oggetto_id=1)
        self.big_dim = (500, 500, 500)
        self.narrow_x = (70, 200, 200)
        self.narrow_y = (200, 60, 200)

    def _permutationi_attese(self, obj):
        w, d, h = obj.width, obj.depth, obj.height
        perms = [(w, d, h)]
        if obj.rotazione_su_z:
            perms.append((d, w, h))
        if obj.rotazione_su_x:
            perms.append((w, h, d))
        if obj.rotazione_su_y:
            perms.append((h, d, w))
        if obj.rotazione_su_x and obj.rotazione_su_y and obj.rotazione_su_z:
            perms.append((d, h, w))
            perms.append((h, w, d))
        return sorted(set(perms), key=lambda p: p[0])

    def _chiama_e_verifica(self, obj, container_dim, deve_riuscire, dim_attese=None):
        risultato = _prova_tutte_orientazioni(obj, 0, 0, 0, [], container_dim, {})
        self.assertEqual(risultato, deve_riuscire)
        if deve_riuscire and dim_attese:
            self.assertEqual((obj.width, obj.depth, obj.height), dim_attese)

    def test_rotazione_non_consentita(self):
        obj = Obj("TEST", 120, 80, 50, oggetto_id=1, orientation_allowed=False)
        self._chiama_e_verifica(obj, self.narrow_x, deve_riuscire=False)
        self._chiama_e_verifica(obj, self.big_dim, deve_riuscire=True, dim_attese=(120, 80, 50))

    def test_solo_rotazione_su_z(self):
        obj = Obj("TEST", 120, 80, 50, oggetto_id=1,
                  rotazione_su_x=False, rotazione_su_y=False, rotazione_su_z=True)
        self._chiama_e_verifica(obj, self.narrow_x, deve_riuscire=False)
        obj2 = Obj("TEST", 120, 80, 50, oggetto_id=1,
                   rotazione_su_x=False, rotazione_su_y=False, rotazione_su_z=True)
        self._chiama_e_verifica(obj2, self.big_dim, deve_riuscire=True, dim_attese=(80, 120, 50))

    def test_solo_rotazione_su_x(self):
        obj = Obj("TEST", 80, 120, 50, oggetto_id=1,
                  rotazione_su_x=True, rotazione_su_y=False, rotazione_su_z=False)
        self._chiama_e_verifica(obj, self.narrow_y, deve_riuscire=True, dim_attese=(80, 50, 120))

    def test_solo_rotazione_su_y(self):
        obj = Obj("TEST", 120, 80, 50, oggetto_id=1,
                  rotazione_su_x=False, rotazione_su_y=True, rotazione_su_z=False)
        self._chiama_e_verifica(obj, self.narrow_x, deve_riuscire=True, dim_attese=(50, 80, 120))

    def test_tutte_sei_permutazioni(self):
        obj = Obj("TEST", 120, 80, 50, oggetto_id=1,
                  rotazione_su_x=True, rotazione_su_y=True, rotazione_su_z=True)
        self._chiama_e_verifica(obj, self.narrow_x, deve_riuscire=True, dim_attese=(50, 80, 120))

    def test_dimensioni_uguali_dedup(self):
        obj = Obj("CUBO", 100, 100, 80, oggetto_id=1,
                  rotazione_su_x=True, rotazione_su_y=True, rotazione_su_z=True)
        perms = self._permutationi_attese(obj)
        self.assertEqual(len(perms), 3)

    def test_cubo_tutte_uguali(self):
        obj = Obj("CUBO", 100, 100, 100, oggetto_id=1,
                  rotazione_su_x=True, rotazione_su_y=True, rotazione_su_z=True)
        perms = self._permutationi_attese(obj)
        self.assertEqual(len(perms), 1)
        self.assertEqual(perms[0], (100, 100, 100))


# =============================================================================
# TEST: VINCOLI TRA OGGETTI (VincoloTraOggetti)
# =============================================================================

class TestVincoloTraOggettiSopra(TestCaseBase):
    """Test del vincolo SOPRA tra oggetti."""

    def test_sopra_a_deve_stare_sopra_b(self):
        contenitore = self.crea_contenitore()
        a = self.crea_oggetto("A-SOPRA", 500, 500, 300, peso=10)
        b = self.crea_oggetto("B-SOTTO", 800, 800, 200, peso=20)
        self.crea_vincolo_tra(a, b, "sopra")

        piano, risultato = self.crea_piano_e_ottimizza(contenitore, [(a, 1), (b, 1)])

        self.assertTrue(risultato.successo)
        posizionati = list(piano.oggetti_posizionati.all())
        self.assertEqual(len(posizionati), 2)

        a_pos = next(p for p in posizionati if p.oggetto_id == a.id)
        b_pos = next(p for p in posizionati if p.oggetto_id == b.id)

        b_top = b_pos.coordinata_z_mm + b_pos.dimensione_z_mm
        self.assertGreaterEqual(a_pos.coordinata_z_mm, b_top)
        overlaps_x = not (
            a_pos.coordinata_x_mm >= b_pos.coordinata_x_mm + b_pos.dimensione_x_mm
            or b_pos.coordinata_x_mm >= a_pos.coordinata_x_mm + a_pos.dimensione_x_mm
        )
        overlaps_y = not (
            a_pos.coordinata_y_mm >= b_pos.coordinata_y_mm + b_pos.dimensione_y_mm
            or b_pos.coordinata_y_mm >= a_pos.coordinata_y_mm + a_pos.dimensione_y_mm
        )
        self.assertTrue(overlaps_x and overlaps_y)

    def test_vincolo_sopra_deroga_area_e_sbalzo(self):
        """Un vincolo valido consente A sopra B anche se A ha area maggiore."""
        base = Obj("B-0", 120, 80, 100, oggetto_id=2)
        sopra = Obj("A-0", 120, 100, 100, oggetto_id=1)
        base.x, base.y, base.z = 0, 0, 0

        self.assertTrue(
            _prova_tutte_orientazioni(
                sopra,
                0,
                0,
                100,
                [base],
                (500, 500, 500),
                {1: {2: None}},
            ),
            "Il vincolo A sopra B deve prevalere sulla regola dell'area.",
        )

    def test_configurazione_esclusa_vieta_stacking_diretto(self):
        """Una configurazione esclusa non può ricadere nella regola area."""
        base = Obj("I03-0", 70, 160, 100, oggetto_id=3)
        sopra = Obj("I02-0", 80, 120, 100, oggetto_id=2)
        base.x, base.y, base.z = 0, 0, 0

        vincoli = {2: {3: set()}}
        self.assertFalse(
            _prova_tutte_orientazioni(
                sopra,
                0,
                0,
                100,
                [base],
                (500, 500, 500),
                vincoli,
            ),
            "Una coppia esplicitamente esclusa deve essere vietata.",
        )
        from caricamento.engine.tre_d.priority_policy import valida_vincoli_sopra
        sopra.x, sopra.y, sopra.z = 0, 0, 100
        self.assertFalse(
            valida_vincoli_sopra([sopra, base], [sopra, base], vincoli)["vincoli_completi"]
        )

    def test_configurazione_valida_e_esclusiva(self):
        """Con un vincolo dimensionale si accetta solo la combinazione valida."""
        base = Obj("I03-0", 70, 160, 100, oggetto_id=3)
        sopra = Obj("I02-0", 80, 120, 100, oggetto_id=2)
        base.x, base.y, base.z = 0, 0, 0
        vincoli = {
            2: {
                3: {
                    ((80, 120, 100), (70, 160, 100)),
                },
            },
        }

        self.assertTrue(
            _prova_tutte_orientazioni(
                sopra, 0, 0, 100, [base],
                (500, 500, 500), vincoli,
            )
        )

        altra_base = Obj("I03-1", 80, 120, 100, oggetto_id=3)
        altra_base.x, altra_base.y, altra_base.z = 0, 0, 0
        altro_sopra = Obj("I02-1", 70, 160, 100, oggetto_id=2)
        self.assertFalse(
            _prova_tutte_orientazioni(
                altro_sopra, 0, 0, 100, [altra_base],
                (500, 500, 500), vincoli,
            ),
            "Una dimensione non elencata nel vincolo non è autorizzata.",
        )
        from caricamento.engine.tre_d.priority_policy import valida_vincoli_sopra
        altro_sopra.x, altro_sopra.y, altro_sopra.z = 0, 0, 100
        self.assertFalse(
            valida_vincoli_sopra(
                [altro_sopra, altra_base],
                [altro_sopra, altra_base],
                vincoli,
            )["vincoli_completi"]
        )

    def test_assenza_vincolo_usa_regole_standard(self):
        """Senza relazione tra i codici resta attiva la regola dell'area."""
        base = Obj("I03-0", 70, 160, 100, oggetto_id=3)
        sopra = Obj("I02-0", 80, 120, 100, oggetto_id=2)
        base.x, base.y, base.z = 0, 0, 0

        self.assertTrue(
            _prova_tutte_orientazioni(
                sopra, 0, 0, 100, [base],
                (500, 500, 500), {},
            ),
            "In assenza di vincolo devono valere le regole geometriche standard.",
        )

    def test_z_positiva_senza_supporto_viene_rifiutata(self):
        """Nessuna modalità può accettare un oggetto volante."""
        volante = Obj("VOLANTE-0", 100, 100, 100, oggetto_id=1)

        for aggressiva in (False, True):
            self.assertFalse(
                _prova_tutte_orientazioni(
                    volante,
                    0,
                    0,
                    100,
                    [],
                    (500, 500, 500),
                    {},
                    compattazione_aggressiva=aggressiva,
                ),
                f"Un oggetto senza supporto non può essere accettato in modalità aggressiva={aggressiva}",
            )

    def test_compattazione_aggressiva_consente_stacking_traslato_con_80_percento(self):
        """Un vincolo A sopra A consente lo scostamento solo in aggressiva.

        Il caso riproduce n.11 sopra n.7: la traslazione lascia l'80% della
        base di A appoggiata su B e non collide con l'oggetto adiacente.
        """
        adiacente = Obj("I01-4", 100, 120, 100, oggetto_id=1)
        base = Obj("I01-7", 100, 120, 100, oggetto_id=1)
        sopra = Obj("I01-11", 100, 120, 100, oggetto_id=1)
        adiacente.x, adiacente.y, adiacente.z = 0, 0, 100
        base.x, base.y, base.z = 80, 0, 0

        vincoli = {1: {1: None}}
        self.assertFalse(
            _prova_tutte_orientazioni(
                sopra, 100, 0, 100, [adiacente, base],
                (500, 500, 500), vincoli,
                compattazione_aggressiva=False,
            )
        )
        self.assertTrue(
            _prova_tutte_orientazioni(
                sopra, 100, 0, 100, [adiacente, base],
                (500, 500, 500), vincoli,
                compattazione_aggressiva=True,
            )
        )
        self.assertEqual((sopra.x, sopra.y, sopra.z), (100, 0, 100))

    def test_compattazione_aggressiva_rifiuta_contatto_inferiore_al_50_percento(self):
        """La modalità aggressiva non accetta un appoggio relazionale del 49%."""
        base = Obj("I01-7", 100, 120, 100, oggetto_id=1)
        sopra = Obj("I01-11", 100, 120, 100, oggetto_id=1)
        base.x, base.y, base.z = 49, 0, 0

        self.assertFalse(
            _prova_tutte_orientazioni(
                sopra, 100, 0, 100, [base],
                (500, 500, 500), {1: {1: None}},
                compattazione_aggressiva=True,
            )
        )

    def test_stacking_senza_vincolo_rispetta_area(self):
        """Senza vincolo, un oggetto con area maggiore non può stare sopra."""
        base = Obj("B-0", 120, 80, 100, oggetto_id=2)
        sopra = Obj("A-0", 120, 100, 100, oggetto_id=1)
        base.x, base.y, base.z = 0, 0, 0

        self.assertFalse(
            _prova_tutte_orientazioni(
                sopra,
                0,
                0,
                100,
                [base],
                (500, 500, 500),
                {},
            )
        )


# =============================================================================
# TEST: FUNZIONI HELPER
# =============================================================================

class TestConfigurazioneOttimizzazione(TestCaseBase):
    """Test per la configurazione dell'ottimizzatore."""

    def test_from_dict_con_valori_parziali(self):
        config = ConfigurazioneOttimizzazione.from_dict({
            "strategia_ottimizzazione": {
                "ordinamento_casuale": True,
            },
        })
        self.assertTrue(config.ordinamento_casuale)
        self.assertEqual(config.algoritmo_base, "Algoritmo 3D Semplificato")
        # La distribuzione pesi è disattivata di default perché il vincolo
        # sulle sezioni può impedire il piazzamento di oggetti che invece
        # entrerebbero nel contenitore.
        self.assertFalse(config.distribuzione_pesi_attiva)

    def test_from_dict_con_config_vuota(self):
        config = ConfigurazioneOttimizzazione.from_dict({})
        self.assertEqual(config.algoritmo_base, "Algoritmo 3D Semplificato")
        self.assertFalse(config.distribuzione_pesi_attiva)

    def test_configurazione_personalizzata(self):
        contenitore = self.crea_contenitore()
        oggetto = self.crea_oggetto("CFG-TEST", 500, 400, 300, peso=10)

        config = ConfigurazioneOttimizzazione(
            ordinamento_casuale=True,
        )
        piano, risultato = self.crea_piano_e_ottimizza(contenitore, [(oggetto, 1)], config=config)
        self.assertTrue(risultato.successo)

    def test_quantita_richiesta_superiore_alle_istanze_posizionate_rende_parziale(self):
        """Con quantita=2, una sola istanza salvata è un piano parziale."""
        contenitore = self.crea_contenitore()
        oggetto = self.crea_oggetto("REPORT-PARZIALE", 500, 400, 300, peso=10)
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Quantita parziale",
            contenitore=contenitore,
        )
        OggettoDaCaricare.objects.create(
            piano_di_carico=piano,
            oggetto=oggetto,
            quantita=2,
        )

        item = ItemPacked(
            oggetto_id=oggetto.id,
            codice=oggetto.codice,
            coordinata_x_mm=0,
            coordinata_y_mm=0,
            coordinata_z_mm=0,
            dimensione_x_mm=500,
            dimensione_y_mm=400,
            dimensione_z_mm=300,
            rotazione_applicata="XYZ",
            peso_kg=Decimal("10"),
            colore="#4488ff",
        )
        packer = Mock()
        packer.results = [item]
        packer.unfitted_codes = [f"{oggetto.codice}-1"]
        packer.priority_report = {
            "priorita_completa": True,
            "vincoli": {"vincoli_completi": True},
        }
        packer.genera_metriche.return_value = {}

        with patch(
            "caricamento.engine.orchestratore_tre_d.TreDPacker",
            return_value=packer,
        ):
            risultato = esegui_ottimizzazione_tre_d(piano.id)

        piano.refresh_from_db()
        self.assertFalse(risultato.successo)
        self.assertEqual(piano.stato, "parziale")
        self.assertEqual(
            risultato.messaggio,
            "Parziale: 1 di 2 oggetti posizionati: "
            "i rimanenti non trovano spazio nel contenitore.",
        )

    def test_report_incompleto_non_rende_parziale_un_piano_completo(self):
        """Lo stato dipende dalle istanze, non dai report diagnostici."""
        contenitore = self.crea_contenitore()
        oggetto = self.crea_oggetto("REPORT-COMPLETO", 500, 400, 300, peso=10)
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Report completo",
            contenitore=contenitore,
        )
        OggettoDaCaricare.objects.create(
            piano_di_carico=piano,
            oggetto=oggetto,
            quantita=1,
        )

        item = ItemPacked(
            oggetto_id=oggetto.id,
            codice=oggetto.codice,
            coordinata_x_mm=0,
            coordinata_y_mm=0,
            coordinata_z_mm=0,
            dimensione_x_mm=500,
            dimensione_y_mm=400,
            dimensione_z_mm=300,
            rotazione_applicata="XYZ",
            peso_kg=Decimal("10"),
            colore="#4488ff",
        )
        packer = Mock()
        packer.results = [item]
        packer.unfitted_codes = []
        packer.priority_report = {
            "priorita_completa": False,
            "vincoli": {"vincoli_completi": False},
        }
        packer.genera_metriche.return_value = {}

        with patch(
            "caricamento.engine.orchestratore_tre_d.TreDPacker",
            return_value=packer,
        ):
            risultato = esegui_ottimizzazione_tre_d(piano.id)

        piano.refresh_from_db()
        self.assertTrue(risultato.successo)
        self.assertEqual(piano.stato, "completato")
        self.assertEqual(
            risultato.messaggio,
            "Completato: 1 oggetti posizionati.",
        )


class TestBuildLookupVincoliTra(TestCaseBase):
    """Test per _build_lookup_vincoli_tra."""

    def test_lookup_filtra_entrambi_gli_estremi_presenti(self):
        a = self.crea_oggetto("L-F-A")
        b = self.crea_oggetto("L-F-B")
        c = self.crea_oggetto("L-F-C")
        vincolo_valido = self.crea_vincolo_tra(a, b, "sopra")
        vincolo_esterno = self.crea_vincolo_tra(a, c, "sopra")

        queryset = VincoloTraOggetti.objects.filter(
            id__in=[vincolo_valido.id, vincolo_esterno.id]
        )
        lookup = _build_lookup_vincoli_tra(queryset, oggetto_ids={a.id, b.id})

        self.assertEqual([entry[0] for entry in lookup[a.id]], [b.id])

    def test_lookup_direzionale_per_sopra(self):
        a = self.crea_oggetto("L-A")
        b = self.crea_oggetto("L-B")
        self.crea_vincolo_tra(a, b, "sopra")

        lookup = _build_lookup_vincoli_tra(VincoloTraOggetti.objects.filter(attivo=True))

        targets_a = [t[0] for t in lookup.get(a.id, [])]
        self.assertIn(b.id, targets_a)
        self.assertNotIn(a.id, [t[0] for t in lookup.get(b.id, [])])

    def test_carica_vincoli_sopra_filtra_oggetti_fuori_piano(self):
        contenitore = self.crea_contenitore()
        a = self.crea_oggetto("L-PIANO-A")
        b = self.crea_oggetto("L-PIANO-B")
        c = self.crea_oggetto("L-FUORI")
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Piano filtro vincoli",
            contenitore=contenitore,
        )
        OggettoDaCaricare.objects.create(piano_di_carico=piano, oggetto=a)
        OggettoDaCaricare.objects.create(piano_di_carico=piano, oggetto=b)
        self.crea_vincolo_tra(a, b, "sopra")
        self.crea_vincolo_tra(a, c, "sopra")

        lookup = _carica_vincoli_sopra(piano)

        self.assertEqual(list(lookup[a.id]), [b.id])

    def test_lookup_solo_vincoli_attivi(self):
        a = self.crea_oggetto("L-C")
        b = self.crea_oggetto("L-D")
        self.crea_vincolo_tra(a, b, "sopra", attivo=False)

        lookup = _build_lookup_vincoli_tra(VincoloTraOggetti.objects.filter(attivo=True))
        self.assertEqual(len(lookup), 0)


class TestOrigineSalvataggioPosizioni(TestCaseBase):
    """La sincronizzazione della lista non deve falsare l'origine del piano."""

    def _salva_posizioni(self, piano, origine):
        client = APIClient()
        client.force_authenticate(user=self.user)
        return client.post(
            "/api/piani/{}/salva_posizioni_manuali/".format(piano.id),
            {
                "origine": origine,
                "oggetti": [{
                    "oggetto_id": piano.oggetti_da_caricare.first().oggetto_id,
                    "codice": piano.oggetti_da_caricare.first().oggetto.codice,
                    "posizione_cm": {"x": 0, "y": 0, "z": 0},
                    "dimensioni_cm": {"x": 50, "y": 40, "z": 30},
                    "colore": "#4488ff",
                    "rotazione": "XYZ",
                }],
            },
            format="json",
        )

    def test_sincronizzazione_non_sovrascrive_algoritmo_automatico(self):
        contenitore = self.crea_contenitore()
        oggetto = self.crea_oggetto("ORIG-A", 500, 400, 300)
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Piano automatico",
            contenitore=contenitore,
            algoritmo="Algoritmo 3D Semplificato",
        )
        OggettoDaCaricare.objects.create(piano_di_carico=piano, oggetto=oggetto)

        response = self._salva_posizioni(piano, "sincronizzazione")

        self.assertEqual(response.status_code, 200)
        piano.refresh_from_db()
        self.assertEqual(piano.algoritmo, "Algoritmo 3D Semplificato")

    def test_sincronizzazione_preserva_piano_parziale(self):
        contenitore = self.crea_contenitore()
        oggetto = self.crea_oggetto("ORIG-P", 500, 400, 300)
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Piano parziale",
            contenitore=contenitore,
            stato="parziale",
            algoritmo="Algoritmo 3D Semplificato",
        )
        OggettoDaCaricare.objects.create(piano_di_carico=piano, oggetto=oggetto)

        response = self._salva_posizioni(piano, "sincronizzazione")

        self.assertEqual(response.status_code, 200)
        piano.refresh_from_db()
        self.assertEqual(piano.stato, "parziale")
        self.assertEqual(piano.algoritmo, "Algoritmo 3D Semplificato")

    def test_salvataggio_manuale_marca_piano_manuale(self):
        contenitore = self.crea_contenitore()
        oggetto = self.crea_oggetto("ORIG-M", 500, 400, 300)
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Piano da modificare",
            contenitore=contenitore,
            algoritmo="Algoritmo 3D Semplificato",
        )
        OggettoDaCaricare.objects.create(piano_di_carico=piano, oggetto=oggetto)

        response = self._salva_posizioni(piano, "manuale")

        self.assertEqual(response.status_code, 200)
        piano.refresh_from_db()
        self.assertEqual(piano.algoritmo, "manuale")

    def test_alternativa_dopo_manuale_riporta_piano_ad_automatico(self):
        """Flusso utente: salvo a mano (piano = "manuale"), poi applico una
        soluzione alternativa (sincronizzazione con algoritmo): il piano deve
        tornare ad essere automatico, non restare etichettato "manuale"."""
        contenitore = self.crea_contenitore()
        oggetto = self.crea_oggetto("ORIG-ALT", 500, 400, 300)
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Piano alt dopo manuale",
            contenitore=contenitore,
            algoritmo="Algoritmo 3D Semplificato 🎲 Monte Carlo",
        )
        OggettoDaCaricare.objects.create(piano_di_carico=piano, oggetto=oggetto)

        # 1) Salvataggio manuale: il piano diventa "manuale"
        response = self._salva_posizioni(piano, "manuale")
        self.assertEqual(response.status_code, 200)
        piano.refresh_from_db()
        self.assertEqual(piano.algoritmo, "manuale")

        # 2) Applicazione alternativa: deve riportare l'algoritmo automatico
        client = APIClient()
        client.force_authenticate(user=self.user)
        response = client.post(
            "/api/piani/{}/salva_posizioni_manuali/".format(piano.id),
            {
                "origine": "sincronizzazione",
                "algoritmo": "Algoritmo 3D Semplificato 🎲 Monte Carlo",
                "oggetti": [{
                    "oggetto_id": oggetto.id,
                    "codice": oggetto.codice,
                    "posizione_cm": {"x": 0, "y": 0, "z": 0},
                    "dimensioni_cm": {"x": 50, "y": 40, "z": 30},
                    "colore": "#4488ff",
                    "rotazione": "XYZ",
                }],
            },
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        piano.refresh_from_db()
        self.assertEqual(piano.algoritmo, "Algoritmo 3D Semplificato 🎲 Monte Carlo")

    def test_alternativa_senza_algoritmo_non_altere_etichetta(self):
        """Se il payload di sincronizzazione non include algoritmo (client
        legacy), l'etichetta del piano non viene toccata."""
        contenitore = self.crea_contenitore()
        oggetto = self.crea_oggetto("ORIG-ALT2", 500, 400, 300)
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Piano alt legacy",
            contenitore=contenitore,
            algoritmo="Algoritmo 3D Semplificato",
        )
        OggettoDaCaricare.objects.create(piano_di_carico=piano, oggetto=oggetto)

        response = self._salva_posizioni(piano, "sincronizzazione")

        self.assertEqual(response.status_code, 200)
        piano.refresh_from_db()
        self.assertEqual(piano.algoritmo, "Algoritmo 3D Semplificato")

    def test_salva_persiste_tutte_le_posizioni_visibili(self):
        """La quantità richiesta non elimina posizioni già visibili nella scena."""
        contenitore = self.crea_contenitore(x=2000, y=2000, z=2000)
        oggetto = self.crea_oggetto("VISIBLE-A", 500, 400, 300)
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Piano scena visibile",
            contenitore=contenitore,
            stato="parziale",
            algoritmo="Algoritmo 3D Semplificato",
        )
        # Una sola unità richiesta, ma due istanze sono già presenti nella
        # scena visualizzata. Entrambe sono valide e devono essere persistite.
        OggettoDaCaricare.objects.create(
            piano_di_carico=piano,
            oggetto=oggetto,
            quantita=1,
        )

        response = self._salva_posizioni_payload(piano, [
            {
                "oggetto_id": oggetto.id,
                "codice": oggetto.codice,
                "posizione_cm": {"x": 0, "y": 0, "z": 0},
                "dimensioni_cm": {"x": 50, "y": 40, "z": 30},
            },
            {
                "oggetto_id": oggetto.id,
                "codice": oggetto.codice,
                "posizione_cm": {"x": 100, "y": 100, "z": 0},
                "dimensioni_cm": {"x": 50, "y": 40, "z": 30},
            },
        ], "sincronizzazione")

        self.assertEqual(response.status_code, 200)
        posizioni = list(piano.oggetti_posizionati.order_by("coordinata_x_mm"))
        self.assertEqual(len(posizioni), 2)
        self.assertEqual(
            [(p.coordinata_x_mm, p.coordinata_y_mm, p.coordinata_z_mm)
             for p in posizioni],
            [(0, 0, 0), (1000, 1000, 0)],
        )
        self.assertEqual(
            [(p.dimensione_x_mm, p.dimensione_y_mm, p.dimensione_z_mm)
             for p in posizioni],
            [(500, 400, 300), (500, 400, 300)],
        )

    def _salva_posizioni_payload(self, piano, oggetti, origine):
        client = APIClient()
        client.force_authenticate(user=self.user)
        return client.post(
            "/api/piani/{}/salva_posizioni_manuali/".format(piano.id),
            {"origine": origine, "oggetti": oggetti},
            format="json",
        )


    def test_fallback_obsolete_riga_id_links_to_correct_row(self):
        """Quando il riga_id è obsoleto (piano tecnico eliminato),
        il backend dovrebbe trovare la riga per oggetto_id e collegarla."""
        contenitore = self.crea_contenitore()
        oggetto = self.crea_oggetto("FALL-A", 500, 400, 300)
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Piano fallback",
            contenitore=contenitore,
        )
        riga = OggettoDaCaricare.objects.create(
            piano_di_carico=piano, oggetto=oggetto, quantita=2,
            colore="#FF0000",
        )

        response = self._salva_posizioni_payload(piano, [{
            "oggetto_id": oggetto.id,
            "codice": oggetto.codice,
            "posizione_cm": {"x": 0, "y": 0, "z": 0},
            "dimensioni_cm": {"x": 50, "y": 40, "z": 30},
            "colore": "#aaaaaa",
            "rotazione": "XYZ",
            # riga_id obsoleta (ID 99999 non esiste nel piano)
            "riga_id": 99999,
        }], "sincronizzazione")

        self.assertEqual(response.status_code, 200)
        op = piano.oggetti_posizionati.first()
        self.assertIsNotNone(op)
        # Il fallback dovrebbe aver trovato la riga per oggetto_id
        self.assertEqual(op.riga_origine_id, riga.id)
        # Il colore dovrebbe essere quello della riga, non quello dello snapshot
        self.assertEqual(op.colore, "#FF0000")

    def test_obsolete_riga_id_uses_row_color_not_snapshot_color(self):
        """Il colore della riga ha precedenza su quello dello snapshot
        quando il riga_id è obsoleto ma la riga viene trovata."""
        contenitore = self.crea_contenitore()
        oggetto = self.crea_oggetto("FALL-B", 600, 500, 400)
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Piano fallback colore",
            contenitore=contenitore,
        )
        riga = OggettoDaCaricare.objects.create(
            piano_di_carico=piano, oggetto=oggetto, quantita=1,
            colore="#00FF00",
        )

        response = self._salva_posizioni_payload(piano, [{
            "oggetto_id": oggetto.id,
            "codice": oggetto.codice,
            "posizione_cm": {"x": 0, "y": 0, "z": 0},
            "dimensioni_cm": {"x": 60, "y": 50, "z": 40},
            "colore": "#000000",  # colore dello snapshot (sbagliato)
            "rotazione": "XYZ",
            "riga_id": 99998,  # obsoleta
        }], "manuale")

        self.assertEqual(response.status_code, 200)
        op = piano.oggetti_posizionati.first()
        self.assertIsNotNone(op)
        self.assertEqual(op.colore, "#00FF00")  # colore dalla riga
        self.assertEqual(op.riga_origine_id, riga.id)

    def test_fallback_does_not_link_two_positions_to_same_riga(self):
        """Quando 2 posizionamenti dello stesso oggetto hanno riga_id
        obsoleti, il fallback dovrebbe collegarne uno alla riga 1 e
        l'altro alla riga 2 (se disponibile), non entrambi alla prima."""
        contenitore = self.crea_contenitore()
        oggetto = self.crea_oggetto("FALL-DUP", 500, 400, 300)
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Piano fallback dup",
            contenitore=contenitore,
        )
        riga1 = OggettoDaCaricare.objects.create(
            piano_di_carico=piano, oggetto=oggetto, quantita=1,
            colore="#FF0000",
        )
        riga2 = OggettoDaCaricare.objects.create(
            piano_di_carico=piano, oggetto=oggetto, quantita=1,
            colore="#0000FF",
        )

        # Due posizionamenti dello stesso oggetto con riga_id obsoleti
        response = self._salva_posizioni_payload(piano, [
            {
                "oggetto_id": oggetto.id,
                "codice": oggetto.codice,
                "posizione_cm": {"x": 0, "y": 0, "z": 0},
                "dimensioni_cm": {"x": 50, "y": 40, "z": 30},
                "colore": "#aaaaaa",
                "rotazione": "XYZ",
                "riga_id": 99991,  # obsoleta
            },
            {
                "oggetto_id": oggetto.id,
                "codice": oggetto.codice,
                "posizione_cm": {"x": 60, "y": 0, "z": 0},
                "dimensioni_cm": {"x": 50, "y": 40, "z": 30},
                "colore": "#bbbbbb",
                "rotazione": "XYZ",
                "riga_id": 99992,  # obsoleta
            },
        ], "sincronizzazione")

        self.assertEqual(response.status_code, 200)
        ops = list(piano.oggetti_posizionati.order_by('id'))
        self.assertEqual(len(ops), 2)
        # I due posizionamenti dovrebbero avere righe diverse
        righe_ids = {op.riga_origine_id for op in ops if op.riga_origine_id}
        self.assertEqual(len(righe_ids), 2,
            "Entrambi i posizionamenti puntano alla stessa riga!")
        # E colori diversi
        colori = {op.colore for op in ops}
        self.assertEqual(len(colori), 2,
            "I due posizionamenti hanno lo stesso colore!")

    def test_payload_che_viola_la_barriera_viene_corretto_in_coda(self):
        """Un payload (es. soluzione alternativa generata prima del fix) che
        mette un oggetto a priorità 0 dentro il blocco dei prioritari viene
        ripiazzato in coda: la priorità ha la precedenza su tutto."""
        contenitore = self.crea_contenitore(x=2000, y=2000, z=2000)
        oggetto = self.crea_oggetto("BARR-A", 500, 400, 300)
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Piano barriera",
            contenitore=contenitore,
        )
        riga_prio1 = OggettoDaCaricare.objects.create(
            piano_di_carico=piano, oggetto=oggetto, quantita=1,
            colore="#FF0000", priorita=1,
        )
        riga_prio0 = OggettoDaCaricare.objects.create(
            piano_di_carico=piano, oggetto=oggetto, quantita=1,
            colore="#00FF00", priorita=0,
        )

        # Il payload viola la barriera: il prio 0 è a X=0 (davanti al prio 1
        # che occupa X=0-500). Il backend deve spostarlo dopo la fine X dei
        # prioritari (X >= 500).
        response = self._salva_posizioni_payload(piano, [
            {
                "oggetto_id": oggetto.id,
                "codice": oggetto.codice,
                "posizione_cm": {"x": 0, "y": 0, "z": 0},
                "dimensioni_cm": {"x": 50, "y": 40, "z": 30},
                "colore": "#aaaaaa",
                "rotazione": "XYZ",
                "riga_id": riga_prio1.id,
            },
            {
                "oggetto_id": oggetto.id,
                "codice": oggetto.codice,
                "posizione_cm": {"x": 0, "y": 0, "z": 0},  # VIOLA!
                "dimensioni_cm": {"x": 50, "y": 40, "z": 30},
                "colore": "#bbbbbb",
                "rotazione": "XYZ",
                "riga_id": riga_prio0.id,
            },
        ], "sincronizzazione")

        self.assertEqual(response.status_code, 200)
        op_prio0 = piano.oggetti_posizionati.get(riga_origine=riga_prio0)
        op_prio1 = piano.oggetti_posizionati.get(riga_origine=riga_prio1)
        barriera = op_prio1.coordinata_x_mm + op_prio1.dimensione_x_mm
        self.assertGreaterEqual(
            op_prio0.coordinata_x_mm,
            barriera,
            "Il prio 0 è rimasto dentro il blocco dei prioritari!",
        )

    def test_salvataggio_manuale_non_bloccato_dalla_barriera(self):
        """Nel flusso MANUALE l'utente ha la precedenza: se sposta a mano un
        oggetto a priorità 0 dentro i blocchi dei prioritari, il salvataggio
        non deve essere rifiutato dalla barriera di fase (prima del fix il
        backend rispondeva 400 "non collocabili in coda")."""
        contenitore = self.crea_contenitore(x=2000, y=2000, z=2000)
        oggetto = self.crea_oggetto("BARR-MAN", 500, 400, 300)
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Piano barriera manuale",
            contenitore=contenitore,
        )
        riga_prio1 = OggettoDaCaricare.objects.create(
            piano_di_carico=piano, oggetto=oggetto, quantita=1,
            colore="#FF0000", priorita=1,
        )
        riga_prio0 = OggettoDaCaricare.objects.create(
            piano_di_carico=piano, oggetto=oggetto, quantita=1,
            colore="#00FF00", priorita=0,
        )

        # L'utente ha spostato a mano il prio0 a X=0, DENTRO il blocco prio1
        # (X=0-500). Nel flusso manuale la scelta va rispettata, non
        # rifiutata con un 400.
        response = self._salva_posizioni_payload(piano, [
            {
                "oggetto_id": oggetto.id,
                "codice": oggetto.codice,
                "posizione_cm": {"x": 0, "y": 0, "z": 0},
                "dimensioni_cm": {"x": 50, "y": 40, "z": 30},
                "colore": "#aaaaaa",
                "rotazione": "XYZ",
                "riga_id": riga_prio1.id,
            },
            {
                "oggetto_id": oggetto.id,
                "codice": oggetto.codice,
                "posizione_cm": {"x": 0, "y": 0, "z": 0},  # viola la barriera
                "dimensioni_cm": {"x": 50, "y": 40, "z": 30},
                "colore": "#bbbbbb",
                "rotazione": "XYZ",
                "riga_id": riga_prio0.id,
            },
        ], "manuale")

        # Il salvataggio manuale deve riuscire e conservare le coordinate
        # inviate dall'utente.
        self.assertEqual(response.status_code, 200)
        op_prio0 = piano.oggetti_posizionati.get(riga_origine=riga_prio0)
        self.assertEqual(op_prio0.coordinata_x_mm, 0,
            "La modifica manuale non è stata salvata!")

    def test_payload_con_barriera_gia_rispettata_non_viene_modificato(self):
        """Se il payload rispetta già la barriera, le coordinate restano
        quelle inviate (nessun riordinamento superfluo)."""
        contenitore = self.crea_contenitore(x=2000, y=2000, z=2000)
        oggetto = self.crea_oggetto("BARR-B", 500, 400, 300)
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Piano barriera ok",
            contenitore=contenitore,
        )
        riga_prio1 = OggettoDaCaricare.objects.create(
            piano_di_carico=piano, oggetto=oggetto, quantita=1,
            colore="#FF0000", priorita=1,
        )
        riga_prio0 = OggettoDaCaricare.objects.create(
            piano_di_carico=piano, oggetto=oggetto, quantita=1,
            colore="#00FF00", priorita=0,
        )

        response = self._salva_posizioni_payload(piano, [
            {
                "oggetto_id": oggetto.id,
                "codice": oggetto.codice,
                "posizione_cm": {"x": 0, "y": 0, "z": 0},
                "dimensioni_cm": {"x": 50, "y": 40, "z": 30},
                "colore": "#aaaaaa",
                "rotazione": "XYZ",
                "riga_id": riga_prio1.id,
            },
            {
                "oggetto_id": oggetto.id,
                "codice": oggetto.codice,
                "posizione_cm": {"x": 60, "y": 0, "z": 0},  # già in coda
                "dimensioni_cm": {"x": 50, "y": 40, "z": 30},
                "colore": "#bbbbbb",
                "rotazione": "XYZ",
                "riga_id": riga_prio0.id,
            },
        ], "sincronizzazione")

        self.assertEqual(response.status_code, 200)
        op_prio0 = piano.oggetti_posizionati.get(riga_origine=riga_prio0)
        # Coordinata invariata: 60 cm = 600 mm
        self.assertEqual(op_prio0.coordinata_x_mm, 600)


class TestImpostazioniOttimizzatoreAPI(TestCase):
    """Le preferenze dell'ottimizzatore sono persistenti e isolate per utente."""

    def setUp(self):
        self.user = User.objects.create_user(
            username="settings-user",
            password="test-password",
        )
        self.altro_user = User.objects.create_user(
            username="settings-other-user",
            password="test-password",
        )
        self.client = APIClient()

    def test_get_autenticato_restituisce_configurazione_vuota(self):
        self.client.force_authenticate(user=self.user)

        response = self.client.get("/api/impostazioni_ottimizzatore/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, {"impostazioni": {}})

    def test_put_persist_get_e_isola_per_utente(self):
        payload = {
            "strategia_ottimizzazione": {
                "ordinamento_casuale": True,
                "compattazione_aggressiva": True,
            },
            "output_ottimizzazione": {
                "modalita_rotazione": "eccentrica",
            },
            "manuale": {
                "strategia_piazzamento": "colonne",
                "massima_sporgenza_pct": 50,
            },
        }
        self.client.force_authenticate(user=self.user)

        put_response = self.client.put(
            "/api/impostazioni_ottimizzatore/",
            payload,
            format="json",
        )
        get_response = self.client.get("/api/impostazioni_ottimizzatore/")

        self.assertEqual(put_response.status_code, 200)
        self.assertEqual(get_response.status_code, 200)
        self.assertEqual(get_response.data["impostazioni"], payload)
        self.assertEqual(
            UserProfile.objects.get(user=self.user).impostazioni_ottimizzatore,
            payload,
        )

        self.client.force_authenticate(user=self.altro_user)
        altro_response = self.client.get("/api/impostazioni_ottimizzatore/")
        self.assertEqual(altro_response.status_code, 200)
        self.assertEqual(altro_response.data, {"impostazioni": {}})

    def test_put_rifiuta_sezioni_non_previste(self):
        self.client.force_authenticate(user=self.user)

        response = self.client.put(
            "/api/impostazioni_ottimizzatore/",
            {"impostazioni_sistema": {"debug": True}},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("sezioni", response.data)


class TestVincoliCombinati(TestCaseBase):
    """Test con combinazioni di vincoli per-oggetto e tra-oggetti."""

    def test_sopra_con_sovrapponibile_false(self):
        contenitore = self.crea_contenitore()
        a = self.crea_oggetto("CONF-A", 400, 400, 300, peso=10)
        b = self.crea_oggetto("CONF-B", 800, 800, 200, peso=20)
        self.set_vincoli(b, sovrapponibile=False)
        self.crea_vincolo_tra(a, b, "sopra")

        piano, risultato = self.crea_piano_e_ottimizza(contenitore, [(a, 1), (b, 1)])

        self.assertTrue(risultato.successo)
        posizionati = piano.oggetti_posizionati.all()
        codici = [p.oggetto.codice for p in posizionati]
        self.assertIn(b.codice, codici)
        self.assertIn(a.codice, codici)

    def test_vincolo_sopra_tutte_config_escluse(self):
        contenitore = self.crea_contenitore()
        a = self.crea_oggetto("CONF-EXC", 400, 400, 300, peso=10)
        b = self.crea_oggetto("CONF-BASE", 400, 400, 300, peso=10)
        self.crea_vincolo_tra(
            a, b, "sopra",
            dettagli_posizionamento={
                "configurazioni": [
                    {"valida": False, "dimsA": [400, 400, 300], "dimsB": [400, 400, 300]},
                ]
            },
        )

        piano, risultato = self.crea_piano_e_ottimizza(contenitore, [(a, 1), (b, 1)])
        self.assertTrue(risultato.successo)

        def _sopra_di(item, other):
            return (
                item.coordinata_z_mm >= other.coordinata_z_mm + other.dimensione_z_mm
                and not (item.coordinata_x_mm >= other.coordinata_x_mm + other.dimensione_x_mm or
                         other.coordinata_x_mm >= item.coordinata_x_mm + item.dimensione_x_mm)
                and not (item.coordinata_y_mm >= other.coordinata_y_mm + other.dimensione_y_mm or
                         other.coordinata_y_mm >= item.coordinata_y_mm + item.dimensione_y_mm)
            )

        a_items = [o for o in risultato.oggetti_posizionati if o.codice == a.codice]
        b_items = [o for o in risultato.oggetti_posizionati if o.codice == b.codice]
        for a_item in a_items:
            for b_item in b_items:
                self.assertFalse(
                    _sopra_di(a_item, b_item),
                    "A non deve essere sopra B quando tutte le configurazioni sono escluse",
                )


class TestPrioritaContenitoreVariabile(TestCase):
    """La priorità resta esplicita e il limite Z dipende dal contenitore."""

    def test_vincolo_non_promuove_la_base_e_prioritario_viene_prima(self):
        """I02 p1 deve essere processato prima di I01 p0 anche con
        il vincolo I02 sopra I01; il vincolo non promuove tutte le basi.
        """
        from caricamento.engine.tre_d.priority_policy import (
            priorita_effettive,
            riordina_per_fasi,
        )

        oggetti = [
            *(
                Obj(
                    f"I01-{indice}",
                    120,
                    100,
                    100,
                    oggetto_id=1,
                    priorita=0,
                )
                for indice in range(15)
            ),
            *(
                Obj(
                    f"I02-{indice}",
                    120,
                    80,
                    100,
                    oggetto_id=2,
                    priorita=1,
                )
                for indice in range(5)
            ),
        ]
        vincoli = {2: {1: None}}

        self.assertEqual(priorita_effettive(oggetti, vincoli), {1: 0, 2: 1})
        riordina_per_fasi(oggetti, vincoli, preserve_inner_order=False)
        self.assertTrue(all(obj.oggetto_id == 2 for obj in oggetti[:5]))
        self.assertTrue(all(obj.oggetto_id == 1 for obj in oggetti[5:]))

        posizionati = load_truck_v2(
            oggetti,
            vincoli_sopra=vincoli,
            container_dim=(1360, 248, 270),
        )
        self.assertEqual(len(posizionati), 20)
        self.assertEqual(sum(obj.oggetto_id == 2 for obj in posizionati), 5)
        self.assertEqual(sum(obj.oggetto_id == 1 for obj in posizionati), 15)
        i02 = [obj for obj in posizionati if obj.oggetto_id == 2]
        i01 = [obj for obj in posizionati if obj.oggetto_id == 1]
        self.assertLessEqual(
            max(obj.x for obj in i02),
            min(obj.x for obj in i01),
            "Il singolo I02 deve restare nella fase I02, prima degli I01.",
        )

    def test_priorita_zero_mai_dentro_i_blocchi_delle_priorita_uno(self):
        """La barriera di fase vieta di piazzare oggetti a priorità 0
        dentro i blocchi delle priorità 1 (anche in stacking). I lotti a
        priorità 0 devono stare fisicamente dopo la fine X dei prioritari.
        """
        from caricamento.engine.tre_d.priority_policy import (
            priorita_effettive,
        )

        # Due lotti dello stesso codice: q5 a priorità 1 (piazzati per
        # primi, da X=0) e q1 a priorità 0 (deve finire in coda, mai
        # impilato dentro il blocco prioritario).
        oggetti = [
            *(
                Obj(
                    f"CART-{indice}",
                    100,
                    100,
                    100,
                    oggetto_id=1,
                    riga_origine_id=10,
                    priorita=1,
                )
                for indice in range(5)
            ),
            Obj(
                "CART-5",
                100,
                100,
                100,
                oggetto_id=1,
                riga_origine_id=11,
                priorita=0,
            ),
        ]

        self.assertEqual(
            priorita_effettive(oggetti),
            {10: 1, 11: 0},
        )

        posizionati = load_truck_v2(
            oggetti,
            container_dim=(500, 200, 300),
        )
        self.assertEqual(len(posizionati), 6)

        prio1 = [obj for obj in posizionati if obj.riga_origine_id == 10]
        prio0 = [obj for obj in posizionati if obj.riga_origine_id == 11]
        self.assertEqual(len(prio0), 1)

        max_end_prio1 = max(obj.x + obj.width for obj in prio1)
        # La priorità 0 non può iniziare prima della fine X dei prioritari:
        # mai dentro i blocchi (né a pavimento né impilata sopra di essi).
        self.assertGreaterEqual(
            prio0[0].x,
            max_end_prio1 - 0.5,
            "Il lotto a priorità 0 è finito dentro il blocco prioritario!",
        )

    def test_altezza_del_contenitore_e_dinamica(self):
        """Lo stesso carico usa il limite Z del camion/cassa selezionato.

        Entrambi gli oggetti sono nella stessa fase di priorità: il secondo
        si impila sul primo perché la barriera di fase vieta di piazzare un
        oggetto a priorità 0 sopra un blocco a priorità 1 (non sarebbe
        scaricabile senza smontare il carico).
        """
        for altezza, secondo_deve_entrare in ((270, False), (370, True)):
            oggetti = [
                Obj("PRIORITARIO-0", 100, 100, 200, oggetto_id=1, priorita=1),
                Obj("SUCCESSIVO-0", 100, 100, 100, oggetto_id=2, priorita=1),
            ]
            posizionati = load_truck_v2(
                oggetti,
                container_dim=(100, 100, altezza),
                preserve_order=True,
            )
            self.assertEqual(len(posizionati), 2 if secondo_deve_entrare else 1)
            if secondo_deve_entrare:
                secondo = next(obj for obj in posizionati if obj.oggetto_id == 2)
                self.assertEqual(secondo.z, 200)
            else:
                self.assertEqual(oggetti[1].z, -1)


class TestStrategieEndToEnd(TestCase):
    """Matrice comune per le quattro strategie automatiche."""

    def _make_objects(self):
        return [
            *(
                Obj(
                    f"I01-{index}",
                    120,
                    100,
                    100,
                    oggetto_id=1,
                    priorita=0,
                )
                for index in range(15)
            ),
            *(
                Obj(
                    f"I02-{index}",
                    120,
                    80,
                    100,
                    oggetto_id=2,
                    priorita=1,
                )
                for index in range(5)
            ),
        ]

    def _run(self, config, container_dim=(1360, 248, 270)):
        objects = self._make_objects()
        result = strategy_for_config(config).execute(
            objects,
            {2: {1: None}},
            container_dim,
        )
        fitted = [obj for obj in result if obj.z >= 0]
        return objects, result, fitted

    def test_tutte_le_strategie_rispettano_priorita_e_integrita(self):
        configs = {
            "deterministica": ConfigurazioneOttimizzazione(),
            "monte_carlo": ConfigurazioneOttimizzazione(
                ordinamento_casuale=True,
            ),
            "backtracking": ConfigurazioneOttimizzazione(
                backtracking_avanzato=True,
            ),
            "ibrida": ConfigurazioneOttimizzazione(
                ordinamento_casuale=True,
                backtracking_avanzato=True,
            ),
        }

        for name, config in configs.items():
            with self.subTest(strategy=name):
                objects, result, fitted = self._run(config)
                i02 = [obj for obj in fitted if obj.oggetto_id == 2]
                i01 = [obj for obj in fitted if obj.oggetto_id == 1]

                requested_ids = {obj.id for obj in objects}
                fitted_ids = {obj.id for obj in fitted}
                self.assertEqual(len(result), 20)
                self.assertEqual(len(fitted), 20)
                self.assertEqual(len(fitted_ids), len(fitted))
                self.assertEqual(fitted_ids, requested_ids)
                self.assertEqual(len(i02), 5)
                self.assertEqual(len(i01), 15)
                self.assertLessEqual(
                    max(obj.x for obj in i02),
                    min(obj.x for obj in i01),
                )
                self.assertTrue(all(obj.z >= 0 for obj in fitted))

    def test_tutte_le_strategie_rispettano_limiti_e_collisioni(self):
        configs = (
            ConfigurazioneOttimizzazione(),
            ConfigurazioneOttimizzazione(ordinamento_casuale=True),
            ConfigurazioneOttimizzazione(backtracking_avanzato=True),
            ConfigurazioneOttimizzazione(
                ordinamento_casuale=True,
                backtracking_avanzato=True,
            ),
        )
        containers = {
            "camion": (1360, 248, 270),
            "cassa": (1360, 248, 370),
        }

        for container_name, container in containers.items():
            for config in configs:
                with self.subTest(
                    container=container_name,
                    strategy=type(strategy_for_config(config)).__name__,
                ):
                    objects, _, fitted = self._run(config, container)
                    requested_ids = {obj.id for obj in objects}
                    fitted_ids = {obj.id for obj in fitted}
                    i02 = [obj for obj in fitted if obj.oggetto_id == 2]
                    i01 = [obj for obj in fitted if obj.oggetto_id == 1]
                    self.assertEqual(fitted_ids, requested_ids)
                    self.assertEqual(len(fitted_ids), len(fitted))
                    self.assertTrue(i02 and i01)
                    self.assertLessEqual(
                        max(obj.x for obj in i02),
                        min(obj.x for obj in i01),
                    )
                    for obj in fitted:
                        self.assertGreaterEqual(obj.x, 0)
                        self.assertGreaterEqual(obj.y, 0)
                        self.assertGreaterEqual(obj.z, 0)
                        self.assertLessEqual(obj.x + obj.width, container[0])
                        self.assertLessEqual(obj.y + obj.depth, container[1])
                        self.assertLessEqual(obj.z + obj.height, container[2])

                    for index, first in enumerate(fitted):
                        for second in fitted[index + 1:]:
                            overlap_xy = (
                                first.x < second.x + second.width
                                and first.x + first.width > second.x
                                and first.y < second.y + second.depth
                                and first.y + first.depth > second.y
                            )
                            overlap_z = (
                                first.z < second.z + second.height
                                and first.z + first.height > second.z
                            )
                            self.assertFalse(
                                overlap_xy and overlap_z,
                                f"Collisione tra {first.id} e {second.id}",
                            )


class TestOwnershipIsolation(TestCase):
    """Gli endpoint REST non devono esporre i dati di un altro utente."""

    def setUp(self):
        self.user = User.objects.create_user(username="isolation-a", password="test-password")
        self.other = User.objects.create_user(username="isolation-b", password="test-password")
        self.client = APIClient()
        self.container = Contenitore.objects.create(
            owner=self.other,
            nome="Other container",
            lunghezza_mm=2000,
            larghezza_mm=2000,
            altezza_mm=2000,
            carico_massimo_kg=Decimal("1000"),
        )
        self.item = Oggetto.objects.create(
            owner=self.other,
            codice="OTHER-ITEM",
            lunghezza_mm=500,
            larghezza_mm=400,
            altezza_mm=300,
            peso_kg=Decimal("10"),
            quantita_disponibile=1,
        )
        self.plan = PianoDiCarico.objects.create(
            owner=self.other,
            nome="Other plan",
            contenitore=self.container,
        )

    def test_other_user_resources_are_not_listed_or_retrievable(self):
        self.client.force_authenticate(user=self.user)

        self.assertEqual(self.client.get("/api/contenitori/").status_code, 200)
        self.assertEqual(self.client.get("/api/contenitori/").data["count"], 0)
        self.assertEqual(self.client.get(f"/api/contenitori/{self.container.id}/").status_code, 404)
        self.assertEqual(self.client.get(f"/api/oggetti/{self.item.id}/").status_code, 404)
        self.assertEqual(self.client.get(f"/api/piani/{self.plan.id}/").status_code, 404)

    def test_other_user_resources_cannot_be_used_in_nested_actions(self):
        self.client.force_authenticate(user=self.user)

        response = self.client.post(
            f"/api/piani/{self.plan.id}/oggetti_da_caricare/",
            {"oggetto_id": self.item.id, "quantita": 1},
            format="json",
        )
        self.assertEqual(response.status_code, 404)


class TestApiErrorHandling(TestCase):
    """Le API restituiscono errori leggibili senza perdere i dettagli legacy."""

    def setUp(self):
        self.user = User.objects.create_user(username="errors-user", password="test-password")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_validation_error_has_standard_envelope_and_request_id(self):
        response = self.client.put(
            "/api/impostazioni_ottimizzatore/",
            {"sezione_non_valida": {}},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.data["success"])
        self.assertEqual(response.data["error"]["code"], "validation_error")
        self.assertTrue(response.data["error"]["request_id"])
        self.assertEqual(response["X-Request-ID"], response.data["error"]["request_id"])
        # Compatibilità con i client che usavano il campo DRF originale.
        self.assertIn("sezioni", response.data)

    def test_stato_con_task_q_fallito_non_resta_in_elaborazione(self):
        container = Contenitore.objects.create(
            owner=self.user,
            nome="Container task fallito",
            lunghezza_mm=2000,
            larghezza_mm=2000,
            altezza_mm=2000,
            carico_massimo_kg=Decimal("1000"),
        )
        piano = PianoDiCarico.objects.create(
            owner=self.user,
            nome="Piano task fallito",
            contenitore=container,
            stato=StatoPiano.IN_ELABORAZIONE,
            task_id="task-fallito-test",
        )
        with patch("django_q.models.Task.objects.filter") as task_filter:
            task_filter.return_value.first.return_value = Mock(success=False)
            response = self.client.get(f"/api/piani/{piano.id}/stato/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["stato"], StatoPiano.ERRORE)
        self.assertIn("worker", response.data["messaggio_errore"])


class TestDemoTrialFingerprint(TestCase):
    """Un fingerprint impedisce di assegnare trial demo aggiuntivi."""

    def setUp(self):
        self.factory = RequestFactory()
        self.ip = "203.0.113.10"
        self.browser = "browser-hash-demo"
        self.cookie = "cookie-token-demo"
        impostazioni = ImpostazioniSistema.get()
        impostazioni.demo_attiva = True
        impostazioni.soglia_controlli_demo = 1
        impostazioni.save()

    def _request(self):
        request = self.factory.post("/")
        request.META["HTTP_X_REAL_IP"] = self.ip
        request.COOKIES["cb_fp"] = self.browser
        request.COOKIES["cb_demo"] = self.cookie
        return request

    def _active_user(self, username):
        user = User.objects.create_user(username=username, password="password-demo")
        profile = user.profile
        profile.trial_start = timezone.now()
        profile.trial_end = timezone.now() + timedelta(days=7)
        profile.save(update_fields=["trial_start", "trial_end"])
        return user

    def test_used_fingerprint_blocks_second_trial_while_first_is_active(self):
        user = self._active_user("first-demo")
        _save_demo_fingerprints(self._request(), user)

        self.assertTrue(_check_demo_abuse(self._request()))

    def test_admin_can_disable_anti_abuse_checks_for_testing(self):
        user = self._active_user("first-demo-disabled-checks")
        _save_demo_fingerprints(self._request(), user)
        impostazioni = ImpostazioniSistema.get()
        impostazioni.controlli_demo_attivi = False
        impostazioni.save()

        self.assertFalse(_check_demo_abuse(self._request()))

    def test_staff_fingerprint_does_not_block_demo_trials(self):
        user = User.objects.create_user(
            username="staff-fingerprint", password="password-demo", is_staff=True
        )
        _save_demo_fingerprints(self._request(), user)

        self.assertFalse(_check_demo_abuse(self._request()))

    def test_existing_account_can_login_from_a_matching_device(self):
        user = self._active_user("portable-demo")
        _save_demo_fingerprints(self._request(), user)

        self.client.cookies["cb_fp"] = self.browser
        self.client.cookies["cb_demo"] = self.cookie
        response = self.client.post(
            "/",
            {"username": user.username, "password": "password-demo"},
            REMOTE_ADDR="127.0.0.1",
            HTTP_X_REAL_IP=self.ip,
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/workspace/")

    def test_new_account_requires_matching_password_confirmation(self):
        # L'email è obbligatoria per i nuovi account (verifica email).
        response = self.client.post(
            "/",
            {
                "username": "new-demo-confirm",
                "password": "password-demo",
                "email": "confirm@test.it",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "inserisci due volte la stessa password")
        self.assertFalse(User.objects.filter(username="new-demo-confirm").exists())

        response = self.client.post(
            "/",
            {
                "username": "new-demo-confirm",
                "password": "password-demo",
                "password_confirm": "password-demo",
                "email": "confirm@test.it",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Email di verifica inviata")
        self.assertTrue(User.objects.filter(username="new-demo-confirm").exists())

    def test_existing_account_can_login_when_new_demo_signups_are_disabled(self):
        user = self._active_user("existing-demo-disabled-signups")
        impostazioni = ImpostazioniSistema.get()
        impostazioni.demo_attiva = False
        impostazioni.save()

        response = self.client.post(
            "/",
            {"username": user.username, "password": "password-demo"},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/workspace/")

    def test_expired_trial_api_returns_json_403(self):
        user = User.objects.create_user(username="expired-demo", password="password-demo")
        profile = user.profile
        profile.trial_start = timezone.now() - timedelta(days=15)
        profile.trial_end = timezone.now() - timedelta(days=1)
        profile.save(update_fields=["trial_start", "trial_end"])
        self.client.force_login(user)

        # Le letture sono permesse (il frontend carica lo stato pagamenti),
        # le scritture vengono bloccate con 403 trial_expired.
        lettura = self.client.get("/api/contenitori/")
        self.assertEqual(lettura.status_code, 200)

        scrittura = self.client.post("/api/contenitori/")
        self.assertEqual(scrittura.status_code, 403)
        self.assertEqual(scrittura.json()["error"]["code"], "trial_expired")

    def test_disactivated_user_cannot_reuse_an_existing_session(self):
        user = self._active_user("disabled-demo")
        self.client.force_login(user)
        user.is_active = False
        user.save(update_fields=["is_active"])

        response = self.client.get("/workspace/")

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/?next=/workspace/")

    def test_new_google_identity_is_blocked_by_used_fingerprint(self):
        from types import SimpleNamespace

        from allauth.core.exceptions import ImmediateHttpResponse
        from caricamento.adapter import DynamicGoogleAdapter

        user = self._active_user("google-demo")
        _save_demo_fingerprints(self._request(), user)
        request = self._request()

        with self.assertRaises(ImmediateHttpResponse) as raised:
            with patch("caricamento.views._check_demo_abuse", return_value=True):
                DynamicGoogleAdapter().pre_social_login(
                    request,
                    SimpleNamespace(is_existing=False),
                )

        self.assertEqual(raised.exception.response.status_code, 302)
        self.assertEqual(raised.exception.response["Location"], "/?trial=used")

    def test_disabled_local_account_shows_disabled_message_instead_of_recreating(self):
        user = self._active_user("disabled-login")
        user.is_active = False
        user.save(update_fields=["is_active"])

        response = self.client.post(
            "/",
            {"username": user.username, "password": "password-demo"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Questo account è stato disattivato")
        self.assertEqual(User.objects.filter(username=user.username).count(), 1)

    def test_google_new_identity_is_blocked_when_demo_is_disabled(self):
        from types import SimpleNamespace

        from allauth.core.exceptions import ImmediateHttpResponse
        from caricamento.adapter import DynamicGoogleAdapter

        impostazioni = ImpostazioniSistema.get()
        impostazioni.demo_attiva = False
        impostazioni.save()

        with self.assertRaises(ImmediateHttpResponse) as raised:
            DynamicGoogleAdapter().pre_social_login(
                self._request(),
                SimpleNamespace(is_existing=False),
            )

        self.assertEqual(raised.exception.response.status_code, 302)
        self.assertEqual(raised.exception.response["Location"], "/?demo=disabled")

    def test_demo_threshold_cannot_exceed_available_signals(self):
        impostazioni = ImpostazioniSistema.get()
        impostazioni.soglia_controlli_demo = 4

        with self.assertRaises(ValidationError):
            impostazioni.full_clean()


class TestSecurityHardening(TestCase):
    """Test delle correzioni di sicurezza (IP reale, magic bytes PNG, XSS)."""

    def setUp(self):
        self.factory = RequestFactory()

    def test_client_ip_prefers_x_real_ip_over_forwarded(self):
        request = self.factory.get("/")
        request.META["REMOTE_ADDR"] = "10.0.0.5"
        request.META["HTTP_X_FORWARDED_FOR"] = "1.2.3.4, 10.0.0.5"
        request.META["HTTP_X_REAL_IP"] = "203.0.113.9"

        self.assertEqual(get_client_ip(request), "203.0.113.9")

    def test_client_ip_ignores_spoofed_forwarded_and_uses_remote_addr(self):
        request = self.factory.get("/")
        request.META["REMOTE_ADDR"] = "10.0.0.5"
        request.META["HTTP_X_FORWARDED_FOR"] = "1.2.3.4"

        self.assertEqual(get_client_ip(request), "10.0.0.5")

    def test_icon_upload_rejects_non_png_content(self):
        staff = User.objects.create_user(
            username="staff-icons", password="password-demo", is_staff=True
        )
        self.client.force_login(staff)

        fake = SimpleUploadedFile(
            "finto.png", b"MZ not a real png", content_type="image/png"
        )
        response = self.client.post("/api/icone-upload/", {"file": fake})

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["error"]["code"], "invalid_file_type")

    def test_icon_upload_accepts_real_png_signature(self):
        staff = User.objects.create_user(
            username="staff-icons-ok", password="password-demo", is_staff=True
        )
        self.client.force_login(staff)

        png_buffer = io.BytesIO()
        Image.new("RGBA", (1, 1), (0, 0, 0, 0)).save(png_buffer, format="PNG")
        png = SimpleUploadedFile(
            "ok.png",
            png_buffer.getvalue(),
            content_type="image/png",
        )
        with tempfile.TemporaryDirectory() as tmp, patch(
            "caricamento.views.ICON_UPLOAD_DIR", tmp
        ):
            response = self.client.post("/api/icone-upload/", {"file": png})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success"], True)
        self.assertEqual(response.json()["filename"], "ok.png")

    def test_workspace_escapes_dettagli_posizionamento(self):
        user = User.objects.create_user(username="xss-user", password="password-demo")
        self.client.force_login(user)

        a = Oggetto.objects.create(
            owner=user, codice="XSS-A", lunghezza_mm=10, larghezza_mm=10,
            altezza_mm=10, peso_kg=Decimal("1.00"),
        )
        b = Oggetto.objects.create(
            owner=user, codice="XSS-B", lunghezza_mm=10, larghezza_mm=10,
            altezza_mm=10, peso_kg=Decimal("1.00"),
        )
        payload = "</script><script>alert('XSS-SEC')</script>"
        VincoloTraOggetti.objects.create(
            oggetto_a=a,
            oggetto_b=b,
            tipo_relazione="sopra",
            dettagli_posizionamento={"note": payload},
        )

        response = self.client.get("/workspace/")

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertNotIn(payload, content)
        # escapejs trasforma '<' e '>' in \u003C / \u003E: niente breakout dello script.
        self.assertIn("\\u003C/script\\u003E", content)

    def test_client_ip_rejects_spoofed_x_real_ip_from_public_peer(self):
        request = self.factory.get("/")
        request.META["REMOTE_ADDR"] = "203.0.113.50"  # IP pubblico: peer non fidato
        request.META["HTTP_X_REAL_IP"] = "10.0.0.7"

        self.assertEqual(get_client_ip(request), "203.0.113.50")

    def test_bulk_vincoli_ignores_structural_fields(self):
        user = User.objects.create_user(username="bulk-user", password="password-demo")
        other = User.objects.create_user(username="bulk-other", password="password-demo")

        o1 = Oggetto.objects.create(
            owner=user, codice="BULK-1", lunghezza_mm=10, larghezza_mm=10,
            altezza_mm=10, peso_kg=Decimal("1.00"),
        )
        o2 = Oggetto.objects.create(
            owner=other, codice="BULK-2", lunghezza_mm=10, larghezza_mm=10,
            altezza_mm=10, peso_kg=Decimal("1.00"),
        )

        client = APIClient()
        client.force_authenticate(user=user)
        response = client.post(
            "/api/oggetti/bulk_vincoli/",
            {"ids": [o1.id], "vincoli": {"oggetto_id": o2.id, "fragile": True}},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        vincolo = VincoloOggetto.objects.get(oggetto=o1)
        # Il campo strutturale oggetto_id è ignorato: nessuna riassegnazione cross-tenant.
        self.assertEqual(vincolo.oggetto_id, o1.id)
        # Il campo della whitelist viene comunque applicato.
        self.assertTrue(vincolo.fragile)


# =============================================================================
# TEST: SPATIAL INDEX (uniform grid)
# =============================================================================

class TestSpatialGrid(TestCase):
    """Regressioni per l'indice spaziale: over-approximation e ordine."""

    def _obj(self, i, x, y, z, w, d, h):
        o = Obj(f"G-{i}", w, d, h, oggetto_id=1)
        o.x, o.y, o.z = x, y, z
        return o

    def test_query_volume_non_perde_candidati_e_preserva_ordine(self):
        grid = SpatialGrid(cell_size=50.0)
        specs = [
            (0, 0, 0, 100, 80, 50),
            (100, 0, 0, 60, 80, 100),
            (0, 80, 50, 100, 80, 50),
            (200, 100, 0, 50, 50, 50),
            (250, 0, 100, 40, 40, 40),
        ]
        objs = [self._obj(i, *s) for i, s in enumerate(specs)]
        for o in objs:
            grid.add(o)

        query = (25, 0, 0, 150, 120, 100)
        x0, y0, z0, x1, y1, z1 = query
        ref = [
            o for o in objs
            if o.x < x1 and o.x + o.width > x0
            and o.y < y1 and o.y + o.depth > y0
            and o.z < z1 and o.z + o.height > z0
        ]
        cand = grid.query_volume(*query)
        # Mai candidati in meno.
        self.assertTrue(set(ref) <= set(cand))
        # I candidati restano nell'ordine di inserimento (== ordine di placed).
        self.assertEqual(cand, [o for o in objs if o in cand])

    def test_query_top_e_bottom_rispettano_l_epsilon(self):
        grid = SpatialGrid(cell_size=50.0)
        base = self._obj(0, 0, 0, 0, 100, 80, 100)
        sopra = self._obj(1, 0, 0, 100, 60, 60, 50)
        altro = self._obj(2, 0, 0, 150, 60, 60, 50)
        for o in (base, sopra, altro):
            grid.add(o)

        # top == 100 -> base (e solo base)
        self.assertEqual(grid.query_top(100), [base])
        # bottom == 100 -> sopra
        self.assertEqual(grid.query_bottom(100), [sopra])
        # Nessun oggetto con top == 75 (entro epsilon)
        self.assertEqual(grid.query_top(75), [])

    def test_add_remove_simmetria(self):
        grid = SpatialGrid(cell_size=50.0)
        a = self._obj(0, 0, 0, 0, 100, 80, 50)
        b = self._obj(1, 100, 0, 0, 60, 80, 100)
        grid.add(a)
        grid.add(b)
        self.assertEqual(grid.size, 2)
        grid.remove(a)
        self.assertEqual(grid.size, 1)
        cand = grid.query_volume(0, 0, 0, 200, 200, 200)
        self.assertEqual(cand, [b])


class TestParitySpatialIndex(TestCase):
    """Parity test: con la griglia attiva il risultato e' bit-identico.

    Usa i codici reali I01-I05 (non il caso 700): il piano reale n. 348 con
    9/13/18/17 istanze e il caso deferimento 20/8/13/19.
    """

    DATI = [
        ("CART-I03", 160, 70, 100, 9, 3),
        ("CART-I01", 120, 100, 100, 13, 1),
        ("CART-I02", 120, 80, 100, 18, 2),
        ("cart-i05", 60, 80, 100, 17, 5),
    ]

    DATI_DEFERIMENTO = [
        ("CART-I03", 70, 160, 100, 20, 3),
        ("CART-I01", 100, 120, 100, 8, 1),
        ("CART-I02", 80, 120, 100, 13, 2),
        ("cart-i05", 60, 80, 100, 19, 5),
    ]

    CONTAINER = (1360, 248, 270)

    def _build(self, dati):
        objects = []
        for codice, width, depth, height, quantita, oggetto_id in dati:
            for indice in range(quantita):
                objects.append(
                    Obj(
                        f"{codice}-{indice}",
                        width,
                        depth,
                        height,
                        oggetto_id=oggetto_id,
                        orientation_allowed=True,
                        rotazione_su_x=False,
                        rotazione_su_y=False,
                        rotazione_su_z=True,
                    )
                )
        return objects

    @staticmethod
    def _snapshot(objects):
        return [
            (
                o.id,
                o.x, o.y, o.z,
                o.width, o.depth, o.height,
                getattr(o, "support_ratio", 1.0),
                float(getattr(o, "_peso_kg", 0.0)),
                float(getattr(o, "_peso_sopra_kg", 0.0)),
            )
            for o in objects
        ]

    def _run_deterministic(self, dati, preserve_order=False):
        return load_truck_v2(
            self._build(dati),
            vincoli_sopra={},
            container_dim=self.CONTAINER,
            preserve_order=preserve_order,
        )

    def _assert_parity(self, run_off, run_on):
        self.assertEqual(self._snapshot(run_off), self._snapshot(run_on))

    def test_load_truck_v2_parita(self):
        for dati in (self.DATI, self.DATI_DEFERIMENTO):
            for preserve in (False, True):
                with self.subTest(dati=dati[0][0], preserve=preserve):
                    with patch(
                        "caricamento.engine.tre_d.packer_3d_v2.SPATIAL_GRID_ENABLED",
                        False,
                    ):
                        off = self._run_deterministic(dati, preserve)
                    with _grid_on():
                        on = self._run_deterministic(dati, preserve)
                    self._assert_parity(off, on)

    def test_deterministic_strategy_parita(self):
        for dati in (self.DATI, self.DATI_DEFERIMENTO):
            with self.subTest(dati=dati[0][0]):
                with patch(
                    "caricamento.engine.tre_d.packer_3d_v2.SPATIAL_GRID_ENABLED",
                    False,
                ):
                    off = DeterministicStrategy().execute(
                        self._build(dati), {}, self.CONTAINER
                    )
                with _grid_on():
                    on = DeterministicStrategy().execute(
                        self._build(dati), {}, self.CONTAINER
                    )
                self._assert_parity(off, on)

    def test_random_packer_parita(self):
        stato = random.getstate()
        try:
            with patch(
                "caricamento.engine.tre_d.packer_3d_v2.SPATIAL_GRID_ENABLED",
                False,
            ):
                random.seed(0)
                off = run_packing_random(
                    self._build(self.DATI),
                    vincoli_sopra={},
                    num_restarts=5,
                    container_dim=self.CONTAINER,
                )
            with _grid_on():
                random.seed(0)
                on = run_packing_random(
                    self._build(self.DATI),
                    vincoli_sopra={},
                    num_restarts=5,
                    container_dim=self.CONTAINER,
                )
        finally:
            random.setstate(stato)
        self._assert_parity(off, on)


class _grid_on:
    """Attiva la griglia con soglia 0 per esercitarla fin dal primo piazzamento."""

    def __enter__(self):
        self._patches = [
            patch(
                "caricamento.engine.tre_d.packer_3d_v2.SPATIAL_GRID_ENABLED",
                True,
            ),
            patch(
                "caricamento.engine.tre_d.packer_3d_v2.SPATIAL_GRID_THRESHOLD",
                0,
            ),
            patch(
                "caricamento.engine.tre_d.placement_rules.SPATIAL_GRID_THRESHOLD",
                0,
            ),
        ]
        for p in self._patches:
            p.start()
        return self

    def __exit__(self, exc_type, exc, tb):
        for p in reversed(self._patches):
            p.stop()
        return False


class TestPagineLegali(TestCase):
    """Le pagine legali pubbliche rendono con i dati del titolare."""

    URL_PAGINE = [
        "/privacy/",
        "/cookie-policy/",
        "/termini/",
        "/rimborsi/",
    ]

    def test_pagine_legali_pubbliche_renderizzano(self):
        for url in self.URL_PAGINE:
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200, url)
            self.assertContains(response, "Carico 3D")

    def test_privacy_mostra_email_titolare(self):
        imp = ImpostazioniSistema.get()
        imp.privacy_titolare = "Titolare di Test"
        imp.privacy_email = "privacy@test.it"
        imp.privacy_sito_url = "http://127.0.0.1:8000"
        imp.save()

        response = self.client.get("/privacy/")

        self.assertContains(response, "Titolare di Test")
        self.assertContains(response, "privacy@test.it")

    def test_slug_sconosciuto_redirige_alla_home(self):
        response = self.client.get("/privacy/inesistente/")
        self.assertEqual(response.status_code, 404)


class TestApiPrivacySettings(TestCase):
    """API dati Privacy/Titolare: GET libero, POST solo admin."""

    def setUp(self):
        self.user = User.objects.create_user(
            username="privacy-user",
            password="test-password",
        )
        self.admin = User.objects.create_user(
            username="privacy-admin",
            password="test-password",
            is_staff=True,
        )
        self.client = APIClient()

    def test_get_restituisce_default(self):
        response = self.client.get("/api/privacy-settings/")

        self.assertEqual(response.status_code, 200)
        privacy = response.json()["privacy"]
        self.assertEqual(privacy["titolare"], "Carico 3D")
        self.assertEqual(privacy["sito_url"], "http://127.0.0.1:8000")

    def test_post_solo_admin(self):
        payload = {
            "privacy": {
                "titolare": "Webapp SRL",
                "email": "info@webapp.it",
                "sede": "Via Roma 1, Milano",
                "piva": "12345678901",
                "sito_url": "https://webapp.it",
            }
        }

        self.assertTrue(self.client.login(username=self.user.username, password="test-password"))
        denied = self.client.post(
            "/api/privacy-settings/", payload, format="json"
        )
        self.assertEqual(denied.status_code, 403)
        self.client.logout()

        self.assertTrue(self.client.login(username=self.admin.username, password="test-password"))
        ok = self.client.post(
            "/api/privacy-settings/", payload, format="json"
        )
        self.assertEqual(ok.status_code, 200)

        imp = ImpostazioniSistema.get()
        self.assertEqual(imp.privacy_titolare, "Webapp SRL")
        self.assertEqual(imp.privacy_email, "info@webapp.it")
        self.assertEqual(imp.privacy_sede, "Via Roma 1, Milano")
        self.assertEqual(imp.privacy_piva, "12345678901")
        self.assertEqual(imp.privacy_sito_url, "https://webapp.it")

    def test_post_valida_campi(self):
        self.assertTrue(self.client.login(username=self.admin.username, password="test-password"))
        response = self.client.post(
            "/api/privacy-settings/",
            {"privacy": {"titolare": "", "email": "non-email", "sito_url": "ftp://x"}},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
